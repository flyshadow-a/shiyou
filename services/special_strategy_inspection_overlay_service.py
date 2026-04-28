from __future__ import annotations

import copy
import os
from collections import defaultdict
from functools import lru_cache
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from core.app_paths import external_path
from services.file_db_adapter import shared_storage_dir
from services.special_strategy_state_db import (
    load_latest_strategy_result_snapshot,
    load_latest_strategy_run,
    load_strategy_result_snapshot_by_run,
    load_strategy_run_by_id,
)


DISPLAY_TO_CONTEXT_TIME = {
    "": "当前",
    "当前": "当前",
    "第0年": "当前",
    "0年": "当前",
    "+0年": "当前",
    "+5年": "第5年",
    "5年": "第5年",
    "第5年": "第5年",
    "+10年": "第10年",
    "10年": "第10年",
    "第10年": "第10年",
    "+15年": "第15年",
    "15年": "第15年",
    "第15年": "第15年",
    "+20年": "第20年",
    "20年": "第20年",
    "第20年": "第20年",
    "+25年": "第25年",
    "25年": "第25年",
    "第25年": "第25年",
}

INSPECTION_LEVEL_ORDER = {
    "": 0,
    "-": 0,
    "I": 1,
    "II": 2,
    "III": 3,
    "IV": 4,
}


def _txt(value: Any) -> str:
    return "" if value is None else str(value).strip()


def _norm_time_value(value: Any) -> str:
    text = _txt(value).replace(" ", "")
    return DISPLAY_TO_CONTEXT_TIME.get(text, text)


def _norm_time_label(display_year: str | None) -> str:
    return _norm_time_value(display_year) or "当前"


def _norm_level(level: Any) -> str:
    text = _txt(level).upper().replace("Ⅱ", "II").replace("Ⅲ", "III").replace("Ⅳ", "IV")
    return text if text in ("II", "III", "IV") else ""


def _pick_worse_level(a: str, b: str) -> str:
    aa = _norm_level(a)
    bb = _norm_level(b)
    return bb if INSPECTION_LEVEL_ORDER.get(bb, 0) > INSPECTION_LEVEL_ORDER.get(aa, 0) else aa


def _member_key(joint_a: str, joint_b: str) -> str:
    a = _txt(joint_a)
    b = _txt(joint_b)
    if not a and not b:
        return ""
    pair = sorted([a, b])
    return f"{pair[0]}|{pair[1]}"


def _normalize_facility_code(facility_code: str) -> str:
    return (facility_code or "").strip().upper() or "WC19-1D"


def _time_matches(row_time: Any, target_time: str) -> bool:
    row_norm = _norm_time_value(row_time)
    target_norm = _norm_time_value(target_time)
    return row_norm == target_norm


def _load_snapshot_payload(facility_code: str, run_id: int | None = None) -> dict[str, Any] | None:
    snapshot = load_strategy_result_snapshot_by_run(int(run_id)) if run_id else load_latest_strategy_result_snapshot(facility_code)
    if not snapshot:
        return None
    payload = snapshot.get("result_json")
    return payload if isinstance(payload, dict) else None


def _extract_workbook_path_from_run_payload(payload: dict[str, Any] | None) -> str:
    if not isinstance(payload, dict):
        return ""

    for key in ("intermediate_workbook", "pipeline_workbook", "workbook_path"):
        text = _txt(payload.get(key))
        if text and os.path.exists(text):
            return os.path.normpath(text)

    state = payload.get("state")
    if isinstance(state, dict):
        for key in ("intermediate_workbook", "pipeline_workbook", "workbook_path"):
            text = _txt(state.get(key))
            if text and os.path.exists(text):
                return os.path.normpath(text)

    result_json = payload.get("result_json")
    if isinstance(result_json, dict):
        state = result_json.get("state")
        if isinstance(state, dict):
            text = _txt(state.get("intermediate_workbook"))
            if text and os.path.exists(text):
                return os.path.normpath(text)

    return ""


def _candidate_pipeline_paths(facility_code: str, run_id: int | None = None) -> list[str]:
    code = _normalize_facility_code(facility_code)
    candidates: list[str] = []

    try:
        run_payload = load_strategy_run_by_id(int(run_id)) if run_id else load_latest_strategy_run(code)
        path = _extract_workbook_path_from_run_payload(run_payload)
        if path:
            candidates.append(path)
    except Exception as exc:
        print("[InspectionOverlay] load run workbook path failed:", exc)

    try:
        snapshot_payload = _load_snapshot_payload(code, run_id=run_id)
        path = _extract_workbook_path_from_run_payload(snapshot_payload)
        if path:
            candidates.append(path)
    except Exception as exc:
        print("[InspectionOverlay] load snapshot workbook path failed:", exc)

    shared_root = shared_storage_dir("special_strategy_runtime")
    if shared_root:
        candidates.append(os.path.join(shared_root, code, "special_strategy.pipeline.xlsx"))

    candidates.append(external_path("special_strategy_runtime", code, "special_strategy.pipeline.xlsx"))

    out: list[str] = []
    seen = set()
    for p in candidates:
        if not p:
            continue
        np = os.path.normpath(str(p))
        if np in seen:
            continue
        seen.add(np)
        if os.path.exists(np):
            out.append(np)

    return out


def _norm_header_key(text: Any) -> str:
    return _txt(text).replace(" ", "").replace("\n", "").replace("_", "").lower()


def _get_by_keys(row: dict[str, Any], *keys: str) -> Any:
    normalized = {_norm_header_key(k): v for k, v in row.items()}
    for key in keys:
        nk = _norm_header_key(key)
        if nk in normalized:
            return normalized[nk]
    return ""


def _read_sheet_rows(ws, header_row: int, data_start_row: int) -> list[dict[str, Any]]:
    header_values = next(ws.iter_rows(min_row=header_row, max_row=header_row, values_only=True))
    headers = ["" if v is None else str(v).strip() for v in header_values]

    rows: list[dict[str, Any]] = []
    for values in ws.iter_rows(min_row=data_start_row, values_only=True):
        if not any(v not in (None, "", " ") for v in values):
            continue

        row: dict[str, Any] = {}
        for idx, header in enumerate(headers):
            if not header:
                continue
            row[header] = values[idx] if idx < len(values) else None
        rows.append(row)

    return rows


@lru_cache(maxsize=32)
def _read_pipeline_overlay_cached(workbook_path: str, mtime: float, context_year: str) -> dict[str, Any]:
    member_level_by_key: dict[str, str] = {}
    node_level_by_joint: dict[str, str] = {}
    node_level_by_joint_brace: dict[str, str] = {}
    member_items_by_key = defaultdict(list)
    node_items_by_joint = defaultdict(list)

    wb = load_workbook(workbook_path, data_only=True, read_only=True)

    try:
        if "构件检验策略" in wb.sheetnames:
            ws_member = wb["构件检验策略"]
            member_rows = _read_sheet_rows(ws_member, header_row=1, data_start_row=2)

            for row in member_rows:
                if not _time_matches(_get_by_keys(row, "检验时间节点", "time_node"), context_year):
                    continue

                inspect_level = _norm_level(_get_by_keys(row, "检验等级", "inspect_level"))
                if not inspect_level:
                    continue

                joint_a = _txt(_get_by_keys(row, "Joint A", "JointA", "joint_a"))
                joint_b = _txt(_get_by_keys(row, "Joint B", "JointB", "joint_b"))
                key = _member_key(joint_a, joint_b)
                if not key:
                    continue

                member_level_by_key[key] = _pick_worse_level(member_level_by_key.get(key, ""), inspect_level)
                member_items_by_key[key].append({
                    "joint_a": joint_a,
                    "joint_b": joint_b,
                    "member_type": _txt(_get_by_keys(row, "MemberType", "member_type")),
                    "risk_level": _txt(_get_by_keys(row, "构件风险等级", "risk_level")),
                    "inspect_level": inspect_level,
                    "time_node": _txt(_get_by_keys(row, "检验时间节点", "time_node")),
                })

        if "节点检验策略" in wb.sheetnames:
            ws_node = wb["节点检验策略"]
            node_rows = _read_sheet_rows(ws_node, header_row=2, data_start_row=3)

            seen_node_items = set()
            for row in node_rows:
                if not _time_matches(_get_by_keys(row, "检验时间节点", "time_node"), context_year):
                    continue

                inspect_level = _norm_level(_get_by_keys(row, "检验等级", "inspect_level"))
                if not inspect_level:
                    continue

                # 注意：当前 Excel 表头就是 JoitID，不是 JointID
                joint_id = _txt(_get_by_keys(row, "JoitID", "JointID", "Joint ID", "joint_id"))
                brace = _txt(_get_by_keys(row, "Brace", "brace"))

                if joint_id:
                    node_level_by_joint[joint_id] = _pick_worse_level(
                        node_level_by_joint.get(joint_id, ""),
                        inspect_level,
                    )

                if joint_id or brace:
                    jb_key = f"{joint_id}|{brace}"
                    node_level_by_joint_brace[jb_key] = _pick_worse_level(
                        node_level_by_joint_brace.get(jb_key, ""),
                        inspect_level,
                    )

                dedup_key = (joint_id, brace, inspect_level)
                if dedup_key in seen_node_items:
                    continue
                seen_node_items.add(dedup_key)

                if joint_id:
                    node_items_by_joint[joint_id].append({
                        "joint_id": joint_id,
                        "brace": brace,
                        "joint_type": _txt(_get_by_keys(row, "JointType", "joint_type")),
                        "risk_level": _txt(_get_by_keys(row, "节点风险等级", "risk_level")),
                        "inspect_level": inspect_level,
                        "time_node": _txt(_get_by_keys(row, "检验时间节点", "time_node")),
                    })
    finally:
        try:
            wb.close()
        except Exception:
            pass

    return {
        "source": "pipeline_xlsx",
        "source_workbook": workbook_path,
        "context_year": context_year,
        "member_level_by_key": dict(member_level_by_key),
        "node_level_by_joint": dict(node_level_by_joint),
        "node_level_by_joint_brace": dict(node_level_by_joint_brace),
        "member_items_by_key": dict(member_items_by_key),
        "node_items_by_joint": dict(node_items_by_joint),
    }


def _read_pipeline_overlay(workbook_path: str, context_year: str) -> dict[str, Any]:
    mtime = os.path.getmtime(workbook_path)
    return copy.deepcopy(_read_pipeline_overlay_cached(workbook_path, mtime, context_year))


def _iter_member_rows_from_context(context: dict[str, Any]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []

    for key in (
        "member_inspection_strategy_rows",
        "member_inspection_rows_current",
        "member_inspection_rows",
    ):
        for row in context.get(key, []) or []:
            if isinstance(row, dict):
                rows.append(row)

    for row in context.get("member_risk_rows", []) or []:
        if isinstance(row, dict) and _txt(row.get("inspect_level")):
            rows.append(row)

    return rows


def _iter_node_rows_from_context(context: dict[str, Any]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []

    for key in (
        "node_inspection_strategy_rows",
        "node_inspection_rows_current",
        "node_inspection_rows_future",
        "node_inspection_rows",
    ):
        for row in context.get(key, []) or []:
            if isinstance(row, dict):
                rows.append(row)

    for row in context.get("node_risk_rows_current", []) or []:
        if isinstance(row, dict) and _txt(row.get("inspect_level")):
            rows.append(row)

    return rows


def _load_context_overlay_fallback(
    facility_code: str,
    run_id: int | None,
    display_year: str,
) -> dict[str, Any]:
    payload = _load_snapshot_payload(facility_code, run_id=run_id)
    if not payload:
        return _empty_overlay(facility_code, run_id, display_year, _norm_time_label(display_year))

    context = payload.get("context") or {}
    context_year = _norm_time_label(display_year)

    member_level_by_key = {}
    node_level_by_joint = {}
    node_level_by_joint_brace = {}
    member_items_by_key = defaultdict(list)
    node_items_by_joint = defaultdict(list)

    for row in _iter_member_rows_from_context(context):
        if not _time_matches(row.get("time_node"), context_year):
            continue
        inspect_level = _norm_level(row.get("inspect_level"))
        if not inspect_level:
            continue

        joint_a = _txt(row.get("joint_a"))
        joint_b = _txt(row.get("joint_b"))
        key = _member_key(joint_a, joint_b)
        if not key:
            continue

        member_level_by_key[key] = _pick_worse_level(member_level_by_key.get(key, ""), inspect_level)
        member_items_by_key[key].append({
            "joint_a": joint_a,
            "joint_b": joint_b,
            "member_type": _txt(row.get("member_type")),
            "inspect_level": inspect_level,
            "time_node": _txt(row.get("time_node")),
        })

    seen_node_items = set()
    for row in _iter_node_rows_from_context(context):
        if not _time_matches(row.get("time_node"), context_year):
            continue
        inspect_level = _norm_level(row.get("inspect_level"))
        if not inspect_level:
            continue

        joint_id = _txt(row.get("joint_id"))
        brace = _txt(row.get("brace"))

        if joint_id:
            node_level_by_joint[joint_id] = _pick_worse_level(node_level_by_joint.get(joint_id, ""), inspect_level)

        if joint_id or brace:
            jb_key = f"{joint_id}|{brace}"
            node_level_by_joint_brace[jb_key] = _pick_worse_level(
                node_level_by_joint_brace.get(jb_key, ""),
                inspect_level,
            )

        dedup_key = (joint_id, brace, inspect_level)
        if dedup_key in seen_node_items:
            continue
        seen_node_items.add(dedup_key)

        if joint_id:
            node_items_by_joint[joint_id].append({
                "joint_id": joint_id,
                "brace": brace,
                "joint_type": _txt(row.get("joint_type")),
                "inspect_level": inspect_level,
                "time_node": _txt(row.get("time_node")),
            })

    return {
        "facility_code": facility_code,
        "run_id": run_id,
        "display_year": display_year,
        "context_year": context_year,
        "source": "snapshot_context_fallback",
        "source_workbook": "",
        "member_level_by_key": dict(member_level_by_key),
        "node_level_by_joint": dict(node_level_by_joint),
        "node_level_by_joint_brace": dict(node_level_by_joint_brace),
        "member_items_by_key": dict(member_items_by_key),
        "node_items_by_joint": dict(node_items_by_joint),
    }


def _empty_overlay(
    facility_code: str,
    run_id: int | None,
    display_year: str,
    context_year: str,
) -> dict[str, Any]:
    return {
        "facility_code": facility_code,
        "run_id": run_id,
        "display_year": display_year,
        "context_year": context_year,
        "source": "empty",
        "source_workbook": "",
        "member_level_by_key": {},
        "node_level_by_joint": {},
        "node_level_by_joint_brace": {},
        "member_items_by_key": {},
        "node_items_by_joint": {},
    }


def load_strategy_inspection_overlay(
    facility_code: str,
    *,
    run_id: int | None = None,
    display_year: str = "当前",
) -> dict[str, Any]:
    code = _normalize_facility_code(facility_code)
    context_year = _norm_time_label(display_year)

    for workbook_path in _candidate_pipeline_paths(code, run_id=run_id):
        try:
            overlay = _read_pipeline_overlay(workbook_path, context_year)
            overlay.update({
                "facility_code": code,
                "run_id": run_id,
                "display_year": display_year,
                "context_year": context_year,
            })

            print(
                "[InspectionOverlay] source=pipeline_xlsx",
                "year=", context_year,
                "workbook=", workbook_path,
                "member=", len(overlay.get("member_level_by_key") or {}),
                "node=", len(overlay.get("node_level_by_joint") or {}),
            )
            return overlay

        except Exception as exc:
            print("[InspectionOverlay] read pipeline workbook failed:", workbook_path, exc)

    overlay = _load_context_overlay_fallback(code, run_id, display_year)
    print(
        "[InspectionOverlay] source=", overlay.get("source"),
        "year=", overlay.get("context_year"),
        "member=", len(overlay.get("member_level_by_key") or {}),
        "node=", len(overlay.get("node_level_by_joint") or {}),
    )
    return overlay


def get_member_inspect_level(
    overlay: dict[str, Any],
    joint_a: str,
    joint_b: str,
) -> str:
    key = _member_key(joint_a, joint_b)
    return _txt((overlay.get("member_level_by_key") or {}).get(key))


def get_node_inspect_level(
    overlay: dict[str, Any],
    joint_id: str,
) -> str:
    return _txt((overlay.get("node_level_by_joint") or {}).get(_txt(joint_id)))


def get_node_brace_inspect_level(
    overlay: dict[str, Any],
    joint_id: str,
    brace: str,
) -> str:
    key = f"{_txt(joint_id)}|{_txt(brace)}"
    return _txt((overlay.get("node_level_by_joint_brace") or {}).get(key))