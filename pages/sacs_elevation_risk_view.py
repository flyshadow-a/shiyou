# -*- coding: utf-8 -*-
from __future__ import annotations

import os
from typing import Dict, List, Tuple, Optional

from PyQt5.QtCore import Qt, QRectF, QTimer, QPointF
from PyQt5.QtGui import QBrush, QColor, QPainter, QPen, QFont, QImage, QPixmap, QPolygonF
from PyQt5.QtWidgets import (
    QGraphicsEllipseItem,
    QGraphicsLineItem,
    QGraphicsRectItem,
    QGraphicsScene,
    QGraphicsSimpleTextItem,
    QGraphicsView,
    QDialog,
    QVBoxLayout,
    QHBoxLayout,
    QLabel,
    QPushButton,
    QWidget,
    QSlider,
)

from services.file_db_adapter import (
    FileBackendError,
    is_file_db_configured,
    list_storage_paths,
    list_storage_paths_by_prefix,
)
from core.app_paths import external_path, first_existing_path
from core.dialog_utils import exec_dialog_safely

import traceback
from pages.sacs_storage_service import get_job_runtime_dir, get_job_source_dir
import openpyxl
import csv
import json

try:
    import pandas as pd
except Exception:
    pd = None


def read_sacs_elevation_lines(file_path: str) -> List[str]:
    encodings = ["utf-8", "utf-8-sig", "gb18030", "gbk", "latin-1"]
    for enc in encodings:
        try:
            with open(file_path, "r", encoding=enc) as f:
                return f.readlines()
        except UnicodeDecodeError:
            continue
        except Exception:
            break
    with open(file_path, "r", encoding="utf-8", errors="ignore") as f:
        return f.readlines()




def _sacs_float(text: str, default: float = 0.0) -> float:
    """SACS 固定列数值解析：空白字段按 0 处理。"""
    value = str(text or "").strip()
    if not value:
        return default
    try:
        return float(value)
    except Exception:
        return default


def _parse_sacs_group_od(line: str):
    text = str(line or "")
    if not text.startswith("GRUP"):
        return None
    gid = text[5:8].strip()
    if not gid:
        return None
    try:
        od_text = text[14:24].strip()
        od = float(od_text) if od_text else 0.0
    except Exception:
        od = 0.0
    return gid, od


def _parse_sacs_joint_line(line: str):
    """
    解析 SACS JOINT 固定列。
    SACS 中空白坐标表示 0；后续 33/40/47 列附近可能有 cm 偏移字段，
    需要按 VBA ReadSACS 逻辑加回坐标。
    """
    text = str(line or "")
    if not text.startswith("JOINT"):
        return None

    nid = text[6:10].strip()
    if not nid:
        return None

    x = _sacs_float(text[11:18], 0.0)
    y = _sacs_float(text[18:25], 0.0)
    z = _sacs_float(text[25:32], 0.0)

    # 兼容 JOINT 后续 cm 偏移字段：VBA 中是 /100 加回。
    x += _sacs_float(text[32:39], 0.0) / 100.0
    y += _sacs_float(text[39:46], 0.0) / 100.0
    z += _sacs_float(text[46:53], 0.0) / 100.0

    return nid, x, y, z


def _parse_sacs_member_line(line: str):
    """
    解析 MEMBER / MEMBER1 固定列。
    - MEMBER OFFSETS / MEMB2 OFFSETS 是上一根构件的偏移信息，不是新构件；
    - MEMBER1 后面没有空格，也要按同样的 joint_a/joint_b/group_id 固定列读取；
    - group_id 应读 16:19，不能读 15:18，否则 WB1 会被读成 WB。
    """
    text = str(line or "")
    upper = text.upper()

    if upper.startswith("MEMBER OFFSETS"):
        return None
    if upper.startswith("MEMB2"):
        return None
    if not upper.startswith("MEMBER"):
        return None

    joint_a = text[7:11].strip()
    joint_b = text[11:15].strip()
    group_id = text[16:19].strip()

    if joint_a and joint_b:
        return joint_a, joint_b, group_id
    return None


def parse_sacs_elevation_model(filepath: str):
    nodes = {}
    members = []
    groups_od = {}

    for line in read_sacs_elevation_lines(filepath):
        group = _parse_sacs_group_od(line)
        if group is not None:
            gid, od = group
            groups_od[gid] = od
            continue

        joint = _parse_sacs_joint_line(line)
        if joint is not None:
            nid, x, y, z = joint
            nodes[nid] = (x, y, z)
            continue

        member = _parse_sacs_member_line(line)
        if member is not None:
            members.append(member)

    return nodes, members, groups_od


def _score_elevation_model_path(path: str, facility_code: str = "") -> int:
    if not path:
        return 0
    name = os.path.basename(path).lower()
    path_low = path.lower()
    score = 0
    if name.startswith("sacinp"):
        score += 300
    if name.endswith(".sacinp"):
        score += 220
    if "褰撳墠妯″瀷" in path:
        score += 80
    if "缁撴瀯妯″瀷" in path:
        score += 60
    if facility_code and facility_code.lower() in path_low:
        score += 200
    if "user" in path_low or "鐢ㄦ埛涓婁紶" in path_low:
        score += 20
    return score


def resolve_elevation_model_path(facility_code: str) -> str:
    code = (facility_code or "").strip() or None
    candidates: List[str] = []

    if is_file_db_configured():
        query_specs = [
            ("model_files", f"{facility_code}/褰撳墠妯″瀷/缁撴瀯妯″瀷"),
            ("model_files", f"{facility_code}/褰撳墠妯″瀷/缁撴瀯妯″瀷/鐢ㄦ埛涓婁紶"),
            ("special_strategy", f"{facility_code}/褰撳墠妯″瀷/缁撴瀯妯″瀷"),
            ("special_strategy", f"{facility_code}/褰撳墠妯″瀷/缁撴瀯妯″瀷/鐢ㄦ埛涓婁紶"),
        ]

        for module_code, logical_prefix in query_specs:
            try:
                rows = list_storage_paths_by_prefix(
                    file_type_code="model",
                    module_code=module_code,
                    logical_path_prefix=logical_prefix,
                    facility_code=code,
                )
                if rows:
                    candidates.extend(rows)
            except FileBackendError:
                pass

        if not candidates:
            for module_code, logical_prefix in query_specs:
                try:
                    rows = list_storage_paths_by_prefix(
                        file_type_code="model",
                        module_code=module_code,
                        logical_path_prefix=logical_prefix,
                        facility_code=None,
                    )
                    if rows:
                        candidates.extend(rows)
                except FileBackendError:
                    pass

        if not candidates:
            for module_code in ("model_files", "special_strategy"):
                try:
                    rows = list_storage_paths(
                        file_type_code="model",
                        module_code=module_code,
                        facility_code=code,
                    )
                    if rows:
                        candidates.extend(rows)
                except FileBackendError:
                    pass

    if not candidates:
        roots = [
            external_path("upload", "model_files"),
            first_existing_path("upload", "model_files"),
        ]
        for root in roots:
            if not root or not os.path.isdir(root):
                continue
            for dir_path, _, file_names in os.walk(root):
                for file_name in file_names:
                    if file_name.lower().startswith("sacinp") or file_name.lower().endswith(".sacinp"):
                        candidates.append(os.path.join(dir_path, file_name))

    usable = [os.path.normpath(path) for path in candidates if path and os.path.isfile(path)]
    if not usable:
        return ""
    usable.sort(key=lambda p: (_score_elevation_model_path(p, facility_code), os.path.getmtime(p)), reverse=True)
    return usable[0]


class RiskNodeItem(QGraphicsEllipseItem):
    def __init__(self, rect: QRectF, tooltip: str = "", hover_callback=None, *args, **kwargs):
        super().__init__(rect, *args, **kwargs)
        self.setAcceptHoverEvents(True)
        self.setToolTip(tooltip)
        self._hover_callback = hover_callback

    def hoverEnterEvent(self, event):
        if self._hover_callback:
            self._hover_callback(self.toolTip())
        super().hoverEnterEvent(event)

    def hoverLeaveEvent(self, event):
        if self._hover_callback:
            self._hover_callback("")
        super().hoverLeaveEvent(event)


class RiskMemberItem(QGraphicsLineItem):
    def __init__(self, x1, y1, x2, y2, tooltip: str = "", hover_callback=None, *args, **kwargs):
        super().__init__(x1, y1, x2, y2, *args, **kwargs)
        self.setAcceptHoverEvents(True)
        self.setToolTip(tooltip)
        self._hover_callback = hover_callback

    def hoverEnterEvent(self, event):
        if self._hover_callback:
            self._hover_callback(self.toolTip())
        super().hoverEnterEvent(event)

    def hoverLeaveEvent(self, event):
        if self._hover_callback:
            self._hover_callback("")
        super().hoverLeaveEvent(event)


class CachedHotspotItem(QGraphicsRectItem):
    """透明热点层，用于缓存 PNG / 预渲染图片的鼠标悬停提示。"""

    def __init__(self, rect: QRectF, tooltip: str = "", *args, **kwargs):
        super().__init__(rect, *args, **kwargs)
        self.setAcceptHoverEvents(True)
        self.setToolTip(str(tooltip or ""))
        self.setPen(QPen(Qt.NoPen))
        # 使用几乎透明的填充，保证图元有可命中的 shape，但视觉上不可见。
        self.setBrush(QBrush(QColor(0, 0, 0, 1)))
        self.setZValue(1000)



class SacsElevationRiskView(QGraphicsView):

    COLOR_BG = QColor(255, 255, 255)

    # 轮廓颜色：单立面更清晰，全部总览略浅但不能太淡
    COLOR_MEMBER_FACE = QColor(0, 0, 0)  # 当前面：黑色
    COLOR_MEMBER_OTHER = QColor(170, 170, 170)  # 杂线：灰色
    COLOR_MARKER_FACE = QColor(0, 0, 0)
    COLOR_MARKER_OTHER = QColor(170, 170, 170)
    COLOR_GRID = QColor("#d7dde6")            # 背景网格
    COLOR_AXIS = QColor("#5c6470")            # 轴线/文字
    COLOR_TEXT = QColor("#5c6470")

    SHOW_NODE_TEXT = False
    SHOW_MEMBER_TEXT = False
    DRAW_NONRISK_NODES = False

    HORIZONTAL_EXAGGERATION_ALL = 2.2
    HORIZONTAL_EXAGGERATION_ROW = 1.75

    def __init__(self, parent=None):
        super().__init__(parent)

        # 直线图其实不需要太强的抗锯齿，关掉后首帧会更轻一些
        self.setRenderHint(QPainter.Antialiasing, False)
        # 文字单独开启抗锯齿，避免历史检测图例预渲染后发虚。
        self.setRenderHint(QPainter.TextAntialiasing, True)
        self.setBackgroundBrush(QBrush(self.COLOR_BG))
        self.setFrameShape(QGraphicsView.NoFrame)
        self.setDragMode(QGraphicsView.NoDrag)
        self.setTransformationAnchor(QGraphicsView.AnchorViewCenter)
        self.setResizeAnchor(QGraphicsView.AnchorViewCenter)

        # 关键：关闭 QGraphicsView 自带滚动条
        self.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        self.setVerticalScrollBarPolicy(Qt.ScrollBarAlwaysOff)

        self.setFocusPolicy(Qt.StrongFocus)
        self.setCursor(Qt.ArrowCursor)
        self.setInteractive(True)

        # 减少重绘负担
        self.setViewportUpdateMode(QGraphicsView.MinimalViewportUpdate)
        self.setOptimizationFlag(QGraphicsView.DontSavePainterState, True)
        self.setOptimizationFlag(QGraphicsView.DontAdjustForAntialiasing, True)
        self.setCacheMode(QGraphicsView.CacheBackground)

        self._reset_pending = False
        self._in_reset_view = False
        self._fit_done = False
        self._zoom_steps = 0

        self._slider_h = None
        self._slider_v = None
        self._initial_scene_rect = QRectF()
        self._last_pan_x = 0
        self._last_pan_y = 0

        self._scene = QGraphicsScene(self)
        self._scene.setItemIndexMethod(QGraphicsScene.NoIndex)
        self.setScene(self._scene)

        self.nodes: Dict[str, Tuple[float, float, float]] = {}
        self.members: List[Tuple[str, str, str]] = []
        self.node_risk_map: Dict[str, str] = {}
        self.member_risk_map: Dict[Tuple[str, str], str] = {}
        self._groups_od: Dict[str, float] = {}

        self._facility_code = ""
        self._model_path = ""
        self._row_name = "ROW A"
        self._info_label = None
        self._workpoint_z = 9.1
        self._level_threshold = 40

        self._inspection_overlay_enabled = False
        self._member_inspect_level_map = {}
        self._node_inspect_level_map = {}
        self._node_brace_inspect_level_map = {}

        self._member_items_by_key = {}
        self._node_items_by_joint = {}
        self._placed_label_rects = []

        self._history_overlay_enabled = False
        self._history_overlay_items = []
        self._history_overlay_legend = []

        # ===== 检验等级叠加显示策略 =====
        # 是否显示每个等级由页面显式控制；默认保持旧行为，显示 II/III/IV。
        self._inspection_visible_levels = {"II", "III", "IV"}
        self._show_inspection_level_ii = True

        # 保留原有字段，兼容其它调用；当前新的结果页不再依赖这些文字牌级别集合。
        self._member_badge_levels = {"III", "IV"}
        self._node_badge_levels = {"III", "IV"}

        # 文字牌抽稀距离，避免同一区域密集堆叠。
        self._badge_min_gap = {
            "II": 62.0,
            "III": 52.0,
            "IV": 42.0,
        }
        self._placed_badge_centers = []

        self._visible_projected_members = []
        self._visible_projected_nodes = {}

        # 模型缓存
        self._cached_model_path = ""
        self._cached_nodes = {}
        self._cached_members = []
        self._cached_groups_od = {}

        # 立面缓存
        self._row_render_cache = {}
        self._last_render_cache_key = None
    # ---------------- UI helper ----------------
    def set_info_label(self, label):
        self._info_label = label

    def bind_sliders(self, slider_h, slider_v):
        self._slider_h = slider_h
        self._slider_v = slider_v

    def reset_pan_state(self):
        self._last_pan_x = 0
        self._last_pan_y = 0

        if self._slider_h is not None:
            self._slider_h.blockSignals(True)
            self._slider_h.setValue(0)
            self._slider_h.blockSignals(False)

        if self._slider_v is not None:
            self._slider_v.blockSignals(True)
            self._slider_v.setValue(0)
            self._slider_v.blockSignals(False)

    def pan_view(self, x_value: int, y_value: int):
        if not self._initial_scene_rect.isValid():
            return

        # 先取当前视口映射到 scene 的实际可见范围
        visible_rect = self.mapToScene(self.viewport().rect()).boundingRect()

        # 在“场景尺寸 - 当前可见尺寸”的基础上，额外给一点缓冲
        max_dx = max(
            0.0,
            (self._initial_scene_rect.width() - visible_rect.width()) / 2.0
            + self._initial_scene_rect.width() * 0.10
        )
        max_dy = max(
            0.0,
            (self._initial_scene_rect.height() - visible_rect.height()) / 2.0
            + self._initial_scene_rect.height() * 0.10
        )

        cx = self._initial_scene_rect.center().x() + max_dx * (x_value / 100.0)
        cy = self._initial_scene_rect.center().y() - max_dy * (y_value / 100.0)

        self.centerOn(QPointF(cx, cy))

    def _add_cached_hotspots(self, hotspots: list | None, scale_x: float = 1.0, scale_y: float = 1.0) -> None:
        """在缓存图片上叠加透明热点，保留 tooltip 交互能力。"""
        if not hotspots:
            return
        for item in hotspots:
            try:
                tooltip = str(item.get("tooltip") or "").strip()
                if not tooltip:
                    continue
                rect_values = item.get("rect") or item.get("pixel_rect") or []
                if isinstance(rect_values, dict):
                    x = float(rect_values.get("x", 0))
                    y = float(rect_values.get("y", 0))
                    w = float(rect_values.get("w", rect_values.get("width", 0)))
                    h = float(rect_values.get("h", rect_values.get("height", 0)))
                else:
                    if len(rect_values) < 4:
                        continue
                    x, y, w, h = [float(v) for v in rect_values[:4]]
                if w <= 0 or h <= 0:
                    continue
                rect = QRectF(x * scale_x, y * scale_y, w * scale_x, h * scale_y)
                # 稍微放大命中范围，避免缓存图缩放后鼠标难以命中。
                pad = max(3.0, min(rect.width(), rect.height()) * 0.35)
                rect = rect.adjusted(-pad, -pad, pad, pad)
                hotspot = CachedHotspotItem(rect, tooltip=tooltip)
                self._scene.addItem(hotspot)
            except Exception:
                continue

    def display_cached_pixmap(
        self,
        pixmap: QPixmap,
        *,
        row_name: str = "",
        description: str = "缓存图",
        hotspots: list | None = None,
    ) -> bool:
        """直接显示内存中的预渲染图片，并可叠加透明 tooltip 热点。"""
        if pixmap is None or pixmap.isNull():
            self._draw_message("缓存图读取失败")
            return False
        try:
            self._scene.blockSignals(True)
            self.setUpdatesEnabled(False)

            self._scene.clear()
            item = self._scene.addPixmap(pixmap)
            item.setPos(0, 0)
            item.setZValue(0)

            self._add_cached_hotspots(hotspots or [])

            self._row_name = (row_name or self._row_name or "XZ 前").strip()
            rect = QRectF(pixmap.rect()).adjusted(-8, -8, 8, 8)
            self._scene.setSceneRect(rect)
            self._initial_scene_rect = QRectF(rect)
            self._fit_done = False
            self._zoom_steps = 0
            self.reset_pan_state()

            if self._info_label is not None:
                self._info_label.setText(
                    f"当前显示：{self._row_name} {description}；滚轮缩放，双击恢复初始视图。"
                )

            if not self._reset_pending:
                self._reset_pending = True
                QTimer.singleShot(0, self.reset_view)
            return True
        finally:
            self._scene.blockSignals(False)
            self.setUpdatesEnabled(True)
            self.viewport().update()

    def build_hotspot_metadata_for_rect(self, source_rect: QRectF, target_width: int, target_height: int) -> list[dict]:
        """把当前 scene 中带 tooltip 的交互图元转换成图片像素坐标热点。"""
        if self.scene() is None or source_rect is None or (not source_rect.isValid()) or source_rect.isNull():
            return []
        sx = float(target_width) / max(1e-6, float(source_rect.width()))
        sy = float(target_height) / max(1e-6, float(source_rect.height()))
        hotspots: list[dict] = []
        for item in self.scene().items():
            try:
                tooltip = str(item.toolTip() or "").strip()
                if not tooltip:
                    continue
                if isinstance(item, CachedHotspotItem):
                    continue
                if not isinstance(item, (RiskNodeItem, RiskMemberItem)):
                    continue
                br = item.sceneBoundingRect()
                if not br.isValid() or br.isNull():
                    continue
                # 线段热点加粗；点热点略放大，便于鼠标命中。
                if isinstance(item, RiskMemberItem):
                    br = br.adjusted(-5, -5, 5, 5)
                    kind = "member"
                else:
                    br = br.adjusted(-4, -4, 4, 4)
                    kind = "node"
                x = (br.left() - source_rect.left()) * sx
                y = (br.top() - source_rect.top()) * sy
                w = br.width() * sx
                h = br.height() * sy
                # 跳过完全落在导出区域外的热点。
                if x + w < 0 or y + h < 0 or x > target_width or y > target_height:
                    continue
                x = max(0.0, min(float(target_width), x))
                y = max(0.0, min(float(target_height), y))
                w = max(1.0, min(float(target_width) - x, w))
                h = max(1.0, min(float(target_height) - y, h))
                hotspots.append({
                    "kind": kind,
                    "tooltip": tooltip,
                    "rect": [round(x, 2), round(y, 2), round(w, 2), round(h, 2)],
                })
            except Exception:
                continue
        return hotspots

    def build_hotspot_metadata_for_export(self, margin: int = 24, scale: float = 1.0) -> dict:
        """生成与 export_current_scene_to_png 导出范围一致的热点元数据。

        如果导出 PNG 时传入了 scale，这里也要传入相同的 scale，
        避免缓存图热点坐标与图片像素尺寸不一致。
        """
        if self.scene() is None:
            return {"hotspots": []}
        rect = self.scene().itemsBoundingRect()
        if (not rect.isValid()) or rect.isNull():
            rect = self.scene().sceneRect()
        if (not rect.isValid()) or rect.isNull():
            return {"hotspots": []}
        rect = rect.adjusted(-margin, -margin, margin, margin)

        try:
            export_scale = float(scale)
        except Exception:
            export_scale = 1.0
        if export_scale <= 0:
            export_scale = 1.0

        width = max(1, int(rect.width() * export_scale))
        height = max(1, int(rect.height() * export_scale))
        return {
            "source_rect": [rect.left(), rect.top(), rect.width(), rect.height()],
            "image_size": [width, height],
            "scale": export_scale,
            "hotspots": self.build_hotspot_metadata_for_rect(rect, width, height),
        }

    def display_cached_image(
        self,
        image_path: str,
        row_name: str = "",
        description: str = "缓存检验等级图",
        hotspots: list | None = None,
    ) -> bool:
        """直接显示服务器中已导出的 PNG 缓存图。

        用途：特检策略主页右上角“查看结果”查看历史最新结果时，
        优先读取图二“生成特检策略报告”时导出的风险等级图，避免重新解析模型和绘制。
        该方法只显示图片，不依赖节点/构件坐标；仍保留滚轮缩放、双击复位、外部滑条平移。
        如果传入 hotspots，会在缓存图上叠加透明热点，保留鼠标悬停提示能力。
        """
        path = os.path.normpath(str(image_path or "").strip())
        if not path or not os.path.exists(path):
            self._draw_message("未找到缓存检验等级图")
            return False

        pixmap = QPixmap(path)
        if pixmap.isNull():
            self._draw_message("缓存检验等级图读取失败")
            return False

        return self.display_cached_pixmap(
            pixmap,
            row_name=row_name,
            description=description,
            hotspots=hotspots or [],
        )

    def export_current_scene_to_png(self, file_path: str, margin: int = 24, scale: float = 1.0) -> str:
        """把当前 QGraphicsScene 导出为 PNG 图片。

        scale 用于提高导出分辨率：
        - 轮廓图通常传 2.5；
        - 检验等级图通常传 2.0；
        - 页面普通导出可以保持默认 1.0。
        """
        if self.scene() is None:
            raise ValueError("当前没有可导出的图形场景")

        rect = self.scene().itemsBoundingRect()
        if (not rect.isValid()) or rect.isNull():
            rect = self.scene().sceneRect()
        if (not rect.isValid()) or rect.isNull():
            raise ValueError("当前没有可导出的立面风险图")

        rect = rect.adjusted(-margin, -margin, margin, margin)

        try:
            export_scale = float(scale)
        except Exception:
            export_scale = 1.0
        if export_scale <= 0:
            export_scale = 1.0

        width = max(1, int(rect.width() * export_scale))
        height = max(1, int(rect.height() * export_scale))

        image = QImage(width, height, QImage.Format_ARGB32)
        image.fill(Qt.white)

        painter = QPainter(image)
        painter.setRenderHint(QPainter.Antialiasing, True)
        painter.setRenderHint(QPainter.TextAntialiasing, True)
        painter.setRenderHint(QPainter.SmoothPixmapTransform, True)
        self.scene().render(
            painter,
            target=QRectF(0, 0, width, height),
            source=rect,
        )
        painter.end()

        folder = os.path.dirname(os.path.normpath(file_path))
        if folder:
            os.makedirs(folder, exist_ok=True)
        if not image.save(file_path, "PNG"):
            raise IOError(f"保存图片失败：{file_path}")
        return os.path.normpath(file_path)

    def _show_hover_info(self, _text: str = ""):
        if self._info_label is None:
            return
        self._info_label.setText(
            f"当前显示：{self._row_name} 立面轮廓图；滚轮缩放，双击恢复初始视图。"
        )

    def reset_view(self):
        if self._in_reset_view:
            return

        self._in_reset_view = True
        self._reset_pending = False
        try:
            rect = self._scene.sceneRect()
            if (not rect.isValid()) or rect.isNull():
                rect = self._scene.itemsBoundingRect()
            if not rect.isValid():
                return

            rect = rect.adjusted(-10, -10, 10, 10)

            # 防止 fitInView 期间再次触发 resizeEvent
            self._fit_done = True

            self._scene.setSceneRect(rect)
            self.resetTransform()
            self.fitInView(rect, Qt.KeepAspectRatio)

            self._initial_scene_rect = QRectF(rect)
            self._zoom_steps = 0
            self.centerOn(rect.center())
            self.reset_pan_state()
            self._show_hover_info()
        finally:
            self._in_reset_view = False

    def _ensure_model_loaded(self):
        cached_model_path = getattr(self, "_cached_model_path", "")
        cached_nodes = getattr(self, "_cached_nodes", {})
        cached_members = getattr(self, "_cached_members", [])
        cached_groups_od = getattr(self, "_cached_groups_od", {})

        # 命中模型缓存：直接复用
        if (
                self._model_path
                and self._model_path == cached_model_path
                and cached_nodes
                and cached_members
        ):
            self.nodes = cached_nodes
            self.members = cached_members
            self._groups_od = cached_groups_od
            return

        # 重新解析模型
        self.nodes, self.members, self._groups_od = self.parse_sacs_full_robust(self._model_path)

        self._cached_model_path = self._model_path
        self._cached_nodes = dict(self.nodes)
        self._cached_members = list(self.members)
        self._cached_groups_od = dict(self._groups_od)

        # 模型变了，立面缓存必须清掉
        self._clear_render_cache()

    def prime_model_cache(self, model_path: str, nodes, members, groups_od) -> None:
        normalized_path = os.path.normpath(str(model_path or "").strip())
        if not normalized_path:
            return
        self._cached_model_path = normalized_path
        self._cached_nodes = dict(nodes or {})
        self._cached_members = list(members or [])
        self._cached_groups_od = dict(groups_od or {})
        self._clear_render_cache()

    # ---------------- 外部入口 ----------------
    def load_for_facility(
            self,
            facility_code: str,
            context: Dict,
            year_label: str,
            row_name: str = "XZ 前",
            workpoint_override: Optional[float] = None,
            level_threshold_override: Optional[int] = None,
            model_path_override: Optional[str] = None,
    ) -> None:
        try:
            self._facility_code = (facility_code or "").strip()
            self._row_name = (row_name or "XZ 前").strip()

            # model_path_override 只在“报告导出原模型/改造后模型轮廓图”时传入。
            # 注意：不能直接 os.path.normpath("")，否则空字符串会被转换成 "."，
            # 后续会把当前目录当作模型文件读取，导致 PermissionError: [Errno 13] Permission denied: '.'.
            raw_override = str(model_path_override or "").strip()
            override_model = os.path.normpath(raw_override) if raw_override else ""

            if override_model and os.path.isfile(override_model):
                self._model_path = override_model
            else:
                # 没有传入有效 override 时，仍按平台编码解析当前模型文件。
                self._model_path = self._resolve_model_path(self._facility_code)

            print("[Elevation] facility_code =", self._facility_code)
            print("[Elevation] row_name =", self._row_name)
            print("[Elevation] model_path_override =", override_model)
            print("[Elevation] resolved model_path =", self._model_path)

            if not self._model_path or not os.path.isfile(self._model_path):
                self._draw_message("未找到结构模型文件")
                return

            self._ensure_model_loaded()

            if not self.nodes or not self.members:
                self._draw_message("模型文件未解析到有效 JOINT/MEMBER")
                return

            self.node_risk_map = {}
            self.member_risk_map = {}

            auto_wp = self._extract_workpoint_from_context(context)
            self._workpoint_z = auto_wp if workpoint_override in (None, "") else self._safe_float(
                workpoint_override, auto_wp
            )

            auto_thr = self._extract_level_threshold_from_context(context)
            if level_threshold_override in (None, ""):
                self._level_threshold = auto_thr
            else:
                try:
                    self._level_threshold = int(level_threshold_override)
                except Exception:
                    self._level_threshold = auto_thr

            # 只有非固定立面时，才去动态检查 row_name
            fixed_rows = {"XZ 前", "XZ 后", "YZ 左", "YZ 右"}
            if self._row_name not in fixed_rows and not self._row_name.startswith("XY "):
                row_names = self.available_row_names()
                if row_names and self._row_name not in row_names:
                    self._row_name = row_names[0]

            self._render_row_elevation()

            self._fit_done = False
            self._zoom_steps = 0

            if not self._reset_pending:
                self._reset_pending = True
                QTimer.singleShot(0, self.reset_view)

        except Exception:
            print("[Elevation] load_for_facility failed")
            traceback.print_exc()
            self._draw_message("立面图加载失败，请查看控制台日志")

    # ---------------- 模型文件路径 ----------------
    def _resolve_model_path(self, facility_code: str) -> str:
        code = (facility_code or "").strip() or None
        candidates: List[str] = []

        if is_file_db_configured():
            query_specs = [
                ("model_files", f"{facility_code}/当前模型/结构模型"),
                ("model_files", f"{facility_code}/当前模型/结构模型/用户上传"),
                ("special_strategy", f"{facility_code}/当前模型/结构模型"),
                ("special_strategy", f"{facility_code}/当前模型/结构模型/用户上传"),
            ]

            for module_code, logical_prefix in query_specs:
                try:
                    rows = list_storage_paths_by_prefix(
                        file_type_code="model",
                        module_code=module_code,
                        logical_path_prefix=logical_prefix,
                        facility_code=code,
                    )
                    if rows:
                        candidates.extend(rows)
                except FileBackendError:
                    pass

            if not candidates:
                for module_code, logical_prefix in query_specs:
                    try:
                        rows = list_storage_paths_by_prefix(
                            file_type_code="model",
                            module_code=module_code,
                            logical_path_prefix=logical_prefix,
                            facility_code=None,
                        )
                        if rows:
                            candidates.extend(rows)
                    except FileBackendError:
                        pass

            if not candidates:
                for module_code in ("model_files", "special_strategy"):
                    try:
                        rows = list_storage_paths(
                            file_type_code="model",
                            module_code=module_code,
                            facility_code=code,
                        )
                        if rows:
                            candidates.extend(rows)
                    except Exception:
                        pass

        upload_roots = [
            external_path("upload", "model_files"),
            first_existing_path("upload", "model_files"),
            r"Y:\upload\model_files",
            self._special_strategy_inputs_dir(),
        ]

        for upload_root in upload_roots:
            if not upload_root or not os.path.isdir(upload_root):
                continue
            for root, _dirs, files in os.walk(upload_root):
                for fn in files:
                    low = fn.lower()
                    if low.startswith("sacinp") or low.endswith(".sacinp") or "sacinp" in low:
                        candidates.append(os.path.join(root, fn))

        uniq = []
        seen = set()
        for p in candidates:
            np = os.path.normpath(str(p))
            if np not in seen:
                seen.add(np)
                uniq.append(np)

        if not uniq:
            return ""

        uniq.sort(
            key=lambda p: (
                self._score_model_candidate(p, facility_code),
                os.path.getmtime(p) if os.path.exists(p) else 0
            ),
            reverse=True,
        )

        for p in uniq:
            if os.path.exists(p):
                return p

        return ""

    def clear_inspection_overlay(self):
        self._inspection_overlay_enabled = False
        self._member_inspect_level_map = {}
        self._node_inspect_level_map = {}
        self._node_brace_inspect_level_map = {}
        self._member_items_by_key = {}
        self._node_items_by_joint = {}
        self._placed_label_rects = []
        self._placed_badge_centers = []
        self._visible_projected_members = []
        self._visible_projected_nodes = {}

    def clear_history_overlay(self):
        self._history_overlay_enabled = False
        self._history_overlay_items = []
        self._history_overlay_legend = []

    def _history_color(self, color_text: str) -> QColor:
        color = QColor(str(color_text or "").strip())
        if color.isValid():
            return color
        return QColor("#E74C3C")

    def _history_tooltip_for_item(self, item: dict) -> str:
        parts = [f"节点: {str(item.get('joint_id') or '').strip()}"]
        round_label = str(item.get('round_label') or '').strip()
        if round_label:
            parts.append(f"检测批次: {round_label}")
        inspect_level = str(item.get('inspect_level') or '').strip()
        if inspect_level:
            parts.append(f"检验等级: {inspect_level}")
        risk_level = str(item.get('risk_level') or '').strip()
        if risk_level:
            parts.append(f"风险等级: {risk_level}")
        brace = str(item.get('brace') or '').strip()
        if brace:
            parts.append(f"Brace: {brace}")
        joint_type = str(item.get('joint_type') or '').strip()
        if joint_type:
            parts.append(f"类型: {joint_type}")
        conclusion = str(item.get('conclusion') or '').strip()
        if conclusion:
            parts.append(f"检验结论: {conclusion}")
        return "\n".join(parts)

    def _draw_history_legend(self):
        """绘制历史检测点图例。

        原来使用 Arial 8 号字，预渲染为图片后再缩放显示时会偏小、发虚。
        这里改为中文字体 + 更大字号 + 白色背景框，提高可读性。
        """
        if not self._history_overlay_legend:
            return

        x0 = 18.0
        y0 = 16.0
        dot_size = 16.0
        title_font = QFont("Microsoft YaHei UI", 14, QFont.Bold)
        item_font = QFont("Microsoft YaHei UI", 12, QFont.Bold)
        title_text = "历史检测点"

        legend_texts = []
        max_chars = len(title_text)
        for entry in self._history_overlay_legend:
            label = str(entry.get("round_label") or "").strip()
            count = entry.get("count")
            text = label
            if count not in (None, ""):
                text += f" ({count})"
            legend_texts.append((entry, text))
            max_chars = max(max_chars, len(text))

        panel_w = max(210.0, 28.0 + max_chars * 15.0)
        panel_h = 46.0 + max(1, len(legend_texts)) * 32.0
        bg = self._scene.addRect(
            QRectF(x0 - 10.0, y0 - 10.0, panel_w, panel_h),
            QPen(QColor("#8fa1b7"), 1.2),
            QBrush(QColor(255, 255, 255, 242)),
        )
        bg.setZValue(198)

        title = QGraphicsSimpleTextItem(title_text)
        title.setBrush(QBrush(QColor("#203040")))
        title.setFont(title_font)
        title.setPos(x0, y0)
        title.setZValue(200)
        self._scene.addItem(title)

        y = y0 + 32.0
        for entry, text in legend_texts:
            color = self._history_color(entry.get("round_color"))
            dot = self._scene.addEllipse(
                QRectF(x0, y + 4.0, dot_size, dot_size),
                QPen(color, 1.4),
                QBrush(color),
            )
            dot.setZValue(200)

            item = QGraphicsSimpleTextItem(text)
            item.setBrush(QBrush(QColor("#203040")))
            item.setFont(item_font)
            item.setPos(x0 + dot_size + 12.0, y - 1.0)
            item.setZValue(200)
            self._scene.addItem(item)
            y += 32.0

    def _draw_history_overlay(self):
        if not self._history_overlay_enabled:
            return
        if not self._history_overlay_items:
            return

        grouped = {}
        for item in self._history_overlay_items:
            joint_id = str(item.get('joint_id') or '').strip()
            if not joint_id:
                continue
            if joint_id not in self._visible_projected_nodes:
                continue
            grouped.setdefault(joint_id, []).append(item)

        for joint_id, rows in grouped.items():
            x, y = self._visible_projected_nodes[joint_id]
            total = len(rows)
            for idx, item in enumerate(rows):
                color = self._history_color(item.get('round_color'))
                if total <= 1:
                    dx = dy = 0.0
                else:
                    shift = 6.0
                    start = -(total - 1) / 2.0
                    dx = (start + idx) * shift
                    dy = - (start + idx) * shift * 0.35
                r = 4.2
                tooltip = self._history_tooltip_for_item(item)
                dot = RiskNodeItem(QRectF(x + dx - r, y + dy - r, 2 * r, 2 * r), tooltip=tooltip)
                dot.setPen(QPen(color, 1.2))
                dot.setBrush(QBrush(color))
                dot.setZValue(120)
                self._scene.addItem(dot)

        self._draw_history_legend()

    def _clear_render_cache(self):
        self._row_render_cache = {}
        self._last_render_cache_key = None

    def _get_row_render_cache_key(self, proj_mode: str):
        return (
            str(self._cached_model_path or self._model_path or "").strip(),
            str(proj_mode or "").strip(),
            round(float(self._workpoint_z), 4),
            int(self._level_threshold),
        )

    def _get_cached_side_face_result(self, proj_mode: str):
        cache_key = self._get_row_render_cache_key(proj_mode)
        cached = self._row_render_cache.get(cache_key)
        if cached is not None:
            return cached

        result = self._collect_side_face_segments(proj_mode)
        self._row_render_cache[cache_key] = result
        return result

    def _member_key_text(self, joint_a: str, joint_b: str) -> str:
        a = str(joint_a or "").strip()
        b = str(joint_b or "").strip()
        if not a and not b:
            return ""
        pair = sorted([a, b])
        return f"{pair[0]}|{pair[1]}"

    def _inspection_color(self, level: str) -> QColor:
        text = str(level or "").strip().upper()
        # II 只用颜色表达，为了和黑色结构线、III/IV 区分，使用偏蓝色。
        if text == "II":
            return QColor("#2D9CDB")
        # III 使用橙色，IV 使用红色，风险/检验等级越高越醒目。
        if text == "III":
            return QColor("#F2994A")
        if text == "IV":
            return QColor("#EB5757")
        return QColor("#5c6470")



    def _append_projected_member(self, bucket, joint_a: str, joint_b: str, p1, p2):
        bucket.append({
            "joint_a": str(joint_a or "").strip(),
            "joint_b": str(joint_b or "").strip(),
            "p1": (float(p1[0]), float(p1[1])),
            "p2": (float(p2[0]), float(p2[1])),
        })

    def _segment_points(self, seg):
        if isinstance(seg, dict):
            return seg["p1"], seg["p2"]
        return seg[0], seg[1]

    def _segment_joints(self, seg):
        if isinstance(seg, dict):
            return str(seg.get("joint_a") or "").strip(), str(seg.get("joint_b") or "").strip()
        return "", ""

    def set_visible_inspection_levels(self, levels) -> None:
        """设置当前立面图需要显示的检验等级。"""
        normalized = set()
        for level in levels or []:
            text = str(level or "").strip().upper()
            if text in ("II", "III", "IV"):
                normalized.add(text)
        if not normalized:
            normalized = {"III", "IV"}
        self._inspection_visible_levels = normalized
        self._show_inspection_level_ii = "II" in normalized
        if self.nodes and self.members:
            self._render_row_elevation()

    def set_show_level_ii(self, show: bool) -> None:
        show = bool(show)
        self._show_inspection_level_ii = show
        levels = {"III", "IV"}
        if show:
            levels.add("II")
        self.set_visible_inspection_levels(levels)

    def visible_inspection_levels(self) -> set[str]:
        return set(self._inspection_visible_levels or set())

    def _display_inspection_level(self, level: str) -> bool:
        return str(level or "").strip().upper() in (self._inspection_visible_levels or set())

    def _member_label_tooltip(self, key: str, level: str) -> str:
        return self._badge_tooltip_for_member(key, level)

    def _node_label_tooltip(self, joint_id: str, level: str) -> str:
        return self._badge_tooltip_for_node(joint_id, level)

    def _rect_overlaps_reserved_labels(self, rect: QRectF, margin: float = 3.0) -> bool:
        test_rect = rect.adjusted(-margin, -margin, margin, margin)
        for old_rect in self._placed_label_rects:
            if test_rect.intersects(old_rect):
                return True
        return False

    def _reserve_label_rect(self, rect: QRectF) -> None:
        self._placed_label_rects.append(rect)

    def _pick_first_free_rect(self, base_x: float, base_y: float, width: float, height: float, candidates: list[tuple[float, float]], margin: float = 3.0):
        for dx, dy in candidates:
            rect = QRectF(base_x + dx, base_y + dy, width, height)
            if not self._rect_overlaps_reserved_labels(rect, margin=margin):
                self._reserve_label_rect(rect)
                return rect
        return QRectF()

    def _arrow_label_candidates(self, ux: float, uy: float) -> list[tuple[float, float]]:
        nx = -uy
        ny = ux
        return [
            (nx * 14 - 12, ny * 14 - 8),
            (-nx * 14 - 12, -ny * 14 - 8),
            (nx * 22 - 12, ny * 22 - 8),
            (-nx * 22 - 12, -ny * 22 - 8),
            (nx * 8 - 12, ny * 8 - 8),
            (-nx * 8 - 12, -ny * 8 - 8),
        ]

    def _member_label_candidates(self, p1, p2) -> list[tuple[float, float]]:
        try:
            x1, y1 = float(p1[0]), float(p1[1])
            x2, y2 = float(p2[0]), float(p2[1])
        except Exception:
            return [(10, -16), (10, 6), (-22, -16), (-22, 6)]
        vx = x2 - x1
        vy = y2 - y1
        dist = (vx * vx + vy * vy) ** 0.5
        if dist < 1e-6:
            return [(10, -16), (10, 6), (-22, -16), (-22, 6)]
        nx = -vy / dist
        ny = vx / dist
        return [
            (nx * 16 - 12, ny * 16 - 8),
            (-nx * 16 - 12, -ny * 16 - 8),
            (nx * 24 - 12, ny * 24 - 8),
            (-nx * 24 - 12, -ny * 24 - 8),
            (nx * 10 - 12, ny * 10 - 8),
            (-nx * 10 - 12, -ny * 10 - 8),
        ]

    def _badge_tooltip_for_member(self, key: str, level: str) -> str:
        rows = self._member_items_by_key.get(key, []) or []
        if rows:
            first = rows[0]
            joint_a = str(first.get("joint_a", "")).strip()
            joint_b = str(first.get("joint_b", "")).strip()
            member_type = str(first.get("member_type", "")).strip()
            risk_level = str(first.get("risk_level", "")).strip()
            time_node = str(first.get("time_node", "")).strip()
            parts = [f"构件: {joint_a} - {joint_b}"]
            if member_type:
                parts.append(f"类型: {member_type}")
            if risk_level:
                parts.append(f"风险等级: {risk_level}")
            parts.append(f"检验等级: {level}")
            if time_node:
                parts.append(f"时间: {time_node}")
            return "\n".join(parts)
        return f"构件检验等级: {level}"

    def _badge_tooltip_for_node(self, joint_id: str, level: str) -> str:
        rows = self._node_items_by_joint.get(str(joint_id).strip(), []) or []
        if rows:
            first = rows[0]
            brace = str(first.get("brace", "")).strip()
            joint_type = str(first.get("joint_type", "")).strip()
            risk_level = str(first.get("risk_level", "")).strip()
            time_node = str(first.get("time_node", "")).strip()
            parts = [f"节点: {joint_id}"]
            if brace:
                parts.append(f"Brace: {brace}")
            if joint_type:
                parts.append(f"类型: {joint_type}")
            if risk_level:
                parts.append(f"风险等级: {risk_level}")
            parts.append(f"检验等级: {level}")
            if time_node:
                parts.append(f"时间: {time_node}")
            return "\n".join(parts)
        return f"节点 {joint_id}\n检验等级: {level}"

    def _can_place_badge(self, x: float, y: float, level: str) -> bool:
        """按文字牌中心距离抽稀，避免 III/IV 文字牌密集堆叠。"""
        min_gap = float(self._badge_min_gap.get(str(level).upper(), 30.0))
        min_gap2 = min_gap * min_gap
        for px, py in self._placed_badge_centers:
            dx = x - px
            dy = y - py
            if dx * dx + dy * dy < min_gap2:
                return False
        self._placed_badge_centers.append((x, y))
        return True

    def _find_brace_scene_point(self, joint_id: str, brace_id: str):
        """根据节点号和 Brace 号，找到箭头的起始方向点。

        优先使用当前立面中可见节点的投影坐标；如果 Brace 节点没有单独记录在
        _visible_projected_nodes 中，则从当前可见构件线段里寻找 joint-brace 这条线，
        用线段另一端作为方向来源。
        """
        joint_id = str(joint_id or "").strip()
        brace_id = str(brace_id or "").strip()
        if not joint_id or not brace_id or brace_id == joint_id:
            return None

        pt = self._visible_projected_nodes.get(brace_id)
        if pt is not None:
            return pt

        for seg in self._visible_projected_members or []:
            ja = str(seg.get("joint_a") or "").strip()
            jb = str(seg.get("joint_b") or "").strip()
            p1 = seg.get("scene_p1")
            p2 = seg.get("scene_p2")
            if not p1 or not p2:
                continue
            if ja == joint_id and jb == brace_id:
                return p2
            if jb == joint_id and ja == brace_id:
                return p1
        return None

    def _draw_direction_arrow(
        self,
        source_pt,
        target_pt,
        *,
        color: QColor,
        tooltip: str = "",
        z_value: float = 58.0,
        level_text: str = "",
    ) -> None:
        """绘制从 Brace 方向指向风险节点的短箭头，并把等级标在箭头上。"""
        try:
            sx, sy = float(source_pt[0]), float(source_pt[1])
            tx, ty = float(target_pt[0]), float(target_pt[1])
        except Exception:
            return

        vx = tx - sx
        vy = ty - sy
        dist = (vx * vx + vy * vy) ** 0.5
        if dist < 1e-6:
            return

        ux = vx / dist
        uy = vy / dist

        arrow_len = max(16.0, min(32.0, dist * 0.42))
        end_gap = 5.5
        x2 = tx - ux * end_gap
        y2 = ty - uy * end_gap
        x1 = tx - ux * (arrow_len + end_gap)
        y1 = ty - uy * (arrow_len + end_gap)

        pen = QPen(color, 1.6)
        pen.setCapStyle(Qt.RoundCap)
        pen.setJoinStyle(Qt.RoundJoin)
        line = RiskMemberItem(x1, y1, x2, y2, tooltip=tooltip)
        line.setPen(pen)
        line.setZValue(z_value)
        self._scene.addItem(line)

        head_len = 7.0
        head_w = 4.2
        bx = x2 - ux * head_len
        by = y2 - uy * head_len
        nx = -uy
        ny = ux
        polygon = QPolygonF([
            QPointF(x2, y2),
            QPointF(bx + nx * head_w, by + ny * head_w),
            QPointF(bx - nx * head_w, by - ny * head_w),
        ])
        head = self._scene.addPolygon(
            polygon,
            QPen(color, 1.0),
            QBrush(color),
        )
        head.setZValue(z_value + 0.5)
        head.setToolTip(str(tooltip or ""))

        level_text = str(level_text or "").strip()
        if level_text:
            self._draw_arrow_level_label(
                (x1 + x2) / 2.0,
                (y1 + y2) / 2.0,
                level_text,
                color,
                tooltip=tooltip,
                z_value=z_value + 0.8,
                ux=ux,
                uy=uy,
            )

    def _draw_arrow_level_label(self, x: float, y: float, level_text: str, color: QColor, *, tooltip: str = "", z_value: float = 59.0, ux: float = 1.0, uy: float = 0.0) -> None:
        text_item = QGraphicsSimpleTextItem(str(level_text or ""))
        text_item.setFont(QFont("Arial", 8, QFont.Bold))
        br = text_item.boundingRect()

        rect = self._pick_first_free_rect(
            x,
            y,
            br.width() + 10,
            br.height() + 6,
            self._arrow_label_candidates(ux, uy),
            margin=4.0,
        )
        if rect.isNull() or rect.width() <= 0:
            return

        bg = self._scene.addRect(
            rect,
            QPen(color, 1.0),
            QBrush(QColor(255, 255, 255, 238)),
        )
        bg.setZValue(z_value)
        bg.setToolTip(str(tooltip or ""))

        text_item.setBrush(QBrush(color))
        text_item.setPos(rect.x() + 5, rect.y() + 3)
        text_item.setZValue(z_value + 0.1)
        text_item.setToolTip(str(tooltip or ""))
        self._scene.addItem(text_item)

    def _draw_member_level_inline_label(self, x: float, y: float, level: str, color: QColor, *, tooltip: str = "", p1=None, p2=None) -> None:
        if not self._can_place_badge(x, y, level):
            return

        text_item = QGraphicsSimpleTextItem(str(level or ""))
        text_item.setFont(QFont("Arial", 9, QFont.Bold))
        br = text_item.boundingRect()
        rect = self._pick_first_free_rect(
            x,
            y,
            br.width() + 12,
            br.height() + 6,
            self._member_label_candidates(p1, p2),
            margin=4.0,
        )
        if rect.isNull() or rect.width() <= 0:
            return

        bg = self._scene.addRect(
            rect,
            QPen(color, 1.0),
            QBrush(QColor(255, 255, 255, 238)),
        )
        bg.setZValue(49)
        bg.setToolTip(str(tooltip or ""))

        text_item.setBrush(QBrush(color))
        text_item.setPos(rect.x() + 6, rect.y() + 3)
        text_item.setZValue(50)
        text_item.setToolTip(str(tooltip or ""))
        self._scene.addItem(text_item)

    def _draw_node_direction_arrows(self, joint_id: str, target_pt, fallback_level: str, fallback_tooltip: str) -> None:
        """为节点检验标注增加 Brace -> Joint 的方向箭头。

        业务含义：
        - 节点检验等级与 “brace -> joint” 来源关系有关；
        - 同一个 joint 可能由不同 brace 接入，且不同 brace 对应的检验等级可能不同；
        - 因此这里必须保留所有不同 brace 来源的箭头，不能只保留一条代表性箭头。

        显示策略：
        - 仅过滤当前页面允许显示的等级（默认 III/IV，点击按钮后包含 II）；
        - 对完全重复的 brace-level 做去重；
        - 每条箭头仍然绘制，等级文字如果空间太密，会自动跳过文字但保留箭头和 tooltip。
        """
        joint_id = str(joint_id or "").strip()
        if not joint_id:
            return

        rows = list(self._node_items_by_joint.get(joint_id, []) or [])
        if not rows:
            return

        seen_keys = set()
        draw_count = 0

        for row in rows:
            brace = str(row.get("brace") or "").strip()
            if not brace:
                continue

            level = str(row.get("inspect_level") or fallback_level or "").strip().upper()
            if level not in ("II", "III", "IV"):
                level = str(fallback_level or "").strip().upper()
            if level not in ("II", "III", "IV"):
                continue
            if not self._display_inspection_level(level):
                continue

            dedup_key = (brace, level)
            if dedup_key in seen_keys:
                continue
            seen_keys.add(dedup_key)

            source_pt = self._find_brace_scene_point(joint_id, brace)
            if source_pt is None:
                continue

            color = self._inspection_color(level)
            tooltip = self._badge_tooltip_for_node(joint_id, level)

            # 多个 brace 指向同一个 joint 时，给箭头做很小的法向偏移，
            # 避免箭头线完全盖在一起；偏移不改变“从 brace 指向 joint”的业务方向。
            offset_source_pt = source_pt
            offset_target_pt = target_pt
            try:
                sx, sy = float(source_pt[0]), float(source_pt[1])
                tx, ty = float(target_pt[0]), float(target_pt[1])
                vx = tx - sx
                vy = ty - sy
                dist = (vx * vx + vy * vy) ** 0.5
                if dist > 1e-6 and draw_count > 0:
                    nx = -vy / dist
                    ny = vx / dist
                    # 0, +4, -4, +8, -8 ... 的轻微错位序列
                    step = (draw_count + 1) // 2
                    sign = 1 if draw_count % 2 == 1 else -1
                    offset = sign * min(10.0, 4.0 * step)
                    offset_source_pt = (sx + nx * offset, sy + ny * offset)
                    offset_target_pt = (tx + nx * offset, ty + ny * offset)
            except Exception:
                offset_source_pt = source_pt
                offset_target_pt = target_pt

            self._draw_direction_arrow(
                offset_source_pt,
                offset_target_pt,
                color=color,
                tooltip=tooltip,
                z_value=57.0,
                level_text=level,
            )
            draw_count += 1

    def _draw_inspection_overlay(self):
        if not self._inspection_overlay_enabled:
            return

        self._placed_label_rects = []
        self._placed_badge_centers = []

        # ---------- 构件：仅显示当前配置中允许的等级；等级直接加粗并标在构件附近 ----------
        drawn_member_keys = set()

        for seg in self._visible_projected_members:
            joint_a = str(seg.get("joint_a") or "").strip()
            joint_b = str(seg.get("joint_b") or "").strip()
            key = self._member_key_text(joint_a, joint_b)
            if not key or key in drawn_member_keys:
                continue

            level = str(self._member_inspect_level_map.get(key, "")).strip().upper()
            if level not in ("II", "III", "IV") or not self._display_inspection_level(level):
                continue

            drawn_member_keys.add(key)

            color = self._inspection_color(level)
            p1 = seg["scene_p1"]
            p2 = seg["scene_p2"]
            tooltip = self._badge_tooltip_for_member(key, level)

            line = RiskMemberItem(p1[0], p1[1], p2[0], p2[1], tooltip=tooltip)
            pen_width = 2.2 if level == "II" else (2.8 if level == "III" else 3.4)
            line.setPen(QPen(color, pen_width))
            line.setZValue(40)
            self._scene.addItem(line)

            mx = (p1[0] + p2[0]) / 2.0
            my = (p1[1] + p2[1]) / 2.0
            self._draw_member_level_inline_label(mx, my, level, color, tooltip=tooltip, p1=p1, p2=p2)

        # ---------- 节点：仅显示当前配置中允许的等级；等级标在箭头上 ----------
        for joint_id, scene_pt in self._visible_projected_nodes.items():
            joint_id = str(joint_id).strip()
            level = str(self._node_inspect_level_map.get(joint_id, "")).strip().upper()
            if level not in ("II", "III", "IV") or not self._display_inspection_level(level):
                continue

            color = self._inspection_color(level)
            x, y = scene_pt
            r = 2.8 if level == "II" else (3.4 if level == "III" else 4.0)
            tooltip = self._badge_tooltip_for_node(joint_id, level)

            self._draw_node_direction_arrows(joint_id, (x, y), level, tooltip)

            dot = RiskNodeItem(QRectF(x - r, y - r, 2 * r, 2 * r), tooltip=tooltip)
            dot.setPen(QPen(color, 1.0))
            dot.setBrush(QBrush(color))
            dot.setZValue(60)
            self._scene.addItem(dot)

    def _draw_level_badge_no_overlap(self, x: float, y: float, level: str, color: QColor):
        text_item = QGraphicsSimpleTextItem(level)
        text_item.setFont(QFont("Arial", 8, QFont.Bold))
        br = text_item.boundingRect()

        # 尝试几个偏移位置
        candidates = [
            (8, -18),
            (8, 6),
            (-24, -18),
            (-24, 6),
            (14, -30),
            (-30, -30),
        ]

        for dx, dy in candidates:
            rect = QRectF(x + dx, y + dy, br.width() + 8, br.height() + 4)

            overlap = False
            for old_rect in self._placed_label_rects:
                if rect.adjusted(-2, -2, 2, 2).intersects(old_rect):
                    overlap = True
                    break

            if overlap:
                continue

            bg = self._scene.addRect(
                rect,
                QPen(color, 1.0),
                QBrush(color),
            )
            bg.setZValue(49)

            text_item = QGraphicsSimpleTextItem(level)
            text_item.setFont(QFont("Arial", 8, QFont.Bold))
            text_item.setBrush(QBrush(QColor("#ffffff")))
            text_item.setPos(rect.x() + 4, rect.y() + 2)
            text_item.setZValue(50)
            self._scene.addItem(text_item)

            self._placed_label_rects.append(rect)
            return

    def _draw_level_badge(self, x: float, y: float, level: str, color: QColor, dy: float = -14.0):
        text_item = QGraphicsSimpleTextItem(level)
        text_item.setFont(QFont("Arial", 8, QFont.Bold))
        text_item.setBrush(QBrush(QColor("#ffffff")))

        br = text_item.boundingRect()
        bg = self._scene.addRect(
            x + 2,
            y + dy,
            br.width() + 8,
            br.height() + 4,
            QPen(color, 1.0),
            QBrush(color),
        )
        bg.setZValue(49)

        text_item.setPos(x + 6, y + dy + 2)
        text_item.setZValue(50)
        self._scene.addItem(text_item)

    def set_inspection_overlay(self, overlay: dict | None):
        overlay = overlay or {}
        self._inspection_overlay_enabled = True
        self._member_inspect_level_map = dict(overlay.get("member_level_by_key") or {})
        self._node_inspect_level_map = dict(overlay.get("node_level_by_joint") or {})
        self._node_brace_inspect_level_map = dict(overlay.get("node_level_by_joint_brace") or {})

        # 新增：保存明细，后面减少拥挤也要用
        self._member_items_by_key = dict(overlay.get("member_items_by_key") or {})
        self._node_items_by_joint = dict(overlay.get("node_items_by_joint") or {})

        # 关键：有模型数据时，立刻重画
        if self.nodes and self.members:
            self._render_row_elevation()

    def set_history_overlay(self, overlay: dict | None):
        overlay = overlay or {}
        self._history_overlay_enabled = True
        self._history_overlay_items = list(overlay.get("items") or [])
        self._history_overlay_legend = list(overlay.get("legend") or [])
        if self.nodes and self.members:
            self._render_row_elevation()

    def _detect_leg_joint_nodes(self) -> Dict[str, Tuple[float, float, float]]:
        if not self.nodes or not self.members:
            return {}

        node_to_max_od = {nid: 0.0 for nid in self.nodes}
        for na, nb, gid in self.members:
            if na in self.nodes and nb in self.nodes:
                od = float(self._groups_od.get(gid, 0.0))
                node_to_max_od[na] = max(node_to_max_od[na], od)
                node_to_max_od[nb] = max(node_to_max_od[nb], od)

        tolerance = 1.0
        elevation_nodes = {
            nid: self.nodes[nid]
            for nid in self.nodes
            if abs(float(self.nodes[nid][2]) - float(self._workpoint_z)) < tolerance
        }
        if not elevation_nodes:
            return {}

        local_max_od = max(node_to_max_od[nid] for nid in elevation_nodes)
        return {
            nid: self.nodes[nid]
            for nid in elevation_nodes
            if node_to_max_od[nid] >= local_max_od * 0.95
        }

    def _get_leg_plane_clusters(self) -> tuple[list[float], list[float]]:
        leg_nodes = self._detect_leg_joint_nodes()

        # 优先用主腿；主腿识别不到时再退回 workpoint 以下节点
        base_nodes = leg_nodes if leg_nodes else self._nodes_below_workpoint()
        return self._get_axis_clusters(base_nodes)

    def _score_model_candidate(self, path: str, facility_code: str) -> int:
        name = os.path.basename(path).lower()
        path_low = path.lower()
        score = 0
        if name.startswith("sacinp"):
            score += 300
        if name.endswith(".sacinp"):
            score += 220
        if "当前模型" in path:
            score += 80
        if "结构模型" in path:
            score += 60
        if facility_code and facility_code.lower() in path_low:
            score += 200
        if "user" in path_low or "用户上传" in path_low:
            score += 20
        return score

    # ---------------- 模型解析 ----------------
    def _read_lines_with_fallback(self, file_path: str) -> List[str]:
        encodings = ["utf-8", "utf-8-sig", "gb18030", "gbk", "latin-1"]
        for enc in encodings:
            try:
                with open(file_path, "r", encoding=enc) as f:
                    return f.readlines()
            except UnicodeDecodeError:
                continue
            except Exception:
                break
        with open(file_path, "r", encoding="utf-8", errors="ignore") as f:
            return f.readlines()

    def parse_sacs_full_robust(self, filepath):
        nodes = {}
        members = []
        groups_od = {}

        lines = self._read_lines_with_fallback(filepath)
        for line in lines:
            group = _parse_sacs_group_od(line)
            if group is not None:
                gid, od = group
                groups_od[gid] = od
                continue

            joint = _parse_sacs_joint_line(line)
            if joint is not None:
                nid, x, y, z = joint
                nodes[nid] = (x, y, z)
                continue

            member = _parse_sacs_member_line(line)
            if member is not None:
                members.append(member)

        return nodes, members, groups_od

    def _project_root(self) -> str:
        return os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

    def _special_strategy_inputs_dir(self) -> str:
        return os.path.join(self._project_root(), "special_strategy_inputs")

    def _resolve_risk_workbook_path(self) -> str:
        root = self._project_root()
        strategy_root = os.path.join(root, "pages", "output_special_strategy")
        candidates = [
            os.path.join(self._special_strategy_inputs_dir(), "special_strategy_template.xlsm"),
            os.path.join(strategy_root, "special_strategy_template.xlsm"),
            os.path.join(root, "special_strategy_template.xlsm"),
        ]
        for p in candidates:
            if os.path.exists(p):
                return p
        return ""

    def _normalize_year_label_for_sheet(self, year_label: str) -> str:
        if year_label == "当前":
            return "当前"
        mapping = {
            "+5年": "第5年",
            "+10年": "第10年",
            "+15年": "第15年",
            "+20年": "第20年",
            "+25年": "第25年",
        }
        return mapping.get(year_label, year_label)

    def _read_sheet_rows(self, workbook_path: str, sheet_name: str, header_row: int, data_start_row: int) -> List[Dict]:
        if not workbook_path or not os.path.exists(workbook_path):
            return []

        import warnings
        with warnings.catch_warnings():
            warnings.filterwarnings("ignore", message="Data Validation extension is not supported and will be removed")
            wb = openpyxl.load_workbook(workbook_path, data_only=True, read_only=True)

        if sheet_name not in wb.sheetnames:
            return []

        ws = wb[sheet_name]

        header_values = next(ws.iter_rows(min_row=header_row, max_row=header_row, values_only=True))
        headers = [("" if v is None else str(v).strip()) for v in header_values]

        rows: List[Dict] = []
        for values in ws.iter_rows(min_row=data_start_row, values_only=True):
            if not any(v not in (None, "", " ") for v in values):
                continue

            row = {}
            for idx, h in enumerate(headers):
                if not h:
                    continue
                row[h] = values[idx] if idx < len(values) else None
            rows.append(row)

        return rows

    def _normalize_existing_path(self, path: str) -> str:
        p = os.path.normpath(str(path or "").strip())
        if not p:
            return ""
        if os.path.exists(p):
            return p

        roots = [
            "",
            external_path(""),
            first_existing_path("upload"),
            first_existing_path("upload", "model_files"),
            self._special_strategy_inputs_dir(),
            r"Y:\shiyou_file_storage",
            r"Y:\upload",
        ]
        for root in roots:
            if not root:
                continue
            candidate = os.path.normpath(os.path.join(root, p))
            if os.path.exists(candidate):
                return candidate
        return ""

    def _collect_candidate_paths(self, context: Dict, keys: List[str]) -> List[str]:
        paths: List[str] = []
        for key in keys:
            value = context.get(key)
            if not value:
                continue

            if isinstance(value, str):
                paths.append(value)
            elif isinstance(value, dict):
                for kk in ("storage_path", "path", "file_path", "absolute_path"):
                    if value.get(kk):
                        paths.append(str(value[kk]))
            elif isinstance(value, list):
                for item in value:
                    if isinstance(item, str):
                        paths.append(item)
                    elif isinstance(item, dict):
                        for kk in ("storage_path", "path", "file_path", "absolute_path"):
                            if item.get(kk):
                                paths.append(str(item[kk]))

        out: List[str] = []
        seen = set()
        for p in paths:
            real = self._normalize_existing_path(p)
            if real and real not in seen:
                seen.add(real)
                out.append(real)
        return out

    def _read_rows_from_table_file(self, path: str, preferred_sheet_keywords: List[str] | None = None) -> List[Dict]:
        if not path or not os.path.exists(path):
            return []

        ext = os.path.splitext(path)[1].lower()

        try:
            if ext == ".csv":
                with open(path, "r", encoding="utf-8-sig", newline="") as f:
                    return list(csv.DictReader(f))

            if ext == ".json":
                with open(path, "r", encoding="utf-8") as f:
                    obj = json.load(f)
                if isinstance(obj, list):
                    return [x for x in obj if isinstance(x, dict)]
                if isinstance(obj, dict):
                    for v in obj.values():
                        if isinstance(v, list) and v and isinstance(v[0], dict):
                            return v
                return []

            if ext in {".xlsx", ".xlsm", ".xls"} and pd is not None:
                sheets = pd.read_excel(path, sheet_name=None)
                rows: List[Dict] = []

                def _sheet_score(name: str) -> int:
                    if not preferred_sheet_keywords:
                        return 0
                    low = str(name).lower()
                    score = 0
                    for kw in preferred_sheet_keywords:
                        if kw.lower() in low:
                            score += 10
                    return score

                ordered = sorted(sheets.items(), key=lambda kv: _sheet_score(kv[0]), reverse=True)
                for sheet_name, df in ordered:
                    if df is None or df.empty:
                        continue
                    clean_df = df.where(pd.notnull(df), None)
                    records = clean_df.to_dict(orient="records")
                    if records:
                        rows.extend(records)
                        if preferred_sheet_keywords and _sheet_score(sheet_name) > 0:
                            return rows
                return rows
        except Exception as exc:
            print("[Elevation] read risk table failed:", path, exc)

        return []

    def _extract_rows_from_context_or_files(
            self,
            context: Dict,
            direct_keys: List[str],
            path_keys: List[str],
            preferred_sheet_keywords: List[str],
    ) -> List[Dict]:
        for key in direct_keys:
            value = context.get(key)
            if isinstance(value, list) and value and isinstance(value[0], dict):
                return value

        paths = self._collect_candidate_paths(context, path_keys)
        for path in paths:
            rows = self._read_rows_from_table_file(path, preferred_sheet_keywords=preferred_sheet_keywords)
            if rows:
                print("[Elevation] loaded risk rows from file:", path, "count=", len(rows))
                return rows

        return []

    def _risk_text_to_num(self, value: object) -> str:
        raw = str(value or "").strip()
        mapping = {
            "一": "1", "二": "2", "三": "3", "四": "4", "五": "5",
            "Ⅰ": "1", "Ⅱ": "2", "Ⅲ": "3", "Ⅳ": "4", "Ⅴ": "5",
            "1": "1", "2": "2", "3": "3", "4": "4", "5": "5",
        }
        return mapping.get(raw, raw)

    def _match_time_node(self, row: Dict, year_label: str) -> bool:
        time_node = str(row.get("time_node") or row.get("year") or row.get("年份") or "").strip()
        if not time_node:
            return year_label == "当前"

        if year_label == "当前":
            return time_node in {"当前", "current", "Current"}

        if year_label.startswith("+") and year_label.endswith("年"):
            n = year_label[1:-1]
            return time_node in {f"第{n}年", year_label}

        return False

    # 下面两个函数保留，后续重新打开风险标注时可直接恢复使用
    def _extract_node_risks_from_context(self, context: Dict, year_label: str) -> Dict[str, str]:
        result: Dict[str, str] = {}

        workbook_path = self._resolve_risk_workbook_path()
        print("[Elevation] risk workbook =", workbook_path)

        rows = self._read_sheet_rows(
            workbook_path=workbook_path,
            sheet_name="节点检验策略",
            header_row=2,
            data_start_row=3,
        )

        target_year = self._normalize_year_label_for_sheet(year_label)
        print("[Elevation] node target_year =", target_year)
        print("[Elevation] node sheet rows =", len(rows))

        matched_rows = 0
        for row in rows:
            joint_id = str(row.get("JoitID") or "").strip()
            if not joint_id:
                continue

            time_node = str(row.get("检验时间节点") or "").strip()
            if time_node != target_year:
                continue

            matched_rows += 1

            risk = self._risk_text_to_num(row.get("节点风险等级"))
            if risk:
                result[joint_id] = risk

        print("[Elevation] node matched_rows =", matched_rows)
        print("[Elevation] node risk_map size =", len(result))
        return result

    def _extract_member_risks_from_context(self, context: Dict, year_label: str) -> Dict[Tuple[str, str], str]:
        result: Dict[Tuple[str, str], str] = {}

        workbook_path = self._resolve_risk_workbook_path()
        print("[Elevation] risk workbook =", workbook_path)

        rows = self._read_sheet_rows(
            workbook_path=workbook_path,
            sheet_name="构件检验策略",
            header_row=1,
            data_start_row=2,
        )

        target_year = self._normalize_year_label_for_sheet(year_label)
        print("[Elevation] member target_year =", target_year)
        print("[Elevation] member sheet rows =", len(rows))

        matched_rows = 0
        for row in rows:
            a = str(row.get("Joint A") or "").strip()
            b = str(row.get("Joint B") or "").strip()
            if not a or not b:
                continue

            time_node = str(row.get("检验时间节点") or "").strip()
            if time_node != target_year:
                continue

            matched_rows += 1

            risk = self._risk_text_to_num(row.get("构件风险等级"))
            if risk:
                result[tuple(sorted((a, b)))] = risk

        print("[Elevation] member matched_rows =", matched_rows)
        print("[Elevation] member risk_map size =", len(result))
        return result

    # ---------------- ROW 面识别 ----------------
    def _cluster_axis_values(self, values: List[float], tol: float) -> List[float]:
        if not values:
            return []
        vals = sorted(values)
        groups = [[vals[0]]]
        for v in vals[1:]:
            if abs(v - groups[-1][-1]) <= tol:
                groups[-1].append(v)
            else:
                groups.append([v])
        return [sum(g) / len(g) for g in groups]

    def _get_axis_clusters(self, node_map: Optional[Dict[str, Tuple[float, float, float]]] = None):
        src = node_map if node_map is not None else self.nodes

        xs = [coord[0] for coord in src.values()]
        ys = [coord[1] for coord in src.values()]
        if not xs or not ys:
            return [], []

        x_span = max(xs) - min(xs)
        y_span = max(ys) - min(ys)

        x_tol = max(0.8, x_span * 0.03)
        y_tol = max(0.8, y_span * 0.03)

        x_clusters = self._cluster_axis_values(xs, x_tol)
        y_clusters = self._cluster_axis_values(ys, y_tol)

        return x_clusters, y_clusters

    def _nodes_below_workpoint(self) -> Dict[str, Tuple[float, float, float]]:
        return {
            nid: coord
            for nid, coord in self.nodes.items()
            if float(coord[2]) <= self._workpoint_z + 1e-6
        }

    def _safe_float(self, value: object, default: float = 9.1) -> float:
        try:
            return float(str(value).strip())
        except Exception:
            return default

    def _extract_workpoint_from_context(self, context: Optional[Dict]) -> float:
        direct_keys = [
            "workpoint", "target_z", "targetZ", "work_point",
            "Workpoint", "工作平面高程", "工作平面高程Workpoint",
        ]

        if isinstance(context, dict):
            for key in direct_keys:
                if key in context and context.get(key) not in ("", None):
                    return self._safe_float(context.get(key), 9.1)

            for value in context.values():
                if isinstance(value, dict):
                    for key in direct_keys:
                        if key in value and value.get(key) not in ("", None):
                            return self._safe_float(value.get(key), 9.1)

        return 9.1

    def _extract_level_threshold_from_context(self, context: Optional[Dict]) -> int:
        direct_keys = [
            "level_threshold",
            "node_limit",
            "node_count_limit",
            "水平层高程节点数量限制",
        ]

        if isinstance(context, dict):
            for key in direct_keys:
                if key in context and context.get(key) not in ("", None):
                    try:
                        return int(float(str(context.get(key)).strip()))
                    except Exception:
                        pass

            for value in context.values():
                if isinstance(value, dict):
                    for key in direct_keys:
                        if key in value and value.get(key) not in ("", None):
                            try:
                                return int(float(str(value.get(key)).strip()))
                            except Exception:
                                pass

        return 40

    def _row_index_to_letters(self, idx: int) -> str:
        # 0 -> A, 1 -> B, ..., 25 -> Z, 26 -> AA
        out = ""
        n = idx
        while True:
            out = chr(ord("A") + (n % 26)) + out
            n = n // 26 - 1
            if n < 0:
                break
        return out

    def _row_letters_to_index(self, text: str) -> int:
        s = (text or "").strip().upper()
        if not s:
            return 0

        value = 0
        for ch in s:
            if not ("A" <= ch <= "Z"):
                return 0
            value = value * 26 + (ord(ch) - ord("A") + 1)
        return max(0, value - 1)

    def _get_horizontal_z_clusters(self) -> List[float]:
        zs = [float(coord[2]) for coord in self.nodes.values() if float(coord[2]) <= self._workpoint_z + 1e-6]
        if not zs:
            return []

        z_span = max(zs) - min(zs)
        z_tol = max(0.5, z_span * 0.01)

        raw_clusters = self._cluster_axis_values(zs, z_tol)
        if not raw_clusters:
            return []

        selected = []
        for c in raw_clusters:
            cnt = sum(1 for z in zs if abs(z - c) <= z_tol)
            if cnt >= self._level_threshold:
                selected.append(c)

        selected.sort(reverse=True)
        return selected

    def _build_dynamic_row_specs(self, node_map: Optional[Dict[str, Tuple[float, float, float]]] = None):
        z_clusters = self._get_horizontal_z_clusters()

        specs = [
            {"name": "XZ 前", "plane_axis": None, "plane_value": None, "proj_mode": "XZ_FRONT"},
            {"name": "XZ 后", "plane_axis": None, "plane_value": None, "proj_mode": "XZ_BACK"},
            {"name": "YZ 左", "plane_axis": None, "plane_value": None, "proj_mode": "YZ_LEFT"},
            {"name": "YZ 右", "plane_axis": None, "plane_value": None, "proj_mode": "YZ_RIGHT"},
        ]

        for zv in z_clusters:
            z_label = f"{zv:.3f}".rstrip("0").rstrip(".")
            specs.append({
                "name": f"XY {z_label}",
                "plane_axis": "Z",
                "plane_value": zv,
                "proj_mode": "XY",
            })

        return specs

    def _select_face_clusters(self, clusters: List[float], side: str) -> List[float]:
        vals = sorted(float(v) for v in clusters)
        if not vals:
            return []

        if side in ("front", "left"):
            return [vals[0]]

        if side in ("back", "right"):
            return [vals[-1]]

        return [vals[0]]

    def _labels_from_segments_by_projection(self, segments, proj_mode: str):
        if not segments:
            return []

        xs = []
        for seg in segments:
            p1, p2 = self._segment_points(seg)
            xs.extend([p1[0], p2[0]])

        if not xs:
            return []

        span = max(xs) - min(xs) if len(xs) >= 2 else 0.0
        tol = max(0.8, span * 0.03)
        clusters = self._cluster_axis_values(xs, tol)

        if proj_mode in ("XZ_FRONT", "XZ_BACK"):
            return [(v, str(i + 1)) for i, v in enumerate(sorted(clusters))]
        else:
            return [(v, self._row_index_to_letters(i)) for i, v in enumerate(sorted(clusters))]

    def _get_xz_upper_zone_min_z(self) -> float:
        """
        前后立面中，顶部/上部甲板区的最低 z。
        取前 5 个水平高程层，补全前后图顶部。
        """
        z_levels = self._get_horizontal_z_clusters()  # 已按从高到低排序
        if len(z_levels) >= 5:
            return float(z_levels[4]) - 0.5
        if len(z_levels) >= 4:
            return float(z_levels[3]) - 0.5
        if z_levels:
            return float(z_levels[-1]) - 0.5
        return float(self._workpoint_z - 20.0)

    def _merge_unique_segments(self, segs_a, segs_b):
        out = []
        seen = set()

        for seg in list(segs_a) + list(segs_b):
            p1, p2 = self._segment_points(seg)
            a = (round(float(p1[0]), 3), round(float(p1[1]), 3))
            b = (round(float(p2[0]), 3), round(float(p2[1]), 3))
            key = tuple(sorted((a, b)))
            if key in seen:
                continue
            seen.add(key)
            out.append(seg)
        return out

    def _collect_side_face_segments(self, proj_mode: str):
        """
        收集四个方向立面图的投影线段：
        - 当前面 -> is_face=True（黑线）
        - 其他杂线 -> is_face=False（灰线）
        - 不再只保留 outer_cluster，避免底部和腿部缺线
        """
        if proj_mode not in ("XZ_FRONT", "XZ_BACK", "YZ_LEFT", "YZ_RIGHT"):
            return [], [], []

        face_value, face_eps = self._direction_face_params(proj_mode)

        seg_map = {}
        point_map = {}

        for na, nb, _gid in self.members:
            if na not in self.nodes or nb not in self.nodes:
                continue

            clipped = self._clip_member_3d_to_workpoint(self.nodes[na], self.nodes[nb])
            if clipped is None:
                continue

            p1, p2 = clipped

            # 是否属于当前面
            is_face = self._member_hits_direction_face_band(
                p1, p2, face_value, face_eps, proj_mode
            )

            # 前后图额外把外腿相关构件提成当前面
            if proj_mode in ("XZ_FRONT", "XZ_BACK"):
                if self._member_hits_xz_outer_leg_band(p1, p2):
                    is_face = True

            # 立面投影
            if proj_mode in ("XZ_FRONT", "XZ_BACK"):
                seg2d_p1 = (float(p1[0]), float(p1[2]))  # XZ
                seg2d_p2 = (float(p2[0]), float(p2[2]))
            else:
                seg2d_p1 = (float(p1[1]), float(p1[2]))  # YZ
                seg2d_p2 = (float(p2[1]), float(p2[2]))

            # 投影后变成点
            if (
                    abs(seg2d_p1[0] - seg2d_p2[0]) < 1e-6
                    and abs(seg2d_p1[1] - seg2d_p2[1]) < 1e-6
            ):
                key = (round(seg2d_p1[0], 3), round(seg2d_p1[1], 3))
                old = point_map.get(key)
                if old is None or (is_face and not old["is_face"]):
                    point_map[key] = {
                        "pt": seg2d_p1,
                        "is_face": is_face,
                    }
                continue

            a = (round(seg2d_p1[0], 3), round(seg2d_p1[1], 3))
            b = (round(seg2d_p2[0], 3), round(seg2d_p2[1], 3))
            key = tuple(sorted((a, b)))

            old = seg_map.get(key)
            item = {
                "joint_a": str(na).strip(),
                "joint_b": str(nb).strip(),
                "p1": seg2d_p1,
                "p2": seg2d_p2,
                "is_face": is_face,
            }

            # 如果同一条投影线重复出现，优先保留黑线版本
            if old is None or (is_face and not old.get("is_face", False)):
                seg_map[key] = item

        segments = list(seg_map.values())
        point_markers = list(point_map.values())
        top_labels = self._labels_from_segments_by_projection(segments, proj_mode)

        return segments, top_labels, point_markers

    def _collect_xz_upper_face_segments(self, proj_mode: str):
        """
        前/后视图上部甲板补线：
        只补“当前立面主面带”附近的上部构件
        """
        z_min = self._get_xz_upper_zone_min_z()
        face_value, face_eps = self._direction_face_params(proj_mode)

        candidates = []

        for na, nb, _gid in self.members:
            if na not in self.nodes or nb not in self.nodes:
                continue

            clipped = self._clip_member_3d_to_workpoint(self.nodes[na], self.nodes[nb])
            if clipped is None:
                continue

            p1, p2 = clipped
            zmax = max(float(p1[2]), float(p2[2]))
            if zmax < z_min:
                continue

            if not self._member_hits_direction_face_band(
                    p1, p2, face_value, face_eps, proj_mode
            ):
                continue

            seg2d_p1 = (float(p1[0]), float(p1[2]))
            seg2d_p2 = (float(p2[0]), float(p2[2]))

            if (
                    abs(seg2d_p1[0] - seg2d_p2[0]) < 1e-6
                    and abs(seg2d_p1[1] - seg2d_p2[1]) < 1e-6
            ):
                continue

            candidates.append({
                "joint_a": na,
                "joint_b": nb,
                "p1": seg2d_p1,
                "p2": seg2d_p2,
            })

        return self._dedupe_projected_segments(candidates)

    def _collect_xz_upper_point_markers(self, proj_mode: str):
        """
        前/后视图里，上部退化成点的构件位置补点
        """
        z_min = self._get_xz_upper_zone_min_z()
        face_value, face_eps = self._direction_face_params(proj_mode)

        candidates = []

        for na, nb, _gid in self.members:
            if na not in self.nodes or nb not in self.nodes:
                continue

            clipped = self._clip_member_3d_to_workpoint(self.nodes[na], self.nodes[nb])
            if clipped is None:
                continue

            p1, p2 = clipped
            zmax = max(float(p1[2]), float(p2[2]))
            if zmax < z_min:
                continue

            if not self._member_hits_direction_face_band(
                    p1, p2, face_value, face_eps, proj_mode
            ):
                continue

            x1, z1 = float(p1[0]), float(p1[2])
            x2, z2 = float(p2[0]), float(p2[2])

            # 只保留投影为点的构件
            if abs(x1 - x2) > 1e-6 or abs(z1 - z2) > 1e-6:
                continue

            candidates.append((x1, z1))

        # 去重点
        out = []
        seen = set()
        for x, z in candidates:
            key = (round(x, 3), round(z, 3))
            if key in seen:
                continue
            seen.add(key)
            out.append((x, z))

        return out

    def _labels_from_face_segments(self, segments, proj_mode: str):
        if not segments:
            return []

        vals = []
        for seg in segments:
            p1, p2 = self._segment_points(seg)
            vals.extend([p1[0], p2[0]])

        if not vals:
            return []

        span = max(vals) - min(vals) if len(vals) >= 2 else 0.0
        tol = max(0.8, span * 0.03)
        used_clusters = self._cluster_axis_values(vals, tol)

        if proj_mode in ("XZ_FRONT", "XZ_BACK"):
            return [(v, str(i + 1)) for i, v in enumerate(sorted(used_clusters))]
        else:
            return [(v, self._row_index_to_letters(i)) for i, v in enumerate(sorted(used_clusters))]

    def _clip_member_3d_to_workpoint(
            self,
            p1: Tuple[float, float, float],
            p2: Tuple[float, float, float],
    ) -> Optional[Tuple[Tuple[float, float, float], Tuple[float, float, float]]]:
        wp = float(self._workpoint_z)

        x1, y1, z1 = p1
        x2, y2, z2 = p2

        # 整根都在 workpoint 上方：不画
        if z1 > wp and z2 > wp:
            return None

        # 整根都在 workpoint 以下：直接保留
        if z1 <= wp and z2 <= wp:
            return p1, p2

        # 穿过 workpoint：裁到 workpoint 平面
        dz = z2 - z1
        if abs(dz) < 1e-12:
            return None

        t = (wp - z1) / dz
        t = max(0.0, min(1.0, t))
        cut = (
            x1 + t * (x2 - x1),
            y1 + t * (y2 - y1),
            wp,
        )

        if z1 > wp:
            return cut, p2
        return p1, cut

    def _segment_sample_bins(
            self,
            p1: Tuple[float, float],
            p2: Tuple[float, float],
            grid: float = 1.5,
            samples: int = 10,
    ) -> set:
        bins = set()
        for i in range(samples + 1):
            t = i / samples
            x = p1[0] + t * (p2[0] - p1[0])
            y = p1[1] + t * (p2[1] - p1[1])
            bins.add((round(x / grid), round(y / grid)))
        return bins

    def _median_value(self, values: List[float]) -> float:
        vals = sorted(float(v) for v in values if v is not None)
        if not vals:
            return 0.0
        n = len(vals)
        mid = n // 2
        if n % 2 == 1:
            return vals[mid]
        return (vals[mid - 1] + vals[mid]) * 0.5

    def _direction_split_params(self, proj_mode: str) -> Tuple[float, float]:
        """
        对方向立面求“前后/左右”分界中值和容差。
        XZ 前后图：按 y 分
        YZ 左右图：按 x 分
        优先使用 workpoint 以下节点，避免上部极端点干扰。
        """
        use_y = proj_mode in ("XZ_FRONT", "XZ_BACK")
        wp = float(self._workpoint_z)

        vals: List[float] = []
        for _nid, (x, y, z) in self.nodes.items():
            if float(z) <= wp + 1e-6:
                vals.append(float(y) if use_y else float(x))

        # 如果 workpoint 以下一个都没有，再退回全节点
        if not vals:
            for _nid, (x, y, z) in self.nodes.items():
                vals.append(float(y) if use_y else float(x))

        if not vals:
            return 0.0, 0.5

        mid = self._median_value(vals)
        span = (max(vals) - min(vals)) if len(vals) >= 2 else 0.0

        # 容差带：防止恰好卡在中间的斜杆被过度裁掉
        eps = max(span * 0.03, 0.5)
        return mid, eps

    def _member_rep_axis_value(
            self,
            p1: Tuple[float, float, float],
            p2: Tuple[float, float, float],
            proj_mode: str,
    ) -> float:
        """
        构件代表深度：
        XZ 视图用 y 中心值
        YZ 视图用 x 中心值
        """
        if proj_mode in ("XZ_FRONT", "XZ_BACK"):
            return (float(p1[1]) + float(p2[1])) * 0.5
        return (float(p1[0]) + float(p2[0])) * 0.5

    def _member_pass_direction_half(
            self,
            axis_value: float,
            split_mid: float,
            eps: float,
            proj_mode: str,
    ) -> bool:
        """
        方向立面的半区过滤：
        - XZ_FRONT：前半区
        - XZ_BACK ：后半区
        - YZ_LEFT ：左半区
        - YZ_RIGHT：右半区
        """
        if proj_mode == "XZ_FRONT":
            return axis_value <= split_mid + eps
        if proj_mode == "XZ_BACK":
            return axis_value >= split_mid - eps
        if proj_mode == "YZ_LEFT":
            return axis_value <= split_mid + eps
        if proj_mode == "YZ_RIGHT":
            return axis_value >= split_mid - eps
        return True

    def _direction_face_params(self, proj_mode: str):
        """
        当前方向立面的主面位置与容差带
        XZ 前后图 -> 按 y 分前后
        YZ 左右图 -> 按 x 分左右
        """
        leg_x_clusters, leg_y_clusters = self._get_leg_plane_clusters()

        if proj_mode in ("XZ_FRONT", "XZ_BACK"):
            vals = sorted(float(v) for v in leg_y_clusters)
        else:
            vals = sorted(float(v) for v in leg_x_clusters)

        if not vals:
            return 0.0, 3.0

        face_value = vals[0] if proj_mode in ("XZ_FRONT", "YZ_LEFT") else vals[-1]
        span = vals[-1] - vals[0] if len(vals) >= 2 else 0.0

        # 容差要稍微放宽，否则还是会缺线
        band_eps = max(span * 0.22, 3.0)
        return face_value, band_eps

    def _member_hits_direction_face_band(
            self,
            p1,
            p2,
            face_value: float,
            eps: float,
            proj_mode: str,
    ) -> bool:
        """
        判断一根 3D 构件是否属于当前方向的立面带
        - XZ_FRONT / XZ_BACK：按 y 判断
        - YZ_LEFT / YZ_RIGHT：按 x 判断

        规则：
        1）任一端点落入面带 -> 算当前面
        2）线段在该轴上的范围与面带相交 -> 也算当前面
        """
        if proj_mode in ("XZ_FRONT", "XZ_BACK"):
            a1, a2 = float(p1[1]), float(p2[1])  # 看 y
        else:
            a1, a2 = float(p1[0]), float(p2[0])  # 看 x

        lo = min(a1, a2)
        hi = max(a1, a2)

        band_lo = face_value - eps
        band_hi = face_value + eps

        if band_lo <= a1 <= band_hi or band_lo <= a2 <= band_hi:
            return True

        if not (hi < band_lo or lo > band_hi):
            return True

        return False

    def _member_hits_xz_outer_leg_band(self, p1, p2) -> bool:
        """
        前后图里，把左右两条外腿附近的杆件也提升为当前面（黑线）。
        这样前/后视图的腿不会缺失。
        """
        leg_x_clusters, _ = self._get_leg_plane_clusters()
        xs = sorted(float(v) for v in leg_x_clusters)

        if len(xs) < 2:
            return False

        x_left = xs[0]
        x_right = xs[-1]
        span = x_right - x_left if len(xs) >= 2 else 0.0
        x_eps = max(span * 0.05, 1.2)

        x1, x2 = float(p1[0]), float(p2[0])
        lo = min(x1, x2)
        hi = max(x1, x2)

        def hit_band(center):
            band_lo = center - x_eps
            band_hi = center + x_eps
            return not (hi < band_lo or lo > band_hi)

        return hit_band(x_left) or hit_band(x_right)
    def _dedupe_projected_segments_keep_face(self, segments):
        kept = {}

        for seg in segments:
            p1 = seg["p1"]
            p2 = seg["p2"]

            a = (round(float(p1[0]), 3), round(float(p1[1]), 3))
            b = (round(float(p2[0]), 3), round(float(p2[1]), 3))
            key = tuple(sorted((a, b)))

            prev = kept.get(key)
            if prev is None:
                kept[key] = seg
                continue

            # 黑线优先
            if seg.get("is_face", False) and not prev.get("is_face", False):
                kept[key] = seg

        return list(kept.values())

    def _dedupe_point_markers_keep_face(self, markers):
        kept = {}

        for mk in markers:
            px, py = mk["pt"]
            key = (round(float(px), 3), round(float(py), 3))

            prev = kept.get(key)
            if prev is None:
                kept[key] = mk
                continue

            if mk.get("is_face", False) and not prev.get("is_face", False):
                kept[key] = mk

        return list(kept.values())

    def _dedupe_projected_segments(self, segs: List[Dict]) -> List[Dict]:
        out = []
        seen = set()

        for seg in segs:
            a = (round(seg["p1"][0], 3), round(seg["p1"][1], 3))
            b = (round(seg["p2"][0], 3), round(seg["p2"][1], 3))
            key = tuple(sorted((a, b)))
            if key in seen:
                continue
            seen.add(key)
            out.append(seg)

        return out

    def _member_pass_direction_face(
            self,
            p1: Tuple[float, float, float],
            p2: Tuple[float, float, float],
            face_value: float,
            eps: float,
            proj_mode: str,
    ) -> bool:
        """
        只保留真正贴近该立面主面的构件。
        这里用“两个端点都靠近主面”这个更严格的条件，
        杂线会明显减少。
        """
        if proj_mode in ("XZ_FRONT", "XZ_BACK"):
            a1, a2 = float(p1[1]), float(p2[1])  # 看 y
        else:
            a1, a2 = float(p1[0]), float(p2[0])  # 看 x

        return (abs(a1 - face_value) <= eps) and (abs(a2 - face_value) <= eps)

    def _top_labels_for_directional_view(self, proj_mode: str):
        leg_x_clusters, leg_y_clusters = self._get_leg_plane_clusters()

        if proj_mode in ("XZ_FRONT", "XZ_BACK"):
            return [(v, str(i + 1)) for i, v in enumerate(sorted(leg_x_clusters))]

        if proj_mode in ("YZ_LEFT", "YZ_RIGHT"):
            return [(v, self._row_index_to_letters(i)) for i, v in enumerate(sorted(leg_y_clusters))]

        return []

    def _collect_directional_view_segments(
            self,
            proj_mode: str,
            z_min: Optional[float] = None,
            keep_all: bool = False,
    ):
        """
        四个方向立面图的主构件收集逻辑：

        新原则：
        1) 不再做全图 bins winner 竞争
        2) 先按“当前立面主面带”选构件
        3) 再做投影
        4) 最后只做完全重复线段去重
        """
        candidates = []

        face_value, face_eps = self._direction_face_params(proj_mode)

        for na, nb, _gid in self.members:
            if na not in self.nodes or nb not in self.nodes:
                continue

            raw1 = self.nodes[na]
            raw2 = self.nodes[nb]

            clipped_3d = self._clip_member_3d_to_workpoint(raw1, raw2)
            if clipped_3d is None:
                continue

            p1, p2 = clipped_3d

            zmax = max(float(p1[2]), float(p2[2]))
            if z_min is not None and zmax < float(z_min):
                continue

            if not keep_all:
                if not self._member_hits_direction_face_band(
                        p1, p2, face_value, face_eps, proj_mode
                ):
                    continue

            if proj_mode in ("XZ_FRONT", "XZ_BACK"):
                seg2d_p1 = (float(p1[0]), float(p1[2]))
                seg2d_p2 = (float(p2[0]), float(p2[2]))
            elif proj_mode in ("YZ_LEFT", "YZ_RIGHT"):
                seg2d_p1 = (float(p1[1]), float(p1[2]))
                seg2d_p2 = (float(p2[1]), float(p2[2]))
            else:
                return [], self._top_labels_for_directional_view(proj_mode)

            # 零长度线段不保留
            if (
                    abs(seg2d_p1[0] - seg2d_p2[0]) < 1e-6
                    and abs(seg2d_p1[1] - seg2d_p2[1]) < 1e-6
            ):
                continue

            candidates.append({
                "joint_a": str(na).strip(),
                "joint_b": str(nb).strip(),
                "p1": seg2d_p1,
                "p2": seg2d_p2,
            })

        candidates = self._dedupe_projected_segments(candidates)
        return candidates, self._top_labels_for_directional_view(proj_mode)

    def available_row_names(self) -> List[str]:
        if not self.nodes:
            return ["XZ 前"]
        specs = self._build_dynamic_row_specs()
        return [item["name"] for item in specs] if specs else ["XZ 前"]

    def _clip_projected_member_to_workpoint(
            self,
            p1: Tuple[float, float],
            p2: Tuple[float, float],
    ) -> Optional[Tuple[Tuple[float, float], Tuple[float, float]]]:
        wp = float(self._workpoint_z)

        x1, z1 = p1
        x2, z2 = p2

        # 整根都在 workpoint 上方：不画
        if z1 > wp and z2 > wp:
            return None

        # 整根都在 workpoint 以下：直接保留
        if z1 <= wp and z2 <= wp:
            return p1, p2

        # 穿过 workpoint：裁到 workpoint 平面
        dz = z2 - z1
        if abs(dz) < 1e-12:
            return None

        t = (wp - z1) / dz
        t = max(0.0, min(1.0, t))
        xc = x1 + t * (x2 - x1)
        cut_point = (xc, wp)

        if z1 > wp:
            return cut_point, p2
        return p1, cut_point

    def _resolve_row_definition(self):
        specs = self._build_dynamic_row_specs()
        if not specs:
            return {
                "name": "XZ 前",
                "plane_axis": "Y",
                "plane_value": 0.0,
                "proj_mode": "XZ",
            }

        target = (self._row_name or "").strip()
        for spec in specs:
            if spec["name"] == target:
                return spec

        return specs[0]

    def _nearest_cluster(self, value: float, clusters: List[float]) -> float:
        if not clusters:
            return value
        return min(clusters, key=lambda c: abs(c - value))

    def _filter_row_members(self):
        spec = self._resolve_row_definition()
        plane_axis = spec["plane_axis"]
        plane_value = spec["plane_value"]
        proj_mode = spec["proj_mode"]

        # XZ/YZ 不再在这里按平面筛
        if proj_mode in ("XZ_FRONT", "XZ_BACK", "YZ_LEFT", "YZ_RIGHT"):
            return {}, [], proj_mode, self._top_labels_for_directional_view(proj_mode)

        # 这里只处理 XY 截面
        z_clusters = self._get_horizontal_z_clusters()
        axis_clusters = z_clusters

        if not axis_clusters:
            return {}, [], proj_mode, []

        target_cluster = self._nearest_cluster(plane_value, axis_clusters)

        selected_nodes = set()
        selected_members = []

        for na, nb, gid in self.members:
            if na not in self.nodes or nb not in self.nodes:
                continue

            xa, ya, za = self.nodes[na]
            xb, yb, zb = self.nodes[nb]

            ca = self._nearest_cluster(za, axis_clusters)
            cb = self._nearest_cluster(zb, axis_clusters)

            if abs(ca - target_cluster) < 1e-6 and abs(cb - target_cluster) < 1e-6:
                if target_cluster > self._workpoint_z + 1e-6:
                    continue
                selected_members.append((na, nb, gid))
                selected_nodes.add(na)
                selected_nodes.add(nb)

        node_map = {nid: self.nodes[nid] for nid in selected_nodes}
        return node_map, selected_members, proj_mode, []

    # ---------------- ROW立面绘制：当前只画轮廓 ----------------
    def _render_row_elevation(self):
        self.setUpdatesEnabled(False)
        self._scene.blockSignals(True)

        try:
            self._scene.clear()
            self._visible_projected_members = []
            self._visible_projected_nodes = {}

            spec = self._resolve_row_definition()
            proj_mode = spec["proj_mode"]

            clipped_segments = []
            projected_nodes = {}
            point_markers = []
            top_labels = []

            directional_modes = ("XZ_FRONT", "XZ_BACK", "YZ_LEFT", "YZ_RIGHT")

            # ---------- 1) 前/后/左/右立面 ----------
            if proj_mode in directional_modes:
                # 关键优化：先读缓存
                clipped_segments, top_labels, point_markers = self._get_cached_side_face_result(proj_mode)

                if not clipped_segments and not point_markers:
                    self._draw_message(f"{self._row_name} 没有可绘制轮廓")
                    return

                # 由当前保留线段反推可见节点
                projected_nodes = {}
                for seg in clipped_segments:
                    ja = str(seg.get("joint_a") or "").strip()
                    jb = str(seg.get("joint_b") or "").strip()

                    if ja in self.nodes:
                        x, y, z = self.nodes[ja]
                        if proj_mode in ("XZ_FRONT", "XZ_BACK"):
                            projected_nodes[ja] = (float(x), float(z))
                        else:
                            projected_nodes[ja] = (float(y), float(z))

                    if jb in self.nodes:
                        x, y, z = self.nodes[jb]
                        if proj_mode in ("XZ_FRONT", "XZ_BACK"):
                            projected_nodes[jb] = (float(x), float(z))
                        else:
                            projected_nodes[jb] = (float(y), float(z))

            # ---------- 2) XY 截面 ----------
            else:
                row_nodes, row_members, proj_mode, top_labels = self._filter_row_members()

                if not row_nodes or not row_members:
                    self._draw_message(f"{self._row_name} 未识别到有效截面")
                    return

                if proj_mode == "XZ":
                    projected_nodes = {nid: (coord[0], coord[2]) for nid, coord in row_nodes.items()}
                elif proj_mode == "YZ":
                    projected_nodes = {nid: (coord[1], coord[2]) for nid, coord in row_nodes.items()}
                else:
                    projected_nodes = {nid: (coord[0], coord[1]) for nid, coord in row_nodes.items()}

                clipped_segments = []
                if proj_mode in ("XZ", "YZ"):
                    for na, nb, _gid in row_members:
                        if na not in projected_nodes or nb not in projected_nodes:
                            continue
                        clipped = self._clip_projected_member_to_workpoint(
                            projected_nodes[na], projected_nodes[nb]
                        )
                        if clipped is not None:
                            clipped_segments.append({
                                "joint_a": na,
                                "joint_b": nb,
                                "p1": clipped[0],
                                "p2": clipped[1],
                                "is_face": True,
                            })
                else:
                    for na, nb, _gid in row_members:
                        if na not in projected_nodes or nb not in projected_nodes:
                            continue
                        clipped_segments.append({
                            "joint_a": na,
                            "joint_b": nb,
                            "p1": projected_nodes[na],
                            "p2": projected_nodes[nb],
                            "is_face": True,
                        })

                if not clipped_segments:
                    self._draw_message(f"{self._row_name} 没有可绘制轮廓")
                    return

            # ---------- 3) 统一规范 point_markers ----------
            normalized_markers = []
            for mk in point_markers:
                if isinstance(mk, dict):
                    pt = mk.get("pt")
                    if pt and len(pt) >= 2:
                        normalized_markers.append({
                            "pt": (float(pt[0]), float(pt[1])),
                            "is_face": bool(mk.get("is_face", False)),
                        })
                elif isinstance(mk, (tuple, list)) and len(mk) >= 2:
                    normalized_markers.append({
                        "pt": (float(mk[0]), float(mk[1])),
                        "is_face": True,
                    })
            point_markers = normalized_markers

            # ---------- 4) 统计边界 ----------
            xs, ys = [], []
            for seg in clipped_segments:
                p1, p2 = seg["p1"], seg["p2"]
                xs.extend([p1[0], p2[0]])
                ys.extend([p1[1], p2[1]])

            for mk in point_markers:
                px, py = mk["pt"]
                xs.append(px)
                ys.append(py)

            if not xs or not ys:
                self._draw_message(f"{self._row_name} 没有可绘制轮廓")
                return

            min_x, max_x = min(xs), max(xs)
            min_y, max_y = min(ys), max(ys)

            view_w = 760.0
            view_h = 760.0
            margin_left = 40.0
            margin_right = 24.0
            margin_top = 32.0
            margin_bottom = 32.0

            dx = max(max_x - min_x, 1e-6)
            dy = max(max_y - min_y, 1e-6)
            avail_w = view_w - margin_left - margin_right
            avail_h = view_h - margin_top - margin_bottom
            scale = min(avail_w / dx, avail_h / dy)

            used_w = dx * scale
            used_h = dy * scale
            x_origin = (view_w - used_w) / 2.0
            y_base = view_h - margin_bottom - (avail_h - used_h) / 2.0

            def map_pt(xv: float, yv: float):
                px = x_origin + (xv - min_x) * scale
                py = y_base - (yv - min_y) * scale
                return px, py

            # ---------- 5) 顶部轴号 ----------
            for axis_value, axis_name in top_labels:
                px, _ = map_pt(axis_value, max_y)

                tick = RiskMemberItem(px, margin_top - 8, px, margin_top - 1)
                tick.setPen(QPen(self.COLOR_AXIS, 1.0))
                self._scene.addItem(tick)

                txt = QGraphicsSimpleTextItem(str(axis_name))
                txt.setBrush(QBrush(self.COLOR_AXIS))
                txt.setFont(QFont("Arial", 9, QFont.Bold))
                txt.setPos(px - 8, 8)
                self._scene.addItem(txt)

            # ---------- 6) 节点映射 ----------
            for nid, (nx, ny) in projected_nodes.items():
                self._visible_projected_nodes[nid] = map_pt(nx, ny)

            # ---------- 7) 分层绘制 ----------
            visible_nodes = {}
            self._visible_projected_members = []

            if proj_mode in directional_modes:
                other_segments = [seg for seg in clipped_segments if not seg.get("is_face", False)]
                face_segments = [seg for seg in clipped_segments if seg.get("is_face", False)]

                other_points = [mk for mk in point_markers if not mk.get("is_face", False)]
                face_points = [mk for mk in point_markers if mk.get("is_face", False)]

                # 灰线先画
                for seg in other_segments:
                    p1, p2 = seg["p1"], seg["p2"]
                    x1, y1 = map_pt(p1[0], p1[1])
                    x2, y2 = map_pt(p2[0], p2[1])

                    item = RiskMemberItem(x1, y1, x2, y2)
                    item.setPen(QPen(self.COLOR_MEMBER_OTHER, 1.0))
                    item.setZValue(1)
                    self._scene.addItem(item)

                # 灰点再画
                for mk in other_points:
                    px, py = mk["pt"]
                    sx, sy = map_pt(px, py)

                    item = RiskMemberItem(sx - 3, sy, sx + 3, sy)
                    item.setPen(QPen(self.COLOR_MARKER_OTHER, 1.0))
                    item.setZValue(2)
                    self._scene.addItem(item)

                # 黑线最后画
                for seg in face_segments:
                    p1, p2 = seg["p1"], seg["p2"]
                    joint_a = str(seg.get("joint_a") or "").strip()
                    joint_b = str(seg.get("joint_b") or "").strip()

                    x1, y1 = map_pt(p1[0], p1[1])
                    x2, y2 = map_pt(p2[0], p2[1])

                    item = RiskMemberItem(x1, y1, x2, y2)
                    item.setPen(QPen(self.COLOR_MEMBER_FACE, 1.45))
                    item.setZValue(10)
                    self._scene.addItem(item)

                    self._visible_projected_members.append({
                        "joint_a": joint_a,
                        "joint_b": joint_b,
                        "scene_p1": (x1, y1),
                        "scene_p2": (x2, y2),
                    })

                    if joint_a:
                        visible_nodes[joint_a] = (x1, y1)
                    if joint_b:
                        visible_nodes[joint_b] = (x2, y2)

                # 黑点最后画
                for mk in face_points:
                    px, py = mk["pt"]
                    sx, sy = map_pt(px, py)

                    item = RiskMemberItem(sx - 3.5, sy, sx + 3.5, sy)
                    item.setPen(QPen(self.COLOR_MARKER_FACE, 1.2))
                    item.setZValue(11)
                    self._scene.addItem(item)

                self._visible_projected_nodes = visible_nodes


            else:

                xy_color = QColor(0, 0, 0)

                for seg in clipped_segments:

                    p1, p2 = seg["p1"], seg["p2"]

                    joint_a = str(seg.get("joint_a") or "").strip()

                    joint_b = str(seg.get("joint_b") or "").strip()

                    x1, y1 = map_pt(p1[0], p1[1])

                    x2, y2 = map_pt(p2[0], p2[1])

                    item = RiskMemberItem(x1, y1, x2, y2)

                    item.setPen(QPen(xy_color, 1.30))

                    item.setZValue(10)

                    self._scene.addItem(item)

                    self._visible_projected_members.append({

                        "joint_a": joint_a,

                        "joint_b": joint_b,

                        "scene_p1": (x1, y1),

                        "scene_p2": (x2, y2),

                    })

                    if joint_a:
                        visible_nodes[joint_a] = (x1, y1)

                    if joint_b:
                        visible_nodes[joint_b] = (x2, y2)

                self._visible_projected_nodes = visible_nodes

            self._scene.setSceneRect(QRectF(0, 0, view_w, view_h))
            self._draw_inspection_overlay()
            self._draw_history_overlay()
            self._show_hover_info()

        except Exception:
            print("[Elevation] _render_row_elevation failed")
            traceback.print_exc()
            self._draw_message("立面图绘制失败，请查看控制台日志")

        finally:
            self._scene.blockSignals(False)
            self.setUpdatesEnabled(True)
            self.viewport().update()

    def _draw_message(self, text: str):
        self._scene.clear()
        item = QGraphicsSimpleTextItem(text)
        item.setBrush(QBrush(QColor(60, 60, 60)))
        item.setFont(QFont("Arial", 10))
        item.setPos(20, 20)
        self._scene.addItem(item)

        rect = item.boundingRect().adjusted(-20, -20, 40, 40)
        self._scene.setSceneRect(rect)

    def _copy_render_state_to(self, target: "SacsElevationRiskView") -> None:
        """把当前立面图状态复制到另一个 SacsElevationRiskView，用于全屏窗口。"""
        target._facility_code = self._facility_code
        target._model_path = self._model_path
        target._row_name = self._row_name
        target._workpoint_z = self._workpoint_z
        target._level_threshold = self._level_threshold

        target.nodes = dict(self.nodes or {})
        target.members = list(self.members or [])
        target.node_risk_map = dict(self.node_risk_map or {})
        target.member_risk_map = dict(self.member_risk_map or {})
        target._groups_od = dict(self._groups_od or {})

        target._inspection_overlay_enabled = bool(self._inspection_overlay_enabled)
        target._member_inspect_level_map = dict(self._member_inspect_level_map or {})
        target._node_inspect_level_map = dict(self._node_inspect_level_map or {})
        target._node_brace_inspect_level_map = dict(self._node_brace_inspect_level_map or {})
        target._member_items_by_key = dict(self._member_items_by_key or {})
        target._node_items_by_joint = dict(self._node_items_by_joint or {})
        target._inspection_visible_levels = set(self._inspection_visible_levels or set())
        target._show_inspection_level_ii = bool(self._show_inspection_level_ii)

        target._history_overlay_enabled = bool(self._history_overlay_enabled)
        target._history_overlay_items = list(self._history_overlay_items or [])
        target._history_overlay_legend = list(self._history_overlay_legend or [])

        target._cached_model_path = self._cached_model_path
        target._cached_nodes = dict(self._cached_nodes or {})
        target._cached_members = list(self._cached_members or [])
        target._cached_groups_od = dict(self._cached_groups_od or {})

    def open_fullscreen_window(self, title: str = "模型立面检验等级图") -> None:
        """在新的最大化窗口中显示当前立面图。

        不改变原页面视图；全屏窗口会复制当前模型、立面、检验等级、历史点位、
        风险标记、鼠标提示等绘图状态，并重新渲染一份。
        """
        dlg = QDialog(self)
        dlg.setWindowTitle(f"{title} - 全屏")
        dlg.resize(1280, 880)

        layout = QVBoxLayout(dlg)
        layout.setContentsMargins(8, 8, 8, 8)
        layout.setSpacing(6)

        top = QWidget(dlg)
        top_lay = QHBoxLayout(top)
        top_lay.setContentsMargins(0, 0, 0, 0)
        top_lay.setSpacing(8)
        title_label = QLabel(f"{title}（全屏）", top)
        title_label.setStyleSheet("font-size:13pt; font-weight:bold; color:#1d2b3a;")
        info_label = QLabel("滚轮缩放，双击恢复初始视图，ESC/关闭按钮退出。", top)
        info_label.setStyleSheet("font-size:10pt; color:#5d6f85;")
        btn_close = QPushButton("关闭", top)
        btn_close.setFixedSize(72, 28)
        btn_close.clicked.connect(dlg.close)
        top_lay.addWidget(title_label, 0)
        top_lay.addWidget(info_label, 1)
        top_lay.addWidget(btn_close, 0)
        layout.addWidget(top, 0)

        view_row = QHBoxLayout()
        view_row.setContentsMargins(0, 0, 0, 0)
        view_row.setSpacing(6)

        full_view = SacsElevationRiskView(dlg)
        full_view.set_info_label(info_label)
        view_row.addWidget(full_view, 1)

        slider_v = QSlider(Qt.Vertical, dlg)
        slider_v.setRange(-100, 100)
        slider_v.setValue(0)
        view_row.addWidget(slider_v, 0)
        layout.addLayout(view_row, 1)

        slider_h = QSlider(Qt.Horizontal, dlg)
        slider_h.setRange(-100, 100)
        slider_h.setValue(0)
        layout.addWidget(slider_h, 0)

        full_view.bind_sliders(slider_h, slider_v)
        slider_h.valueChanged.connect(lambda v: full_view.pan_view(v, slider_v.value()))
        slider_v.valueChanged.connect(lambda v: full_view.pan_view(slider_h.value(), v))

        self._copy_render_state_to(full_view)
        if full_view.nodes and full_view.members:
            full_view._render_row_elevation()
            full_view._fit_done = False
            full_view._zoom_steps = 0
            QTimer.singleShot(0, full_view.reset_view)
        else:
            full_view._draw_message("当前没有可全屏显示的图形")

        dlg.showMaximized()
        exec_dialog_safely(dlg, title="立面图窗口错误", context="立面图全屏窗口", parent=self)

    def resizeEvent(self, event):
        super().resizeEvent(event)
        if self._fit_done or self._in_reset_view or self._reset_pending:
            return
        self._reset_pending = True
        QTimer.singleShot(0, self.reset_view)

    def wheelEvent(self, event):
        dy = event.angleDelta().y()
        if dy == 0:
            event.accept()
            return

        delta = 1 if dy > 0 else -1
        new_steps = self._zoom_steps + delta

        if new_steps < -8 or new_steps > 20:
            event.accept()
            return

        self._zoom_steps = new_steps

        factor = 1.08 if delta > 0 else 1 / 1.08
        self.scale(factor, factor)

        # 缩放后保持当前平移状态
        if self._slider_h is not None and self._slider_v is not None:
            self.pan_view(self._slider_h.value(), self._slider_v.value())

        event.accept()

    def mouseDoubleClickEvent(self, event):
        self.reset_view()
        event.accept()
