#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
inspection_tool.py

Single-file implementation of the converted WC19-1D VBA workflow.

Active runtime path:
    run_sacs_report_from_config.py -> sacs_to_report.py -> inspection_tool.py

Business modules covered in this file:
    1. Structural model parsing
    2. Fatigue input interpretation
    3. Fatigue result interpretation
    4. Collapse result interpretation
    5. Consequence level determination
    6. Fatigue failure level calculation
    7. Collapse failure level calculation
    8. Overall risk level calculation
    9. Inspection strategy generation

Main dependency stack:
    pandas, numpy, openpyxl
"""
from __future__ import annotations

import argparse
import glob
import json
import math
import re
import random
import sys
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from dataclasses import dataclass
from pathlib import Path
from statistics import NormalDist
from typing import Dict, List, Tuple, Optional, Any, Iterable, Sequence, Set

import numpy as np
import pandas as pd
import openpyxl

_STD_NORM = NormalDist()
MAX_VBA_COLLAPSE_FILES = 12
MAX_VBA_FATIGUE_FILES = 7
MAX_VBA_FTGINP_FILES = 8
FATIGUE_AUDIT_COLUMNS = [
    "FileIndex",
    "FilePath",
    "Joint",
    "Action",
    "Brace",
    "Score",
    "CandidateCount",
    "Remark",
    "ManualBrace",
    "Applied",
]

# One-file module map used to keep the converted workflow manageable.
# Each item points to the primary function(s) that implement the corresponding
# VBA business module in this file.
PIPELINE_MODULE_MAP = {
    1: ("结构模型解析", ["parse_sacinp", "classify_structure"]),
    2: ("疲劳分析输入模型解释", ["parse_ftginp_ringmember", "parse_ftginp_ringmembers"]),
    3: ("疲劳分析结果解释", ["parse_ftglst_detail", "parse_ftglst_details"]),
    4: ("倒塌分析结果解释", ["parse_clplog", "parse_clplogs"]),
    5: ("失效后果等级确定", ["_global_level_from_tag", "_local_level_member", "_local_level_joint"]),
    6: ("疲劳失效等级计算", ["fatigue_ctf", "fatigue_beta_current", "fatigue_beta_forecast", "pf_from_beta"]),
    7: ("倒塌失效等级计算", ["collapse_pf", "_min_factor_by_location", "_collapse_rsr"]),
    8: ("整体风险等级计算", ["build_member_risk_vba", "build_joint_risk_vba", "build_joint_forecast_vba_wide"]),
    9: ("检测策略形成", ["build_node_plan_vba", "build_member_plan_vba"]),
}


def _norm_cdf(x: float) -> float:
    return float(_STD_NORM.cdf(float(x)))


def _norm_pdf(x: float) -> float:
    xv = float(x)
    return float(math.exp(-0.5 * xv * xv) / math.sqrt(2.0 * math.pi))


# =========================
# Utilities
# =========================

def _as_float(x: Any) -> Optional[float]:
    if x is None:
        return None
    try:
        return float(x)
    except Exception:
        return None

def _cell(ws, addr: str) -> Any:
    return ws[addr].value

def _safe_str(x: Any) -> str:
    return "" if x is None else str(x).strip()

def _mkdir_for_file(path: str | Path) -> None:
    Path(path).parent.mkdir(parents=True, exist_ok=True)


def _fw_sub(s: str, start: int, width: int) -> str:
    """
    VBA-like Mid(s, start, width) with 1-based start index.
    """
    if width <= 0:
        return ""
    i0 = max(0, start - 1)
    i1 = i0 + width
    if i0 >= len(s):
        return ""
    return s[i0:i1]


def _parse_float_or_none(s: Any) -> Optional[float]:
    if s is None:
        return None
    txt = str(s).strip()
    if txt == "":
        return None
    try:
        return float(txt)
    except Exception:
        m = _RE_FLOAT.search(txt)
        if m:
            try:
                return float(m.group(0))
            except Exception:
                return None
        return None


def _ensure_path_list(path_or_paths: str | Path | Sequence[str | Path]) -> List[Path]:
    if isinstance(path_or_paths, (str, Path)):
        return [Path(path_or_paths)]
    out: List[Path] = []
    for p in path_or_paths:
        out.append(Path(p))
    return out


def _natural_sort_key(path_like: str | Path) -> List[Any]:
    text = str(path_like).replace("\\", "/")
    parts = re.split(r"(\d+)", text)
    key: List[Any] = []
    for p in parts:
        if p.isdigit():
            key.append(int(p))
        else:
            key.append(p.lower())
    return key


def _resolve_multi_inputs(
    path_or_paths: str | Path | Sequence[str | Path],
    dir_pattern: str | None = None,
) -> List[Path]:
    """
    Resolve multi-input arguments:
    - file path
    - directory (expand by dir_pattern)
    - wildcard path (glob)
    Keep user-provided group order and de-duplicate by absolute path.
    """
    raw = _ensure_path_list(path_or_paths)
    resolved: List[Path] = []
    seen: Set[str] = set()

    for raw_p in raw:
        text = str(raw_p)
        has_glob = any(ch in text for ch in ("*", "?", "[", "]"))
        expanded: List[Path] = []

        if has_glob:
            expanded = [Path(p) for p in glob.glob(text)]
            expanded = [p for p in expanded if p.is_file()]
            expanded.sort(key=_natural_sort_key)
        elif raw_p.is_dir():
            pattern = dir_pattern or "*"
            expanded = [p for p in raw_p.glob(pattern) if p.is_file()]
            expanded.sort(key=_natural_sort_key)
        else:
            expanded = [raw_p]

        if not expanded and (has_glob or raw_p.is_dir()):
            raise FileNotFoundError(f"no files resolved from input: {raw_p}")

        for p in expanded:
            if not p.is_file():
                raise FileNotFoundError(f"file not found: {p}")
            rp = str(p.resolve())
            if rp in seen:
                continue
            seen.add(rp)
            resolved.append(p)

    return resolved


def _file_line_count(path: Path) -> int:
    cnt = 0
    with path.open("r", encoding="utf-8", errors="ignore") as f:
        for _ in f:
            cnt += 1
    return cnt


# =========================
# Geometry & Graph
# =========================

def _vec(p1: Tuple[float,float,float], p2: Tuple[float,float,float]) -> Tuple[float,float,float]:
    return (p2[0]-p1[0], p2[1]-p1[1], p2[2]-p1[2])

def _norm(v: Tuple[float,float,float]) -> float:
    return math.sqrt(v[0]*v[0]+v[1]*v[1]+v[2]*v[2])

def _angle_deg(v1: Tuple[float,float,float], v2: Tuple[float,float,float]) -> float:
    n1, n2 = _norm(v1), _norm(v2)
    if n1 == 0 or n2 == 0:
        return 180.0
    dot = (v1[0]*v2[0]+v1[1]*v2[1]+v1[2]*v2[2])/(n1*n2)
    dot = max(-1.0, min(1.0, dot))
    return math.degrees(math.acos(dot))

def is_near_vertical(p1: Tuple[float,float,float], p2: Tuple[float,float,float], tol_deg: float) -> bool:
    v = _vec(p1,p2)
    ang = min(_angle_deg(v,(0,0,1)), _angle_deg(v,(0,0,-1)))
    return ang <= tol_deg


def _round_half_up_scalar(x: float, digits: int) -> float:
    q = Decimal("1").scaleb(-digits)
    return float(Decimal(str(x)).quantize(q, rounding=ROUND_HALF_UP))


def vector_angle_degree_vba(v1: Tuple[float, float, float], v2: Tuple[float, float, float]) -> float:
    """
    VBA VectorAngleDegree emulation:
    cosine is rounded to 5 decimals before Acos.
    """
    vv = v1[0] * v2[0] + v1[1] * v2[1] + v1[2] * v2[2]
    l1 = _norm(v1)
    l2 = _norm(v2)
    if l1 == 0.0 or l2 == 0.0:
        return 0.0
    cosine = _round_half_up_scalar(vv / (l1 * l2), 5)
    cosine = max(-1.0, min(1.0, cosine))
    return float(math.degrees(math.acos(cosine)))

@dataclass(frozen=True)
class Edge:
    idx: int
    a: str
    b: str
    od: float

def build_adjacency(members: pd.DataFrame) -> Dict[str, List[Edge]]:
    adj: Dict[str, List[Edge]] = {}
    for i, r in members.reset_index(drop=True).iterrows():
        a = str(r["A"]); b = str(r["B"])
        od = float(r["OD"]) if pd.notna(r["OD"]) else float("nan")
        e = Edge(int(i), a, b, od)
        adj.setdefault(a, []).append(e)
        adj.setdefault(b, []).append(e)
    return adj

def other_end(e: Edge, joint: str) -> str:
    return e.b if joint == e.a else e.a


# =========================
# Module 1: 结构模型解析
# Source: sacinp.* / SACS model input
# Output: joints, groups, members, sections, then JointType/MemberType
# =========================

_RE_FLOAT = re.compile(r"[-+]?\d+(?:\.\d+)?(?:[Ee][-+]?\d+)?")

def parse_sacinp(path: str | Path) -> Tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    """
    Strict port of VBA ReadSACS:
    - fixed-column Mid parsing only
    - no free-format fallback parse
    - no dedup on Sections/Groups
    - stop reading when CENTER/SURFID/WGTFP/LOADCN appears
    """
    p = Path(path)
    if not p.exists():
        raise FileNotFoundError(f"sacinp not found: {p}")

    sections: List[Dict[str, Any]] = []
    groups: List[Dict[str, Any]] = []
    members_raw: List[Dict[str, Any]] = []
    joints: List[Dict[str, Any]] = []
    # VBA SQL "SELECT * FROM [Sections$] WHERE ID='...'" returns first match.
    # Keep first OD per section ID for GRUP fallback when OD column is blank.
    section_first_od_by_id: Dict[str, float] = {}

    with p.open("r", encoding="utf-8", errors="ignore") as f:
        for raw in f:
            line = raw.rstrip("\r\n")

            if line.startswith("SECT"):
                sid = _fw_sub(line, 6, 7).strip()
                stype = _fw_sub(line, 16, 3).strip()
                if sid == "":
                    continue
                od = None
                # VBA: If Type is TUB/CON, OD = CDbl(Mid(A,50,6))
                if stype in {"TUB", "CON"}:
                    od = _parse_float_or_none(_fw_sub(line, 50, 6))
                sections.append({"ID": sid, "Type": stype if stype else None, "OD": od})
                if od is not None and sid not in section_first_od_by_id:
                    section_first_od_by_id[sid] = float(od)
                continue

            if line.startswith("GRUP"):
                gid = _fw_sub(line, 6, 3).strip()
                if gid == "":
                    continue
                sect_hint = _fw_sub(line, 10, 7).strip()
                od = _parse_float_or_none(_fw_sub(line, 18, 6))
                # VBA fallback: SELECT * FROM [Sections$] WHERE ID=Mid(A,10,7), use field 3 (OD).
                if od is None and sect_hint in section_first_od_by_id:
                    od = section_first_od_by_id[sect_hint]

                # ReadSACS does not populate Groups.Type.
                groups.append({"ID": gid, "OD": od, "Type": None})
                continue

            if line.startswith("MEMBER"):
                token8 = _fw_sub(line, 8, 8).strip()
                token7 = _fw_sub(line, 8, 7).strip().upper()
                # Skip header / OFFSETS auxiliary lines.
                if token8 == "" or token7 == "OFFSETS":
                    continue
                a = _fw_sub(line, 8, 4).strip()
                b = _fw_sub(line, 12, 4).strip()
                gid = _fw_sub(line, 17, 3).strip()

                if a and b and gid:
                    members_raw.append({"A": a, "B": b, "ID": gid, "OD": None, "MemberType": None, "Z1": None, "Z2": None})
                continue

            if line.startswith("JOINT"):
                token8 = _fw_sub(line, 7, 8).strip()
                token7 = _fw_sub(line, 8, 7).strip().upper()
                if token8 == "" or token7 == "OFFSETS":
                    continue
                jid = _fw_sub(line, 7, 4).strip()
                if jid == "":
                    continue

                x = _parse_float_or_none(_fw_sub(line, 12, 7))
                y = _parse_float_or_none(_fw_sub(line, 19, 7))
                z = _parse_float_or_none(_fw_sub(line, 26, 7))
                dx = _parse_float_or_none(_fw_sub(line, 33, 7))
                dy = _parse_float_or_none(_fw_sub(line, 40, 7))
                dz = _parse_float_or_none(_fw_sub(line, 47, 7))

                x = 0.0 if x is None else float(x)
                y = 0.0 if y is None else float(y)
                z = 0.0 if z is None else float(z)
                if dx is not None:
                    x += float(dx) / 100.0
                if dy is not None:
                    y += float(dy) / 100.0
                if dz is not None:
                    z += float(dz) / 100.0

                joints.append({"Joint": jid, "X": x, "Y": y, "Z": z, "JointType": None})
                continue

            if any(k in line for k in ("CENTER", "SURFID", "WGTFP", "LOADCN")):
                break

    joints_df = pd.DataFrame(joints, columns=["Joint","X","Y","Z","JointType"])
    groups_df = pd.DataFrame(groups, columns=["ID","OD","Type"])
    members_df = pd.DataFrame(members_raw, columns=["A","B","ID","OD","MemberType","Z1","Z2"])
    sections_df = pd.DataFrame(sections, columns=["ID","Type","OD"])

    if not groups_df.empty:
        # VBA uses MAX(OD) by group ID when filling member OD.
        od_by_group = groups_df.groupby("ID", dropna=False)["OD"].max().to_dict()
        members_df["OD"] = members_df["ID"].map(od_by_group)

    if not joints_df.empty and not members_df.empty:
        jz = joints_df.set_index("Joint")["Z"].to_dict()
        members_df["Z1"] = members_df["A"].map(jz)
        members_df["Z2"] = members_df["B"].map(jz)

    return joints_df, groups_df, members_df, sections_df


# =========================
# Module 4: 倒塌分析结果解释
# Source: clplog.*
# Output: collapse failure rows + per-file last load factor / RSR basis
# =========================

def _is_vba_numeric_prefix(s: str) -> bool:
    """
    VBA IsNumeric(Left(currentLine, 2)) behavior approximation.
    """
    t = s[:2]
    try:
        float(t)
        return True
    except Exception:
        return False


def parse_clplog(path: str | Path, load_id: Optional[int] = None) -> Tuple[pd.DataFrame, Optional[float]]:
    """
    Strict port of Sheet5.ParseCollapseAnalysis for one file:
    - parse line-by-line with currentLine = Trim(line)
    - update lastLoadFactor from Mid(line,25,5) when IsNumeric(Left(currentLine,2))
      and token count >= 5
    - parse:
      * "*** WARNING - JOINT ..." -> 节点失效
      * "*** MEMBER ... HAS ... AT SEGMENT ..." -> 构件失效
    """
    p = Path(path)
    if not p.exists():
        raise FileNotFoundError(f"clplog not found: {p}")

    records: List[Dict[str, Any]] = []
    last_load_factor: Optional[float] = None

    with p.open("r", encoding="utf-8", errors="ignore") as f:
        for raw_line in f:
            line = raw_line.rstrip("\r\n")
            current_line = line.strip()

            # VBA:
            # If IsNumeric(Left(currentLine, 2)) Then
            #   parts = Split(Application.Trim(currentLine), " ")
            #   If UBound(parts) >= 4 Then lastLoadFactor = Mid(lines(i), 25, 5)
            if current_line != "" and _is_vba_numeric_prefix(current_line):
                parts = re.split(r"\s+", current_line)
                if len(parts) >= 5:
                    last_load_factor = _parse_float_or_none(_fw_sub(line, 25, 5))

            if current_line.startswith("*** WARNING - JOINT"):
                joint_info = ""
                pos_joint = current_line.find("AT JOINT")
                if pos_joint >= 0:
                    rest = current_line[pos_joint + 9 :].strip()
                    if rest:
                        joint_info = rest.split(" ", 1)[0]

                remark = ""
                pos_brace = current_line.find("FOR BRACE")
                if pos_brace >= 0:
                    remark_info = current_line[pos_brace:]
                    pos_load = remark_info.find(" AT LOAD")
                    if pos_load >= 0:
                        # VBA Left(remarkinfo, InStr(...," AT LOAD"))
                        remark = remark_info[: pos_load + 1].rstrip()
                    else:
                        remark = remark_info

                records.append({
                    "LOADID": load_id,
                    "TYPE": "节点失效",
                    "LOCATION": joint_info,
                    "FACTOR": last_load_factor,
                    "REMARK": remark,
                    "SOURCE_FILE": str(p),
                })
                continue

            if current_line.startswith("*** MEMBER"):
                member_info = ""
                pos_member = current_line.find("MEMBER")
                pos_has = current_line.find(" HAS")
                if pos_member >= 0 and pos_has > (pos_member + 7):
                    member_info = current_line[pos_member + 7 : pos_has].strip()

                remark = ""
                pos_seg = current_line.find("AT SEGMENT")
                if pos_seg >= 0:
                    remark = current_line[pos_seg:]

                records.append({
                    "LOADID": load_id,
                    "TYPE": "构件失效",
                    "LOCATION": member_info,
                    "FACTOR": last_load_factor,
                    "REMARK": remark,
                    "SOURCE_FILE": str(p),
                })
                continue

    return (
        pd.DataFrame(records, columns=["LOADID", "TYPE", "LOCATION", "FACTOR", "REMARK", "SOURCE_FILE"]),
        last_load_factor,
    )


def parse_clplogs(paths: Sequence[str | Path]) -> Tuple[pd.DataFrame, pd.DataFrame]:
    """
    Multi-file strict port of Sheet5.ParseCollapseAnalysis loop f_i=2..13:
    - LOADID = file order (1-based, equivalent to f_i-1)
    - max files = 12 (VBA B2:B13)
    - returns:
      1) collapse rows dataframe
      2) per-file summary dataframe (LOADID, LastLoadFactor, SOURCE_FILE)
    """
    frames: List[pd.DataFrame] = []
    summary_rows: List[Dict[str, Any]] = []
    for idx, p in enumerate(_ensure_path_list(paths), start=1):
        df, last_factor = parse_clplog(p, load_id=idx)
        if df.empty:
            summary_rows.append(
                {
                    "LOADID": idx,
                    "LastLoadFactor": last_factor,
                    "SOURCE_FILE": str(Path(p)),
                }
            )
        else:
            frames.append(df)
            summary_rows.append(
                {
                    "LOADID": idx,
                    "LastLoadFactor": last_factor,
                    "SOURCE_FILE": str(Path(p)),
                }
            )
    collapse_df = (
        pd.concat(frames, ignore_index=True)
        if frames
        else pd.DataFrame(columns=["LOADID", "TYPE", "LOCATION", "FACTOR", "REMARK", "SOURCE_FILE"])
    )
    summary_df = pd.DataFrame(summary_rows, columns=["LOADID", "LastLoadFactor", "SOURCE_FILE"])
    return collapse_df, summary_df


# =========================
# Module 3: 疲劳分析结果解释
# Source: ftglst.*
# Output: fatigue damage table used by the later node risk modules
# =========================

_RE_FTG_DETAIL_HEADER = re.compile(r"MEMBER\s+FATIGUE\s+DETAIL\s+REPORT", re.I)
_RE_FTG_DETAIL_COLHDR = re.compile(r"JOINT\s+MEMBER\s+GRUP\s+LOAD", re.I)
_RE_FTG_DETAIL_COLHDR_ALT = re.compile(r"JOINT\s+CHD\s+BRC\s+GRUP\s+LOAD", re.I)

FTG_POS = ["TOP","TOP_LEFT","LEFT","BOT_LEFT","BOT","BOT_RIGHT","RIGHT","TOP_RIGHT"]

def parse_sacs_damage_number(s: str) -> float:
    s = s.strip()
    if not s:
        return 0.0
    try:
        return float(s)
    except ValueError:
        m = re.fullmatch(r"([+-]?\d*\.\d+)([+-]\d+)", s)
        if m:
            return float(f"{m.group(1)}E{m.group(2)}")
        nums = _RE_FLOAT.findall(s)
        return float(nums[0]) if nums else 0.0


def _normalize_spaced_caps(line: str) -> str:
    """
    Normalize report headings where words are printed as spaced capitals,
    e.g. 'M E M B E R  F A T I G U E' -> 'MEMBER FATIGUE'.
    """
    text = line.upper()
    # Join spaced letters inside a word: "M E M B E R" -> "MEMBER".
    # Keep multi-space word boundaries intact at this stage.
    text = re.sub(r"(?<=\b[A-Z]) (?=[A-Z]\b)", "", text)
    text = re.sub(r"\s+", " ", text).strip()
    return text


def _is_int_token(token: str) -> bool:
    return re.fullmatch(r"[+-]?\d+", token) is not None

def _parse_total_damage_values(line: str) -> np.ndarray:
    parts = line.replace("*", " ").split()
    start = 0
    for i, token in enumerate(parts):
        if token.upper() == "DAMAGE":
            start = i + 1
            break
    vals: List[float] = []
    for token in parts[start:]:
        if len(vals) >= 8:
            break
        if re.search(r"\d", token):
            vals.append(parse_sacs_damage_number(token))
    while len(vals) < 8:
        vals.append(0.0)
    return np.array(vals[:8], dtype=float)


def _parse_ftglst_pair_rows(
    path: str | Path,
    default_factor: float = 1.0,
    factor_map: Optional[Dict[str, float]] = None,
) -> List[Dict[str, Any]]:
    """
    Parse ftglst using the same pair logic as VBA Sheet8_FatiguePickup:
    one row = brace-total-damage block + chord-total-damage block.
    """
    p = Path(path)
    if not p.exists():
        raise FileNotFoundError(f"ftglst not found: {p}")

    fmap = factor_map or {}
    out: List[Dict[str, Any]] = []

    in_block = False
    brace_mode = True
    joint = ""
    chd_a = ""
    chd_b = ""
    brc = ""
    grup_b = ""
    grup_c = ""
    brace_damage: Optional[np.ndarray] = None

    with p.open("r", encoding="utf-8", errors="ignore") as f:
        for raw_line in f:
            line = raw_line.rstrip("\r\n")
            norm = _normalize_spaced_caps(line)

            if not in_block:
                # Strict VBA section start:
                # "* *  M E M B E R  F A T I G U E  D E T A I L  R E P O R T  * *"
                if ("* * MEMBER FATIGUE DETAIL REPORT * *" in norm) and ("INTERMEDIATE" not in norm):
                    in_block = True
                continue

            # Strict VBA section end:
            # "*  *  *  M E M B E R  F A T I G U E  R E P O R T  *  *  *"
            if ("* * * MEMBER FATIGUE REPORT * * *" in norm) and ("INTERMEDIATE" not in norm):
                break

            if _fw_sub(line, 1, 4).strip() != "" and ("*" not in line):
                # Strict fixed-width extraction, same as VBA Mid positions.
                if brace_mode:
                    joint = _fw_sub(line, 1, 4).strip()
                    brc = _fw_sub(line, 12, 4).strip()
                    grup_b = _fw_sub(line, 18, 3).strip()
                else:
                    chd_a = _fw_sub(line, 7, 4).strip()
                    chd_b = _fw_sub(line, 12, 4).strip()
                    grup_c = _fw_sub(line, 18, 3).strip()

            if "TOTAL DAMAGE" not in norm:
                continue

            factor_val = _as_float(fmap.get(joint, default_factor))
            if factor_val is None or float(factor_val) == 0:
                factor_val = 1.0
            scaled = _parse_total_damage_values(line) / float(factor_val)

            if brace_mode:
                brace_damage = scaled
            else:
                chord_damage = scaled
                if brace_damage is not None and joint and chd_a and chd_b and brc:
                    out.append(
                        {
                            "joint": joint,
                            "chd_a": chd_a,
                            "chd_b": chd_b,
                            "brc": brc,
                            "selector_key": f"{joint}{brc}",
                            "full_key": f"{joint}{chd_a}{chd_b}{brc}",
                            "grup": grup_b if grup_b else (grup_c if grup_c else None),
                            "brace_damage": brace_damage.copy(),
                            "chord_damage": chord_damage.copy(),
                        }
                    )
                brace_damage = None

            brace_mode = not brace_mode

    return out


def _build_ftglst_pair_row(
    *,
    joint: str,
    chd_a: str,
    chd_b: str,
    brc: str,
    grup: Optional[str],
    brace_damage: np.ndarray,
    chord_damage: np.ndarray,
) -> Optional[Dict[str, Any]]:
    joint_s = _safe_str(joint)
    chd_a_s = _safe_str(chd_a)
    chd_b_s = _safe_str(chd_b)
    brc_s = _safe_str(brc)
    if joint_s == "" or chd_a_s == "" or chd_b_s == "" or brc_s == "":
        return None
    return {
        "joint": joint_s,
        "chd_a": chd_a_s,
        "chd_b": chd_b_s,
        "brc": brc_s,
        "selector_key": f"{joint_s}{brc_s}",
        "full_key": f"{joint_s}{chd_a_s}{chd_b_s}{brc_s}",
        "grup": grup,
        "brace_damage": np.array(brace_damage, dtype=float),
        "chord_damage": np.array(chord_damage, dtype=float),
    }


def _is_ftg_detail_start(line_norm: str) -> bool:
    return ("* * MEMBER FATIGUE DETAIL REPORT * *" in line_norm) and ("INTERMEDIATE" not in line_norm)


def _is_ftg_detail_end(line_norm: str) -> bool:
    return ("* * * MEMBER FATIGUE REPORT * * *" in line_norm) and ("INTERMEDIATE" not in line_norm)


def _parse_ftglst_vba_turn_rows(
    path: str | Path,
    turn: int,
    default_factor: float,
    factor_map: Dict[str, float],
    selectors: Set[str],
) -> Tuple[List[Dict[str, Any]], Dict[str, Dict[str, Any]]]:
    """
    Strict VBA-style parser for one fatigue result file in FatiguePickup.

    turn == 1:
      initialize ARR1 baseline rows.
    turn > 1:
      only collect replacement rows gated by selector state machine
      (CP/Jck/ch logic in VBA).
    """
    p = Path(path)
    if not p.exists():
        raise FileNotFoundError(f"ftglst not found: {p}")

    # VBA local state.
    b_report_section = False
    brace_data = True
    cp = ""
    jck = False
    ch = 0

    joint_num = ""
    chd_a = ""
    chd_b = ""
    brc = ""
    grup_b = ""
    grup_c = ""
    brace_damage_buf: Optional[np.ndarray] = None
    brace_joint_buf = ""
    brace_brc_buf = ""
    brace_grup_buf: Optional[str] = None

    baseline_rows: List[Dict[str, Any]] = []
    replace_map: Dict[str, Dict[str, Any]] = {}

    with p.open("r", encoding="utf-8", errors="ignore") as f:
        for raw_line in f:
            line = raw_line.rstrip("\r\n")
            line_norm = _normalize_spaced_caps(line)

            # VBA termination marker.
            if _is_ftg_detail_end(line_norm):
                break

            if _is_ftg_detail_start(line_norm):
                b_report_section = True

            # Strictly mirror VBA: Aaas = Left(TextLine,4) & Mid(TextLine,12,4)
            aaas_raw = _fw_sub(line, 1, 4) + _fw_sub(line, 12, 4)
            aaas = aaas_raw.strip()

            if turn > 1:
                if cp != aaas_raw and aaas != "":
                    if (aaas in selectors) and b_report_section:
                        cp = aaas_raw
                        jck = True
                        brace_damage_buf = None
                        brace_joint_buf = ""
                        brace_brc_buf = ""
                        brace_grup_buf = None

            # turn==1 always parse report section; turn>1 parse only when Jck=true.
            active = b_report_section if turn == 1 else (b_report_section and jck)
            if not active:
                continue

            # VBA gate in turn>1:
            # If Mid(TextLine, 4, 20) = "*** TOTAL DAMAGE ***" Then ch = ch + 1
            # If ch = 2 Then ch = 0 : Jck = False
            if turn > 1:
                if _fw_sub(line, 4, 20) == "*** TOTAL DAMAGE ***":
                    ch += 1
                if ch == 2:
                    ch = 0
                    jck = False

            # Parse header lines (non-star with left4 token).
            if _fw_sub(line, 1, 4).strip() != "" and ("*" not in line):
                if brace_data:
                    joint_num = _fw_sub(line, 1, 4).strip()
                    brc = _fw_sub(line, 12, 4).strip()
                    grup_b = _fw_sub(line, 18, 3).strip()
                else:
                    chd_a = _fw_sub(line, 7, 4).strip()
                    chd_b = _fw_sub(line, 12, 4).strip()
                    grup_c = _fw_sub(line, 18, 3).strip()

            # TOTAL DAMAGE rows.
            if "*** TOTAL DAMAGE ***" not in line:
                continue

            fac = _as_float(factor_map.get(joint_num, default_factor))
            if fac is None or float(fac) == 0.0:
                fac = 1.0
            scaled = _parse_total_damage_values(line) / float(fac)

            if brace_data:
                brace_damage_buf = np.array(scaled, dtype=float)
                brace_joint_buf = joint_num
                brace_brc_buf = brc
                brace_grup_buf = grup_b if grup_b else None
            else:
                if brace_damage_buf is not None:
                    row = _build_ftglst_pair_row(
                        joint=brace_joint_buf,
                        chd_a=chd_a,
                        chd_b=chd_b,
                        brc=brace_brc_buf,
                        grup=brace_grup_buf if brace_grup_buf else (grup_c if grup_c else None),
                        brace_damage=brace_damage_buf,
                        chord_damage=np.array(scaled, dtype=float),
                    )
                    if row is not None:
                        if turn == 1:
                            baseline_rows.append(row)
                        else:
                            replace_map[str(row["full_key"])] = row
                brace_damage_buf = None

            brace_data = not brace_data

    return baseline_rows, replace_map


def _pair_rows_to_fatigue_df(pair_rows: Sequence[Dict[str, Any]]) -> pd.DataFrame:
    """
    Output both BRACE SIDE and CHORD SIDE 8-direction damages (VBA sheet layout),
    while keeping merged TOP..TOP_RIGHT and Dmax for downstream compatibility.
    """
    brace_cols = [f"BRACE_{p}" for p in FTG_POS]
    chord_cols = [f"CHORD_{p}" for p in FTG_POS]

    rows: List[Dict[str, Any]] = []
    for r in pair_rows:
        brace = np.array(r.get("brace_damage", np.zeros(8, dtype=float)), dtype=float)
        chord = np.array(r.get("chord_damage", np.zeros(8, dtype=float)), dtype=float)
        merged = np.maximum(brace, chord)
        dmax = float(max(np.max(brace), np.max(chord))) if (brace.size or chord.size) else 0.0

        row: Dict[str, Any] = {
            "JOINT": _safe_str(r.get("joint")),
            "CHD_A": _safe_str(r.get("chd_a")),
            "CHD_B": _safe_str(r.get("chd_b")),
            "BRACE": _safe_str(r.get("brc")),
            "MEMBER": f"{_safe_str(r.get('chd_a'))}-{_safe_str(r.get('brc'))}",
            "GRUP": r.get("grup"),
            "Dmax_percent": dmax,
            "Dmax": dmax / 100.0,
        }
        for i, p in enumerate(FTG_POS):
            row[f"BRACE_{p}"] = float(brace[i])
            row[f"CHORD_{p}"] = float(chord[i])
            # Keep legacy merged columns used by current risk pipeline.
            row[p] = float(merged[i])
        rows.append(row)

    out_cols = [
        "JOINT",
        "CHD_A",
        "CHD_B",
        "BRACE",
        *brace_cols,
        *chord_cols,
        "MEMBER",
        "GRUP",
        *FTG_POS,
        "Dmax_percent",
        "Dmax",
    ]
    return pd.DataFrame(rows, columns=out_cols)


# =========================
# Module 2: 疲劳分析输入模型解释
# Source: ftginp.*
# Output: default factor, joint override factor, selector mapping, JSLC gaps
# =========================

def parse_ftginp_ringmember(path: str | Path) -> Dict[str, Any]:
    """
    Strictly mirror Ringmember extraction from one ftginp file:
    - default factor from first FTOPT Mid(22,7)
    - factor overrides from JNTOVR Mid(8,4) / Mid(38,6)
    - selector keys from CONRST/CONSCF as joint+brace
    - stop at FTCASE
    """
    p = Path(path)
    if not p.exists():
        raise FileNotFoundError(f"ftginp not found: {p}")

    default_factor = 1.0
    factor_map: Dict[str, float] = {}
    selectors: Set[str] = set()
    jslc_joints: Set[str] = set()

    inset = True
    jslc_turn = 0
    with p.open("r", encoding="utf-8", errors="ignore") as f:
        for raw_line in f:
            line = raw_line.rstrip("\r\n")
            head6 = _fw_sub(line, 1, 6)
            head4 = _fw_sub(line, 1, 4)

            if inset and _fw_sub(line, 1, 5) == "FTOPT":
                v = _parse_float_or_none(_fw_sub(line, 22, 7))
                if v is not None and float(v) != 0.0:
                    default_factor = float(v)
                inset = False
                continue

            if head6 == "JNTOVR":
                jid = _fw_sub(line, 8, 4).strip()
                fac = _parse_float_or_none(_fw_sub(line, 38, 6))
                if jid != "":
                    if fac is None or float(fac) == 0.0:
                        fac = float(default_factor)
                    factor_map[jid] = float(fac)
                continue

            if head6 == "CONRST":
                jt = _fw_sub(line, 16, 4).strip()
                a = _fw_sub(line, 8, 4).strip()
                b = _fw_sub(line, 12, 4).strip()
                brace = a if (a != "" and a != jt) else b
                if jt != "" and brace != "":
                    selectors.add(jt + brace)
                continue

            if head6 == "CONSCF":
                jt = _fw_sub(line, 16, 4).strip()
                brace = _fw_sub(line, 12, 4).strip()
                if jt != "" and brace != "":
                    selectors.add(jt + brace)
                continue

            if head4 == "JSLC":
                # Strictly mirror VBA Ringmember:
                # Jt = Mid(line, 7 + trun*4, 4), and trun is not reset in-loop.
                for _ in range(100):
                    jt = _fw_sub(line, 7 + jslc_turn * 4, 4).strip()
                    if jt == "":
                        break
                    jslc_joints.add(jt)
                    jslc_turn += 1
                continue

            if head6 == "FTCASE":
                break

    return {
        "file_path": str(p),
        "default_factor": float(default_factor),
        "factor_map": factor_map,
        "selectors": selectors,
        "jslc_joints": jslc_joints,
    }


def build_fatigue_merge_cfg_from_ftginp(paths: Sequence[str | Path]) -> List[Dict[str, Any]]:
    cfg_list: List[Dict[str, Any]] = []
    for p in _ensure_path_list(paths):
        cfg_list.append(parse_ftginp_ringmember(p))
    return cfg_list


def _show_warning_popup(title: str, message: str) -> None:
    """
    Best-effort popup message (Windows). No-op when GUI is unavailable.
    """
    try:
        import ctypes

        MB_OK = 0x00000000
        MB_ICONWARNING = 0x00000030
        MB_SYSTEMMODAL = 0x00001000
        ctypes.windll.user32.MessageBoxW(0, message, title, MB_OK | MB_ICONWARNING | MB_SYSTEMMODAL)
    except Exception:
        return


def _pending_manual_fill_rows(audit_df: pd.DataFrame) -> pd.DataFrame:
    if audit_df is None or audit_df.empty:
        return pd.DataFrame(columns=FATIGUE_AUDIT_COLUMNS)
    if "Action" not in audit_df.columns:
        return pd.DataFrame(columns=FATIGUE_AUDIT_COLUMNS)
    pending = audit_df[audit_df["Action"].astype(str) == "manual_fill_needed"].copy()
    if pending.empty:
        return pending
    # Keep VBA-like order by file index then joint natural order.
    if "FileIndex" in pending.columns:
        pending["FileIndex"] = pd.to_numeric(pending["FileIndex"], errors="coerce")
    if "Joint" in pending.columns:
        pending["Joint"] = pending["Joint"].astype(str)
    if "FileIndex" in pending.columns and "Joint" in pending.columns:
        pending = pending.sort_values(by=["FileIndex", "Joint"], kind="stable")
    elif "Joint" in pending.columns:
        pending = pending.sort_values(by=["Joint"], kind="stable")
    return pending


def _write_manual_fill_csv(audit_df: pd.DataFrame, out_xlsx: str | Path) -> Optional[Path]:
    pending = _pending_manual_fill_rows(audit_df)
    if pending.empty:
        return None
    out_base = Path(out_xlsx).resolve()
    csv_path = out_base.with_name(f"{out_base.stem}.manual_fill.csv")
    export_cols = [c for c in ["FileIndex", "FilePath", "Joint", "ManualBrace", "Remark"] if c in pending.columns]
    pending.loc[:, export_cols].to_csv(csv_path, index=False, encoding="utf-8-sig")
    return csv_path


def _build_manual_fill_warning(audit_df: pd.DataFrame, csv_path: Optional[Path]) -> str:
    pending = _pending_manual_fill_rows(audit_df)
    total = int(len(pending))
    if total <= 0:
        return "未发现需要补全的数据。"
    # Keep popup concise: only action hint, no long details.
    return (
        f"检测到 {total} 条缺失数据。\n"
        "请补全后继续运行。\n"
        "如已启用交互补全，将弹出输入框填写 ManualBrace。"
    )


def _collect_manual_overrides_gui(audit_df: pd.DataFrame) -> Optional[Dict[Tuple[int, str], str]]:
    """
    GUI popup input (tkinter askstring) for ManualBrace.
    Returns:
      - dict: GUI ran (possibly empty if user skipped all)
      - None: GUI unavailable/error, caller may fallback to terminal input
    """
    pending = _pending_manual_fill_rows(audit_df)
    if pending.empty:
        return {}

    try:
        import tkinter as tk
        from tkinter import simpledialog
    except Exception:
        return None

    try:
        root = tk.Tk()
        root.withdraw()
        try:
            root.attributes("-topmost", True)
        except Exception:
            pass
    except Exception:
        return None

    overrides: Dict[Tuple[int, str], str] = {}
    total = int(len(pending))
    try:
        for idx, (_, rr) in enumerate(pending.iterrows(), start=1):
            fi = _as_float(rr.get("FileIndex"))
            if fi is None:
                continue
            file_index = int(fi)
            joint = _safe_str(rr.get("Joint"))
            file_path = _safe_str(rr.get("FilePath"))
            if joint == "":
                continue

            prompt = (
                f"[{idx}/{total}] FileIndex={file_index}\n"
                f"Joint={joint}\n"
                f"Case={Path(file_path).parent.name}\n\n"
                "请输入 ManualBrace（4字符，留空=跳过，输入 q=结束）："
            )
            val = simpledialog.askstring("疲劳输入检查", prompt, parent=root)
            if val is None:
                # User closed dialog: stop asking, continue pipeline with collected values.
                break
            v = val.strip()
            if v == "":
                continue
            if v.lower() == "q":
                break
            overrides[(file_index, joint)] = v[:4]
    finally:
        try:
            root.destroy()
        except Exception:
            pass

    print(f"[INFO] GUI交互补全完成，录入 {len(overrides)} 条 ManualBrace。")
    return overrides


def _collect_manual_overrides_interactive(audit_df: pd.DataFrame) -> Dict[Tuple[int, str], str]:
    """
    Interactive terminal input for manual selector completion.
    Returns {(FileIndex, Joint): Brace}.
    """
    overrides: Dict[Tuple[int, str], str] = {}
    pending = _pending_manual_fill_rows(audit_df)
    if pending.empty:
        return overrides

    # Non-interactive environment (e.g. batch runner / CI): skip prompt safely.
    if not sys.stdin or not sys.stdin.isatty():
        print("[WARN] interactive manual fill skipped: stdin is not a tty.")
        return overrides

    print("")
    print("=== 手工补全 JSLC 节点撑杆（VBA Ringmember 红色标记等效流程）===")
    print("输入规则：")
    print("  - 输入 4 字符 Brace（例如 306L）并回车 -> 应用")
    print("  - 直接回车 -> 跳过当前项")
    print("  - 输入 q -> 结束输入，继续后续计算")

    total = int(len(pending))
    for idx, (_, rr) in enumerate(pending.iterrows(), start=1):
        fi = _as_float(rr.get("FileIndex"))
        if fi is None:
            continue
        file_index = int(fi)
        joint = _safe_str(rr.get("Joint"))
        file_path = _safe_str(rr.get("FilePath"))
        if joint == "":
            continue

        while True:
            prompt = (
                f"[{idx}/{total}] FileIndex={file_index}, Joint={joint}, "
                f"Case={Path(file_path).parent.name} -> ManualBrace: "
            )
            val = input(prompt).strip()
            if val == "":
                break
            if val.lower() == "q":
                print("[INFO] 用户结束手工补全输入，继续后续流程。")
                return overrides
            brace = val[:4]
            if brace == "":
                break
            overrides[(file_index, joint)] = brace
            break

    print(f"[INFO] 交互补全完成，录入 {len(overrides)} 条 ManualBrace。")
    return overrides


def build_ringmember_manual_fill_audit(
    merge_cfg: Sequence[Dict[str, Any]],
) -> pd.DataFrame:
    """
    Mirror VBA Ringmember JSLC check:
    - JSLC listed joint must already exist in CONRST/CONSCF extracted joint set.
    - Missing ones are "manual fill needed" (VBA marks red and pops up message).
    """
    rows: List[Dict[str, Any]] = []
    for idx, cfg in enumerate(merge_cfg, start=1):
        selectors = {_safe_str(x) for x in (cfg.get("selectors") or set()) if _safe_str(x)}
        known_joints: Set[str] = set()
        for s in selectors:
            j, _ = _split_selector_key(s)
            if j:
                known_joints.add(j)

        jslc_joints = {_safe_str(x) for x in (cfg.get("jslc_joints") or set()) if _safe_str(x)}
        missing_joints = [j for j in jslc_joints if j not in known_joints]
        for joint in sorted(missing_joints, key=_natural_sort_key):
            rows.append(
                {
                    "FileIndex": idx,
                    "FilePath": _safe_str(cfg.get("file_path")),
                    "Joint": joint,
                    "Action": "manual_fill_needed",
                    "Brace": "",
                    "Score": np.nan,
                    "CandidateCount": 0,
                    "Remark": "jslc_joint_not_in_conrst_conscf",
                    "ManualBrace": "",
                    "Applied": "",
                }
            )

    return pd.DataFrame(rows, columns=FATIGUE_AUDIT_COLUMNS)


def load_manual_selector_overrides_from_workbook(
    manual_fill_workbook: str | Path | None,
) -> Dict[Tuple[int, str], str]:
    """
    Read user-filled manual brace overrides from workbook sheet:
    - sheet name: FatigueSelectorAudit
    - required keys: FileIndex, Joint
    - brace input column priority: ManualBrace > Brace
    """
    overrides: Dict[Tuple[int, str], str] = {}
    if manual_fill_workbook is None:
        return overrides
    p = Path(manual_fill_workbook)
    if not p.exists():
        print(f"[WARN] manual fill workbook not found: {p}")
        return overrides

    try:
        wb = openpyxl.load_workbook(p, data_only=True, read_only=True)
    except Exception as e:
        print(f"[WARN] failed to open manual fill workbook: {p} ({e})")
        return overrides

    if "FatigueSelectorAudit" not in wb.sheetnames:
        print(f"[WARN] manual fill workbook has no sheet 'FatigueSelectorAudit': {p}")
        return overrides

    ws = wb["FatigueSelectorAudit"]
    header: Dict[str, int] = {}
    for c in range(1, ws.max_column + 1):
        name = _safe_str(ws.cell(1, c).value)
        if name != "":
            header[name] = c

    if "FileIndex" not in header or "Joint" not in header:
        print(f"[WARN] manual fill sheet missing required columns FileIndex/Joint: {p}")
        return overrides

    col_file = header["FileIndex"]
    col_joint = header["Joint"]
    col_manual = header.get("ManualBrace")
    col_brace = header.get("Brace")

    for r in range(2, ws.max_row + 1):
        fi_v = ws.cell(r, col_file).value
        fi_f = _as_float(fi_v)
        if fi_f is None:
            continue
        file_index = int(fi_f)
        joint = _safe_str(ws.cell(r, col_joint).value)
        if joint == "":
            continue

        manual_brace = ""
        if col_manual is not None:
            manual_brace = _safe_str(ws.cell(r, col_manual).value)
        if manual_brace == "" and col_brace is not None:
            manual_brace = _safe_str(ws.cell(r, col_brace).value)
        if manual_brace == "":
            continue
        manual_brace = manual_brace[:4]
        overrides[(file_index, joint)] = manual_brace

    print(f"[INFO] manual selector overrides loaded: {len(overrides)} from {p}")
    return overrides


def apply_manual_selector_overrides(
    merge_cfg: Sequence[Dict[str, Any]],
    manual_overrides: Dict[Tuple[int, str], str],
) -> Tuple[List[Dict[str, Any]], pd.DataFrame]:
    """
    Apply manual overrides by adding selector key joint+brace into each file cfg.
    Returns updated cfg list and an audit dataframe for applied rows.
    """
    cfg_list = [dict(c or {}) for c in merge_cfg]
    applied_rows: List[Dict[str, Any]] = []
    if not cfg_list or not manual_overrides:
        return cfg_list, pd.DataFrame(columns=FATIGUE_AUDIT_COLUMNS)

    for idx, cfg in enumerate(cfg_list, start=1):
        selectors = {_safe_str(x) for x in (cfg.get("selectors") or set()) if _safe_str(x)}
        cfg["selectors"] = selectors
        file_path = _safe_str(cfg.get("file_path"))
        for (fi, joint), brace in manual_overrides.items():
            if fi != idx:
                continue
            j = _safe_str(joint)
            b = _safe_str(brace)[:4]
            if j == "" or b == "":
                continue
            selectors.add(j + b)
            applied_rows.append(
                {
                    "FileIndex": idx,
                    "FilePath": file_path,
                    "Joint": j,
                    "Action": "manual_applied",
                    "Brace": b,
                    "Score": np.nan,
                    "CandidateCount": 0,
                    "Remark": "user_manual_override",
                    "ManualBrace": b,
                    "Applied": "Y",
                }
            )

    applied_df = pd.DataFrame(applied_rows, columns=FATIGUE_AUDIT_COLUMNS)
    return cfg_list, applied_df


def _split_selector_key(key: str) -> Tuple[str, str]:
    k = _safe_str(key)
    if len(k) >= 8:
        return k[:4].strip(), k[4:8].strip()
    half = len(k) // 2
    return k[:half].strip(), k[half:].strip()


def enrich_fatigue_merge_cfg_by_topology(
    merge_cfg: Sequence[Dict[str, Any]],
    members_df: pd.DataFrame,
) -> Tuple[List[Dict[str, Any]], pd.DataFrame]:
    """
    Keep VBA selectors as primary source, and use topology to fill JSLC-missing
    selectors (the red-manual-fill cases in Ringmember).
    """
    cfg_list = [dict(c or {}) for c in merge_cfg]
    if not cfg_list:
        return [], pd.DataFrame(columns=["FileIndex", "FilePath", "Joint", "Action", "Brace", "Score", "CandidateCount", "Remark"])

    # Build undirected adjacency from members.
    adj: Dict[str, List[Dict[str, Any]]] = {}
    if not members_df.empty:
        for _, r in members_df.iterrows():
            a = _safe_str(r.get("A"))
            b = _safe_str(r.get("B"))
            if a == "" or b == "":
                continue
            od = _as_float(r.get("OD"))
            mtype = _safe_str(r.get("MemberType"))
            z1 = _as_float(r.get("Z1"))
            z2 = _as_float(r.get("Z2"))
            e_ab = {"joint": a, "brace": b, "od": od, "member_type": mtype, "z_self": z1, "z_other": z2}
            e_ba = {"joint": b, "brace": a, "od": od, "member_type": mtype, "z_self": z2, "z_other": z1}
            adj.setdefault(a, []).append(e_ab)
            adj.setdefault(b, []).append(e_ba)

    audit_rows: List[Dict[str, Any]] = []

    for idx, cfg in enumerate(cfg_list, start=1):
        selectors = {_safe_str(x) for x in (cfg.get("selectors") or set()) if _safe_str(x)}
        cfg["selectors"] = selectors

        # Existing selector joints from CONRST/CONSCF.
        known_joint_to_brace: Dict[str, Set[str]] = {}
        known_ods: List[float] = []
        for s in selectors:
            j, b = _split_selector_key(s)
            if j == "" or b == "":
                continue
            known_joint_to_brace.setdefault(j, set()).add(b)
            for e in adj.get(j, []):
                if _safe_str(e.get("brace")) == b:
                    odv = _as_float(e.get("od"))
                    if odv is not None:
                        known_ods.append(float(odv))
                    break
        target_od = float(np.median(known_ods)) if known_ods else None

        jslc_joints = {_safe_str(x) for x in (cfg.get("jslc_joints") or set()) if _safe_str(x)}
        for joint in sorted(jslc_joints, key=_natural_sort_key):
            if joint in known_joint_to_brace:
                continue

            cand_edges = adj.get(joint, [])
            if not cand_edges:
                audit_rows.append(
                    {
                        "FileIndex": idx,
                        "FilePath": _safe_str(cfg.get("file_path")),
                        "Joint": joint,
                        "Action": "unresolved",
                        "Brace": "",
                        "Score": np.nan,
                        "CandidateCount": 0,
                        "Remark": "joint_not_in_model_topology",
                    }
                )
                continue

            scored: List[Tuple[float, float, str]] = []
            for e in cand_edges:
                brace = _safe_str(e.get("brace"))
                if brace == "":
                    continue
                score = 0.0
                mtype = _safe_str(e.get("member_type"))
                if mtype == "X-Brace":
                    score += 4.0
                elif mtype == "LEG":
                    score -= 1.5

                if brace.endswith("X"):
                    score += 1.5

                od = _as_float(e.get("od"))
                od_num = float(od) if od is not None else float("nan")
                if np.isfinite(od_num):
                    if od_num >= 100.0:
                        score += 0.5
                    if target_od is not None and target_od > 0:
                        rel = abs(od_num - target_od) / target_od
                        score += max(0.0, 2.0 - 4.0 * rel)

                # Slight preference for non-horizontal members when available.
                z_self = _as_float(e.get("z_self"))
                z_other = _as_float(e.get("z_other"))
                if z_self is not None and z_other is not None and abs(float(z_other) - float(z_self)) > 1e-6:
                    score += 0.2

                scored.append((score, od_num if np.isfinite(od_num) else -1.0, brace))

            if not scored:
                audit_rows.append(
                    {
                        "FileIndex": idx,
                        "FilePath": _safe_str(cfg.get("file_path")),
                        "Joint": joint,
                        "Action": "unresolved",
                        "Brace": "",
                        "Score": np.nan,
                        "CandidateCount": 0,
                        "Remark": "no_valid_neighbors",
                    }
                )
                continue

            scored.sort(key=lambda x: (-x[0], -x[1], _natural_sort_key(x[2])))
            best_score, _, best_brace = scored[0]

            # Conservative threshold to avoid over-inference.
            if best_score < 1.0:
                audit_rows.append(
                    {
                        "FileIndex": idx,
                        "FilePath": _safe_str(cfg.get("file_path")),
                        "Joint": joint,
                        "Action": "unresolved",
                        "Brace": best_brace,
                        "Score": best_score,
                        "CandidateCount": len(scored),
                        "Remark": "low_confidence",
                    }
                )
                continue

            selectors.add(joint + best_brace)
            cfg["selectors"] = selectors
            audit_rows.append(
                {
                    "FileIndex": idx,
                    "FilePath": _safe_str(cfg.get("file_path")),
                    "Joint": joint,
                    "Action": "inferred",
                    "Brace": best_brace,
                    "Score": best_score,
                    "CandidateCount": len(scored),
                    "Remark": "topology_inferred_from_members",
                }
            )

    audit_df = pd.DataFrame(
        audit_rows,
        columns=["FileIndex", "FilePath", "Joint", "Action", "Brace", "Score", "CandidateCount", "Remark"],
    )
    return cfg_list, audit_df


def parse_ftglst_detail(path: str | Path) -> pd.DataFrame:
    # Single-file path keeps VBA pair extraction without cross-file replacement.
    return _pair_rows_to_fatigue_df(_parse_ftglst_pair_rows(path))


def parse_ftglst_details(
    paths: Sequence[str | Path],
    merge_cfg: Optional[Sequence[Dict[str, Any]]] = None,
) -> pd.DataFrame:
    path_list = _ensure_path_list(paths)
    if not path_list:
        return pd.DataFrame(
            columns=[
                "JOINT",
                "CHD_A",
                "CHD_B",
                "BRACE",
                *[f"BRACE_{p}" for p in FTG_POS],
                *[f"CHORD_{p}" for p in FTG_POS],
                "MEMBER",
                "GRUP",
                *FTG_POS,
                "Dmax_percent",
                "Dmax",
            ]
        )

    # Strict VBA-like fallback when no merge config:
    # keep only first file as baseline (no cross-file accumulation).
    if merge_cfg is None:
        return _pair_rows_to_fatigue_df(_parse_ftglst_pair_rows(path_list[0]))

    cfg_list = list(merge_cfg)

    def _cfg_at(idx: int) -> Tuple[float, Dict[str, float], Set[str]]:
        if idx < len(cfg_list):
            cfg = cfg_list[idx] or {}
        else:
            cfg = {}
        default_factor = _as_float(cfg.get("default_factor"))
        if default_factor is None or float(default_factor) == 0:
            default_factor = 1.0
        factor_map_raw = cfg.get("factor_map") or {}
        factor_map_clean: Dict[str, float] = {}
        for k, v in factor_map_raw.items():
            sk = _safe_str(k)
            fv = _as_float(v)
            if sk and fv is not None and float(fv) != 0:
                factor_map_clean[sk] = float(fv)
        selectors = {_safe_str(x) for x in (cfg.get("selectors") or []) if _safe_str(x)}
        return float(default_factor), factor_map_clean, selectors

    # Strict VBA state machine:
    # - turn 1 initializes ARR1 baseline rows
    # - turn>1 collects Dic2 replacement rows and overlays only existing ARR1 keys.
    base_rows: List[Dict[str, Any]] = []
    for i, p in enumerate(path_list, start=1):
        fi, mi, selectors = _cfg_at(i - 1)
        # Align with the actual workbook button-run result:
        # the first fatigue result file is scaled by its default factor,
        # while JNTOVR-based overrides take effect only in later replacement turns.
        if i == 1:
            mi = {}
        turn_rows, replace_map = _parse_ftglst_vba_turn_rows(
            p,
            turn=i,
            default_factor=fi,
            factor_map=mi,
            selectors=selectors,
        )
        if i == 1:
            base_rows = list(turn_rows)
            continue
        if not replace_map or not base_rows:
            continue
        for j, base_row in enumerate(base_rows):
            k = str(base_row.get("full_key", ""))
            rep = replace_map.get(k)
            if rep is None:
                continue
            base_rows[j]["brace_damage"] = np.array(rep.get("brace_damage", np.zeros(8, dtype=float)), dtype=float)
            base_rows[j]["chord_damage"] = np.array(rep.get("chord_damage", np.zeros(8, dtype=float)), dtype=float)

    return _pair_rows_to_fatigue_df(base_rows)


# =========================
# Read matrices & parameters from xlsm / json
# =========================

@dataclass
class RiskMatrix:
    prob_upper: List[float]                 # row 36 C:G (legacy)
    prob_thresholds: List[float]            # row 37 C:G (VBA MATCH thresholds)
    possibility_values: List[int]           # row 39 C:G
    risk_map: Dict[Tuple[int,int], str]     # (consequence 5..1, possibility 5..1) -> '一'..'五'

@dataclass
class RiskMatrixPack:
    rm: RiskMatrix
    A: float
    B: float
    c_delta: float
    c_a: float
    c_b: float
    m: float

def normalize_level(s: str) -> str:
    s = s.strip()
    s = s.replace("Level", "").replace("level", "").strip()
    return s

_RE_DIST = re.compile(r"(\d+)\s*%?\s*level\s*(I{1,3}|IV)", re.I)

def parse_distribution_from_note(note: str, times: Dict[str,str]) -> Dict[str, Dict[str,float]]:
    dist_by_time: Dict[str, Dict[str,float]] = {}
    if note:
        matches = _RE_DIST.findall(note)
        if matches:
            ratios = {}
            total = 0
            for pct, lvl in matches:
                pct_i = int(pct)
                total += pct_i
                ratios[normalize_level(lvl)] = pct_i / 100.0
            if total and abs(total-100) > 1e-6:
                ratios = {k: v/(total/100.0) for k,v in ratios.items()}
            for t in times:
                dist_by_time[t] = ratios
            return dist_by_time

    for t, v in times.items():
        if "/" in v:
            parts = [normalize_level(x) for x in v.split("/")]
            if len(parts) == 2:
                dist_by_time[t] = {parts[0]: 0.8, parts[1]: 0.2}
            else:
                dist_by_time[t] = {normalize_level(v): 1.0}
        else:
            dist_by_time[t] = {normalize_level(v): 1.0}
    return dist_by_time

def _risk_pack_to_plain(pack: RiskMatrixPack) -> Dict[str, Any]:
    risk_rows = []
    for (consequence, possibility), grade in sorted(pack.rm.risk_map.items()):
        risk_rows.append(
            {
                "consequence": int(consequence),
                "possibility": int(possibility),
                "grade": str(grade),
            }
        )
    return {
        "rm": {
            "prob_upper": [float(x) for x in pack.rm.prob_upper],
            "prob_thresholds": [float(x) for x in pack.rm.prob_thresholds],
            "possibility_values": [int(x) for x in pack.rm.possibility_values],
            "risk_rows": risk_rows,
        },
        "A": float(pack.A),
        "B": float(pack.B),
        "c_delta": float(pack.c_delta),
        "c_a": float(pack.c_a),
        "c_b": float(pack.c_b),
        "m": float(pack.m),
    }


def _risk_pack_from_plain(raw: Dict[str, Any]) -> RiskMatrixPack:
    rm_raw = raw["rm"]
    risk_map: Dict[Tuple[int, int], str] = {}
    for row in rm_raw.get("risk_rows", []):
        consequence = int(row["consequence"])
        possibility = int(row["possibility"])
        risk_map[(consequence, possibility)] = str(row["grade"])
    rm = RiskMatrix(
        prob_upper=[float(x) for x in rm_raw.get("prob_upper", [])],
        prob_thresholds=[float(x) for x in rm_raw.get("prob_thresholds", [])],
        possibility_values=[int(x) for x in rm_raw.get("possibility_values", [])],
        risk_map=risk_map,
    )
    return RiskMatrixPack(
        rm=rm,
        A=float(raw["A"]),
        B=float(raw["B"]),
        c_delta=float(raw["c_delta"]),
        c_a=float(raw["c_a"]),
        c_b=float(raw["c_b"]),
        m=float(raw["m"]),
    )


def _policy_to_plain(policy: Dict[str, Any]) -> Dict[str, Any]:
    out: Dict[str, Any] = {}
    for risk, spec in policy.items():
        times = {str(k): _safe_str(v) for k, v in dict(spec.get("times", {})).items()}
        dist = {
            str(t): {str(level): float(prob) for level, prob in dict(levels).items()}
            for t, levels in dict(spec.get("dist", {})).items()
        }
        out[str(risk)] = {
            "cycle": spec.get("cycle"),
            "times": times,
            "dist": dist,
            "note": _safe_str(spec.get("note")),
        }
    return out


def _policy_from_plain(policy_raw: Dict[str, Any]) -> Dict[str, Any]:
    out: Dict[str, Any] = {}
    for risk, spec in dict(policy_raw).items():
        out[str(risk)] = {
            "cycle": spec.get("cycle"),
            "times": {str(k): _safe_str(v) for k, v in dict(spec.get("times", {})).items()},
            "dist": {
                str(t): {str(level): float(prob) for level, prob in dict(levels).items()}
                for t, levels in dict(spec.get("dist", {})).items()
            },
            "note": _safe_str(spec.get("note")),
        }
    return out


def _fatigue_merge_to_plain(fatigue_merge: Sequence[Dict[str, Any]]) -> List[Dict[str, Any]]:
    out: List[Dict[str, Any]] = []
    for item in fatigue_merge:
        out.append(
            {
                "file_path": _safe_str(item.get("file_path")),
                "default_factor": float(item.get("default_factor", 1.0)),
                "factor_map": {
                    str(k): float(v) for k, v in dict(item.get("factor_map", {})).items()
                },
                "selectors": sorted(str(x) for x in set(item.get("selectors", set()))),
            }
        )
    return out


def _fatigue_merge_from_plain(fatigue_merge_raw: Sequence[Dict[str, Any]]) -> List[Dict[str, Any]]:
    out: List[Dict[str, Any]] = []
    for item in fatigue_merge_raw:
        out.append(
            {
                "file_path": _safe_str(item.get("file_path")),
                "default_factor": float(item.get("default_factor", 1.0)),
                "factor_map": {
                    str(k): float(v) for k, v in dict(item.get("factor_map", {})).items()
                },
                "selectors": set(str(x) for x in item.get("selectors", [])),
            }
        )
    return out


def _cfg_to_jsonable(cfg: Dict[str, Any]) -> Dict[str, Any]:
    return {
        "model_path": _safe_str(cfg.get("model_path")),
        "region": _safe_str(cfg.get("region")),
        "global_level_tag": _safe_str(cfg.get("global_level_tag")),
        "collapse_a_const": _as_float(cfg.get("collapse_a_const")),
        "collapse_b_const": _as_float(cfg.get("collapse_b_const")),
        "min_leg_od": float(cfg.get("min_leg_od", 0.0)),
        "x_angle_deviation": _as_float(cfg.get("x_angle_deviation")),
        "wp_z": float(cfg.get("wp_z", 0.0)),
        "no_legs": int(cfg.get("no_legs", 0)),
        "work_points": [[float(x), float(y)] for x, y in cfg.get("work_points", [])],
        "fatigue_merge": _fatigue_merge_to_plain(cfg.get("fatigue_merge", [])),
        "risk_pack": _risk_pack_to_plain(cfg["risk_pack"]),
        "policy_strict": _policy_to_plain(cfg.get("policy_strict", {})),
        "policy_loose": _policy_to_plain(cfg.get("policy_loose", {})),
        "design_life": float(cfg.get("design_life", 26.0)),
        "served_years": float(cfg.get("served_years", 1.0)),
    }


def export_calc_source_json_from_xlsm(source_xlsm: str | Path, out_json: str | Path) -> Path:
    cfg = load_from_template(source_xlsm)
    out_path = Path(out_json)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    out_path.write_text(
        json.dumps(_cfg_to_jsonable(cfg), ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    return out_path


def load_from_params_json(params_json: str | Path) -> Dict[str, Any]:
    p = Path(params_json)
    if not p.exists():
        raise FileNotFoundError(f"params json not found: {p}")
    raw = json.loads(p.read_text(encoding="utf-8"))
    return {
        "template": str(p),
        "model_path": _safe_str(raw.get("model_path")),
        "region": _safe_str(raw.get("region")),
        "global_level_tag": _safe_str(raw.get("global_level_tag")),
        "collapse_a_const": _as_float(raw.get("collapse_a_const")),
        "collapse_b_const": _as_float(raw.get("collapse_b_const")),
        "min_leg_od": float(raw.get("min_leg_od", 0.0)),
        "x_angle_deviation": _as_float(raw.get("x_angle_deviation")),
        "wp_z": float(raw.get("wp_z", 0.0)),
        "no_legs": int(raw.get("no_legs", 0)),
        "work_points": [(float(x), float(y)) for x, y in raw.get("work_points", [])],
        "fatigue_merge": _fatigue_merge_from_plain(raw.get("fatigue_merge", [])),
        "risk_pack": _risk_pack_from_plain(raw["risk_pack"]),
        "policy_strict": _policy_from_plain(raw.get("policy_strict", {})),
        "policy_loose": _policy_from_plain(raw.get("policy_loose", {})),
        "design_life": float(raw.get("design_life", 26.0)),
        "served_years": float(raw.get("served_years", 1.0)),
    }


def load_from_template(template_xlsm: str | Path) -> Dict[str, Any]:
    p = Path(template_xlsm)
    if not p.exists():
        raise FileNotFoundError(f"template xlsm not found: {p}")
    wb = openpyxl.load_workbook(p, data_only=True, keep_vba=True)

    ctrl = wb["控制页面"]
    region = _safe_str(_cell(ctrl,"B45"))
    model_path = _safe_str(_cell(ctrl,"B2"))
    global_level_tag = _safe_str(_cell(ctrl, "B44"))
    collapse_a_const = _as_float(_cell(ctrl, "B46"))
    collapse_b_const = _as_float(_cell(ctrl, "B47"))
    served_years = _as_float(_cell(ctrl, "B48"))
    design_life = _as_float(_cell(ctrl, "B49"))

    x_angle_deviation = _as_float(_cell(ctrl, "B6"))
    if x_angle_deviation is None:
        x_angle_deviation = 15.0
    min_leg_od = float(_cell(ctrl,"B7"))
    wp_z = float(_cell(ctrl,"B8"))
    no_legs = int(_cell(ctrl,"B9"))

    wps=[]
    for r in range(11, 19):
        idx = ctrl.cell(r,1).value
        x = ctrl.cell(r,2).value
        y = ctrl.cell(r,3).value
        if idx is None or x is None or y is None:
            continue
        wps.append((float(x), float(y)))

    # VBA Sheet8_FatiguePickup merge configuration:
    # files from B39:H39, default factors from B38:H38, and per-file mapping
    # columns driven by cl = 12 + (trun-1)*4.
    fatigue_merge: List[Dict[str, Any]] = []
    trun = 0
    for file_col in range(2, 9):  # B..H
        ftg_path = _safe_str(ctrl.cell(39, file_col).value)
        if ftg_path == "":
            continue
        trun += 1
        cl = 12 + (trun - 1) * 4

        default_factor = _as_float(ctrl.cell(38, 1 + trun).value)
        if default_factor is None or float(default_factor) == 0:
            default_factor = 1.0

        factor_map: Dict[str, float] = {}
        for r in range(38, 301):
            jid = _safe_str(ctrl.cell(r, cl).value)
            if jid == "":
                break
            fac = _as_float(ctrl.cell(r, cl + 1).value)
            if fac is None or float(fac) == 0:
                fac = float(default_factor)
            factor_map[jid] = float(fac)

        selectors: Set[str] = set()
        if trun > 1:
            for r in range(38, 301):
                a = _safe_str(ctrl.cell(r, cl + 2).value)
                if a == "":
                    break
                b = _safe_str(ctrl.cell(r, cl + 3).value)
                selectors.add(a + b)

        fatigue_merge.append(
            {
                "file_path": ftg_path,
                "default_factor": float(default_factor),
                "factor_map": factor_map,
                "selectors": selectors,
            }
        )

    rm_ws = wb["风险评级矩阵"]
    prob_upper = [
        float(rm_ws.cell(36,3).value),
        float(rm_ws.cell(36,4).value),
        float(rm_ws.cell(36,5).value),
        float(rm_ws.cell(36,6).value),
        float(rm_ws.cell(36,7).value),
    ]
    prob_thresholds = [
        float(rm_ws.cell(37, 3).value),
        float(rm_ws.cell(37, 4).value),
        float(rm_ws.cell(37, 5).value),
        float(rm_ws.cell(37, 6).value),
        float(rm_ws.cell(37, 7).value),
    ]
    possibility_values = [
        int(rm_ws.cell(39, 3).value),
        int(rm_ws.cell(39, 4).value),
        int(rm_ws.cell(39, 5).value),
        int(rm_ws.cell(39, 6).value),
        int(rm_ws.cell(39, 7).value),
    ]

    risk_map={}
    for row in range(40,45):
        consequence = int(rm_ws.cell(row,2).value)
        for col in range(3,8):
            possibility = int(rm_ws.cell(39, col).value)
            grade = _safe_str(rm_ws.cell(row,col).value)
            risk_map[(consequence, possibility)] = grade
    rm = RiskMatrix(
        prob_upper=prob_upper,
        prob_thresholds=prob_thresholds,
        possibility_values=possibility_values,
        risk_map=risk_map,
    )

    regions = { _safe_str(rm_ws.cell(50,c).value): c for c in (2,3,4) }
    col = regions.get(region, 2)
    A = float(rm_ws.cell(51,col).value)
    B = float(rm_ws.cell(52,col).value)

    jr = wb["节点失效风险等级"]
    c_delta = float(jr.cell(3,12).value)
    c_a = float(jr.cell(3,13).value)
    c_b = float(jr.cell(3,14).value)
    m_val = float(jr.cell(3,15).value)

    strict = wb["检验策略详细矩阵 (严苛）"]
    loose  = wb["检验策略详细矩阵 (宽松)"]

    def parse_policy(ws) -> Dict[str, Any]:
        policy={}
        for r in range(5,10):
            risk = _safe_str(ws.cell(r,1).value)
            cycle = ws.cell(r,2).value
            note = _safe_str(ws.cell(r,7).value)
            times = {
                "N": _safe_str(ws.cell(r,3).value),
                "N+5": _safe_str(ws.cell(r,4).value),
                "N+10": _safe_str(ws.cell(r,5).value),
                "N+15": _safe_str(ws.cell(r,6).value),
            }
            dist = parse_distribution_from_note(note, times)
            policy[risk] = {"cycle": cycle, "times": times, "dist": dist, "note": note}
        return policy

    return {
        "template": str(p),
        "model_path": model_path,
        "region": region,
        "global_level_tag": global_level_tag,
        "collapse_a_const": collapse_a_const,
        "collapse_b_const": collapse_b_const,
        "min_leg_od": min_leg_od,
        "x_angle_deviation": x_angle_deviation,
        "wp_z": wp_z,
        "no_legs": no_legs,
        "work_points": wps,
        "fatigue_merge": fatigue_merge,
        "risk_pack": RiskMatrixPack(rm=rm, A=A, B=B, c_delta=c_delta, c_a=c_a, c_b=c_b, m=m_val),
        "policy_strict": parse_policy(strict),
        "policy_loose": parse_policy(loose),
        "design_life": float(design_life) if design_life is not None else 26.0,
        "served_years": float(served_years) if served_years is not None else 1.0,
    }


# =========================
# Modules 5-8 shared formulas
# 5: consequence level lookup
# 6: fatigue failure probability / level
# 7: collapse failure probability / level
# 8: overall risk grade matrix lookup
# =========================

def collapse_pf(A: float, B: float, Rm: float, VR: float) -> float:
    term1 = math.exp((A - Rm)/B + (VR**2)*(Rm**2)/(2*(B**2)))
    # VBA formula uses NORM.DIST(..., FALSE): standard normal PDF (not CDF).
    term2 = 1.0 - _norm_pdf(VR * Rm / B - 1.0 / VR)
    return float(term1 * term2)

def fatigue_ctf(c_delta: float, c_a: float, c_b: float, m: float) -> float:
    # VBA formula: ((1+c_delta^2)*(1+c_a^2)*(1+c_b^2)^(m^2)-1)^(1/2)
    return math.sqrt((1 + c_delta**2) * (1 + c_a**2) * ((1 + c_b**2) ** (m**2)) - 1.0)

def fatigue_beta_current(design_life: float, damage: float, served_years: float, ctf: float) -> float:
    if damage <= 0:
        return 99.0
    return math.log(design_life / damage / served_years) / math.sqrt(math.log(1.0 + ctf**2))

def fatigue_beta_forecast(damage_future: float, ctf: float) -> float:
    if damage_future <= 0:
        return 99.0
    return math.log(1.0 / damage_future) / math.sqrt(math.log(1.0 + ctf**2))

def pf_from_beta(beta: float) -> float:
    b = float(beta)
    if not np.isfinite(b):
        return float("nan")
    # Numerically stable tail probability for large beta:
    # Pf = Phi(-beta) = 0.5 * erfc(beta / sqrt(2))
    return float(0.5 * math.erfc(b / math.sqrt(2.0)))

def possibility_level(pf: float, prob_upper: List[float]) -> int:
    for i, ub in enumerate(prob_upper):
        if pf <= ub:
            return 5 - i
    return 1


def possibility_level_vba(pf: float, rm: RiskMatrix) -> Optional[int]:
    """
    VBA equivalent of:
    INDEX(row39, MATCH(pf, row37, 1))
    """
    if pf is None:
        return None
    try:
        x = float(pf)
    except Exception:
        return None
    if not np.isfinite(x):
        return None

    thresholds = list(rm.prob_thresholds)
    levels = list(rm.possibility_values)
    if len(thresholds) == 0 or len(levels) == 0:
        return None
    if len(thresholds) != len(levels):
        n = min(len(thresholds), len(levels))
        thresholds = thresholds[:n]
        levels = levels[:n]
        if n == 0:
            return None

    # MATCH(...,1): largest threshold <= x (ascending thresholds).
    idx = int(np.searchsorted(np.array(thresholds, dtype=float), x, side="right") - 1)
    if idx < 0:
        idx = 0
    if idx >= len(levels):
        idx = len(levels) - 1
    return int(levels[idx])

def risk_grade(consequence_level: int, poss_level: int, rm: RiskMatrix) -> str:
    return rm.risk_map.get((consequence_level, poss_level), "")


# =========================
# Modules 5-8 assembled business outputs
# build_member_risk_vba: consequence + collapse + overall member risk
# build_joint_risk_vba: consequence + fatigue + collapse + overall joint risk
# build_joint_forecast_vba_wide: future fatigue / overall node risk by time node
# =========================

def _global_level_from_tag(tag: str) -> int:
    t = _safe_str(tag).upper()
    if t == "L-1":
        return 0
    if t == "L-2":
        return 1
    if t == "L-3":
        return 2
    return 0


def _local_level_member(member_type: str) -> int:
    t = _safe_str(member_type)
    if t == "LEG":
        return 1
    if t == "X-Brace":
        return 2
    return 3


def _local_level_joint(joint_type: str) -> int:
    t = _safe_str(joint_type)
    if t == "LegJoint":
        return 1
    if t == "X Joint":
        return 2
    return 3


def _min_factor_by_location(collapse_df: pd.DataFrame) -> Dict[str, float]:
    out: Dict[str, float] = {}
    if collapse_df.empty:
        return out
    for _, r in collapse_df.iterrows():
        loc = _safe_str(r.get("LOCATION"))
        f = _as_float(r.get("FACTOR"))
        if loc == "" or f is None:
            continue
        if (loc not in out) or (float(f) < float(out[loc])):
            out[loc] = float(f)
    return out


def _collapse_rsr(collapse_df: pd.DataFrame, collapse_summary_df: pd.DataFrame) -> float:
    vals: List[float] = []
    if not collapse_summary_df.empty and "LastLoadFactor" in collapse_summary_df.columns:
        for v in collapse_summary_df["LastLoadFactor"].tolist():
            fv = _as_float(v)
            if fv is not None:
                vals.append(float(fv))
    if not vals and (not collapse_df.empty) and ("FACTOR" in collapse_df.columns):
        for v in collapse_df["FACTOR"].tolist():
            fv = _as_float(v)
            if fv is not None:
                vals.append(float(fv))
    if not vals:
        return float("nan")
    return float(min(vals))


def build_member_risk_vba(
    members_df: pd.DataFrame,
    collapse_df: pd.DataFrame,
    collapse_summary_df: pd.DataFrame,
    cfg: Dict[str, Any],
    rm: RiskMatrix,
) -> pd.DataFrame:
    """
    VBA equivalent of Sheet1.UpdateMemberRisk.
    Combines:
    - Module 5: member consequence level
    - Module 7: member collapse possibility
    - Module 8: member overall risk grade
    """
    wz = float(cfg["wp_z"])
    global_lv = _global_level_from_tag(_safe_str(cfg.get("global_level_tag")))
    a_const = _as_float(cfg.get("collapse_a_const"))
    b_const = _as_float(cfg.get("collapse_b_const"))
    if a_const is None:
        a_const = 0.0
    if b_const is None:
        b_const = 1.0
    vr = 0.1
    rsr = _collapse_rsr(collapse_df, collapse_summary_df)
    loc_min = _min_factor_by_location(collapse_df)

    rows: List[Dict[str, Any]] = []
    for _, r in members_df.iterrows():
        ja = _safe_str(r.get("A"))
        jb = _safe_str(r.get("B"))
        if ja == "":
            continue
        z2 = _as_float(r.get("Z2"))
        if z2 is None:
            continue
        # Match the actual workbook macro behavior, not the intended semantics.
        # UpdateMemberRisk checks Sheet3 columns 7 and 8, while the generated
        # Members sheet stores Z2 in column 7 and leaves column 8 blank.
        # In practice this means the VBA filter behaves like "Z2 < wz".
        if float(z2) >= wz:
            continue

        member_type = _safe_str(r.get("MemberType")) or "Other"
        consequence = global_lv + _local_level_member(member_type)

        pair_key = f"{ja}-{jb}"
        rm_factor = loc_min.get(pair_key, rsr)
        pf = collapse_pf(float(a_const), float(b_const), float(rm_factor), vr) if np.isfinite(rm_factor) else np.nan
        poss = possibility_level_vba(pf, rm)
        grade = risk_grade(consequence, int(poss), rm) if poss is not None else ""

        rows.append(
            {
                "JointA": ja,
                "JointB": jb,
                "MemberType": member_type,
                "ConsequenceLevel": consequence,
                "A": float(a_const),
                "B": float(b_const),
                "Rm": rm_factor,
                "VR": vr,
                "Pf": pf,
                "CollapsePossLevel": poss,
                "RiskGrade": grade,
            }
        )

    return pd.DataFrame(
        rows,
        columns=[
            "JointA",
            "JointB",
            "MemberType",
            "ConsequenceLevel",
            "A",
            "B",
            "Rm",
            "VR",
            "Pf",
            "CollapsePossLevel",
            "RiskGrade",
        ],
    )


def build_joint_risk_vba(
    joints_df: pd.DataFrame,
    fatigue_df: pd.DataFrame,
    collapse_df: pd.DataFrame,
    collapse_summary_df: pd.DataFrame,
    cfg: Dict[str, Any],
    rm: RiskMatrix,
) -> pd.DataFrame:
    """
    VBA equivalent of Sheet1.UpdateJointRisk.
    Combines:
    - Module 5: joint consequence level
    - Module 6: current fatigue possibility
    - Module 7: joint collapse possibility
    - Module 8: current joint overall risk grade
    """
    wz = float(cfg["wp_z"])
    global_lv = _global_level_from_tag(_safe_str(cfg.get("global_level_tag")))
    a_const = _as_float(cfg.get("collapse_a_const"))
    b_const = _as_float(cfg.get("collapse_b_const"))
    if a_const is None:
        a_const = 0.0
    if b_const is None:
        b_const = 1.0
    vr = 0.1
    rsr = _collapse_rsr(collapse_df, collapse_summary_df)
    loc_min = _min_factor_by_location(collapse_df)

    c_delta = 0.3
    c_a = 0.73
    c_b = 0.3
    m_val = 4.0
    ctf = fatigue_ctf(c_delta, c_a, c_b, m_val)
    served_years = float(cfg.get("served_years", 1.0))
    design_life = float(cfg.get("design_life", 26.0))

    # VBA Dic (Sheet2 A:E): under-water joints only, blank type -> Other.
    joint_type_map: Dict[str, str] = {}
    for _, r in joints_df.iterrows():
        jid = _safe_str(r.get("Joint"))
        if jid == "":
            continue
        z = _as_float(r.get("Z"))
        if z is None or float(z) >= wz:
            continue
        jt = _safe_str(r.get("JointType")) or "Other"
        joint_type_map[jid] = jt

    rows: List[Dict[str, Any]] = []
    if fatigue_df.empty:
        return pd.DataFrame(
            columns=[
                "JoitID",
                "JointType",
                "A",
                "B",
                "Rm",
                "VR",
                "Pf_collapse",
                "CollapsePossLevel",
                "D",
                "CTf",
                "beta_fatigue",
                "Pf_fatigue",
                "FatiguePossLevel",
                "PossLevel",
                "RiskGrade",
                "Brace",
                "ConsequenceLevel",
                "c_delta",
                "c_a",
                "c_b",
                "m",
            ]
        )

    for _, fr in fatigue_df.iterrows():
        jid = _safe_str(fr.get("JOINT"))
        if jid == "" or jid not in joint_type_map:
            continue
        joint_type = joint_type_map[jid]
        consequence = global_lv + _local_level_joint(joint_type)

        brace = _safe_str(fr.get("BRACE"))
        if brace == "":
            member_txt = _safe_str(fr.get("MEMBER"))
            if "-" in member_txt:
                brace = member_txt.split("-")[-1]
            else:
                brace = member_txt

        rm_factor = loc_min.get(jid, rsr)
        pf_c = collapse_pf(float(a_const), float(b_const), float(rm_factor), vr) if np.isfinite(rm_factor) else np.nan
        poss_c = possibility_level_vba(pf_c, rm)

        d_percent = _as_float(fr.get("Dmax_percent"))
        if d_percent is None:
            d_percent = 0.0
        beta_f = fatigue_beta_current(design_life, float(d_percent), served_years, ctf)
        pf_f = pf_from_beta(beta_f)
        poss_f = possibility_level_vba(pf_f, rm)

        if poss_c is None:
            poss_comb = poss_f
        elif poss_f is None:
            poss_comb = poss_c
        else:
            poss_comb = min(int(poss_c), int(poss_f))
        grade = risk_grade(consequence, int(poss_comb), rm) if poss_comb is not None else ""

        rows.append(
            {
                "JoitID": jid,
                "JointType": joint_type,
                "A": float(a_const),
                "B": float(b_const),
                "Rm": rm_factor,
                "VR": vr,
                "Pf_collapse": pf_c,
                "CollapsePossLevel": poss_c,
                "D": float(d_percent),
                "CTf": ctf,
                "beta_fatigue": beta_f,
                "Pf_fatigue": pf_f,
                "FatiguePossLevel": poss_f,
                "PossLevel": poss_comb,
                "RiskGrade": grade,
                "Brace": brace,
                "ConsequenceLevel": consequence,
                "c_delta": c_delta,
                "c_a": c_a,
                "c_b": c_b,
                "m": m_val,
            }
        )

    return pd.DataFrame(
        rows,
        columns=[
            "JoitID",
            "JointType",
            "A",
            "B",
            "Rm",
            "VR",
            "Pf_collapse",
            "CollapsePossLevel",
            "D",
            "CTf",
            "beta_fatigue",
            "Pf_fatigue",
            "FatiguePossLevel",
            "PossLevel",
            "RiskGrade",
            "Brace",
            "ConsequenceLevel",
            "c_delta",
            "c_a",
            "c_b",
            "m",
        ],
    )


def build_joint_forecast_vba(
    joint_risk_df: pd.DataFrame,
    cfg: Dict[str, Any],
    rm: RiskMatrix,
) -> pd.DataFrame:
    if joint_risk_df.empty:
        return pd.DataFrame(columns=["JoitID", "Year", "D_future", "beta", "Pf", "PossLevel", "RiskGrade"])

    fuyi = float(cfg.get("served_years", 1.0))
    life = float(cfg.get("design_life", 26.0))
    years_add = [0, 5, 10, 15, 20, 25]
    c_delta = 0.3
    c_a = 0.73
    c_b = 0.3
    m_val = 4.0
    ctf = fatigue_ctf(c_delta, c_a, c_b, m_val)

    rows: List[Dict[str, Any]] = []
    for _, r in joint_risk_df.iterrows():
        jid = _safe_str(r.get("JoitID"))
        consequence = int(_as_float(r.get("ConsequenceLevel")) or 3)
        collapse_lv = _as_float(r.get("CollapsePossLevel"))
        d = float(_as_float(r.get("D")) or 0.0)
        for add in years_add:
            year = fuyi + float(add)
            d_future = year * d / life if life != 0 else np.nan
            beta = fatigue_beta_forecast(d_future, ctf)
            pf = pf_from_beta(beta)
            lv_f = possibility_level_vba(pf, rm)
            if collapse_lv is None:
                lv_comb = lv_f
            elif lv_f is None:
                lv_comb = int(collapse_lv)
            else:
                lv_comb = min(int(collapse_lv), int(lv_f))
            grade = risk_grade(consequence, int(lv_comb), rm) if lv_comb is not None else ""
            rows.append(
                {
                    "JoitID": jid,
                    "Year": year,
                    "D_future": d_future,
                    "beta": beta,
                    "Pf": pf,
                    "PossLevel": lv_comb,
                    "RiskGrade": grade,
                }
            )

    return pd.DataFrame(rows, columns=["JoitID", "Year", "D_future", "beta", "Pf", "PossLevel", "RiskGrade"])


def build_joint_forecast_vba_wide(
    joint_risk_df: pd.DataFrame,
    cfg: Dict[str, Any],
    rm: RiskMatrix,
) -> pd.DataFrame:
    """
    VBA equivalent of Sheet1.JointRiskForeCast / Sheet12.
    Extends the current joint risk into the future time nodes used by Module 9.
    """
    base_cols = ["JoitID", "Brace", "JointType", "ConsequenceLevel", "CollapsePossLevel"]
    horizon_tags = [("N", 0), ("N+5", 5), ("N+10", 10), ("N+15", 15), ("N+20", 20), ("N+25", 25)]
    block_cols: List[str] = []
    for tag, _ in horizon_tags:
        block_cols.extend(
            [
                f"{tag}_D",
                f"{tag}_c_delta",
                f"{tag}_c_a",
                f"{tag}_c_b",
                f"{tag}_m",
                f"{tag}_CTf",
                f"{tag}_beta",
                f"{tag}_Pf",
                f"{tag}_FatiguePossLevel",
                f"{tag}_PossLevel",
                f"{tag}_RiskGrade",
            ]
        )
    out_cols = base_cols + block_cols

    if joint_risk_df.empty:
        return pd.DataFrame(columns=out_cols)

    fuyi = float(cfg.get("served_years", 1.0))
    life = float(cfg.get("design_life", 26.0))
    c_delta = 0.3
    c_a = 0.73
    c_b = 0.3
    m_val = 4.0
    ctf = fatigue_ctf(c_delta, c_a, c_b, m_val)

    rows: List[Dict[str, Any]] = []
    for _, r in joint_risk_df.iterrows():
        jid = _safe_str(r.get("JoitID"))
        brace = _safe_str(r.get("Brace"))
        joint_type = _safe_str(r.get("JointType"))
        consequence = int(_as_float(r.get("ConsequenceLevel")) or 3)
        collapse_lv = _as_float(r.get("CollapsePossLevel"))
        d_now = float(_as_float(r.get("D")) or 0.0)

        one: Dict[str, Any] = {
            "JoitID": jid,
            "Brace": brace,
            "JointType": joint_type,
            "ConsequenceLevel": consequence,
            "CollapsePossLevel": collapse_lv,
        }

        for tag, add in horizon_tags:
            year = fuyi + float(add)
            d_future = year * d_now / life if life != 0 else np.nan
            beta = fatigue_beta_forecast(d_future, ctf)
            pf = pf_from_beta(beta)
            lv_f = possibility_level_vba(pf, rm)
            if collapse_lv is None:
                lv_comb = lv_f
            elif lv_f is None:
                lv_comb = int(collapse_lv)
            else:
                lv_comb = min(int(collapse_lv), int(lv_f))
            grade = risk_grade(consequence, int(lv_comb), rm) if lv_comb is not None else ""

            one[f"{tag}_D"] = d_future
            one[f"{tag}_c_delta"] = c_delta
            one[f"{tag}_c_a"] = c_a
            one[f"{tag}_c_b"] = c_b
            one[f"{tag}_m"] = m_val
            one[f"{tag}_CTf"] = ctf
            one[f"{tag}_beta"] = beta
            one[f"{tag}_Pf"] = pf
            one[f"{tag}_FatiguePossLevel"] = lv_f
            one[f"{tag}_PossLevel"] = lv_comb
            one[f"{tag}_RiskGrade"] = grade

        rows.append(one)

    return pd.DataFrame(rows, columns=out_cols)


# =========================
# Module 9: 检测策略形成
# Sheet18 / Sheet19 equivalent outputs
# build_node_plan_vba -> 节点检测策略
# build_member_plan_vba -> 构件检测策略
# =========================
# Clean section marker for maintainers. The older nearby comment block is a
# historical artifact from an earlier encoding state and can be ignored.
# Module 9: inspection strategy generation starts here.
def _risk_grade_severity(grade: Any) -> Optional[int]:
    """
    Normalize risk grade labels to severity 1..5 (1 highest, 5 lowest).
    Supports Chinese numerals and Roman numerals.
    """
    s = _safe_str(grade)
    if s == "":
        return None

    zh = {"一": 1, "二": 2, "三": 3, "四": 4, "五": 5}
    if s in zh:
        return zh[s]

    s_up = s.upper().replace("级", "").replace("LEVEL", "").strip()
    roman = {"I": 1, "II": 2, "III": 3, "IV": 4, "V": 5}
    if s_up in roman:
        return roman[s_up]

    m = re.search(r"\d+", s_up)
    if m:
        v = int(m.group(0))
        if 1 <= v <= 5:
            return v
    return None


def _inspect_rank(level: str) -> int:
    lv = _safe_str(level).upper()
    if lv == "IV":
        return 3
    if lv == "III":
        return 2
    if lv == "II":
        return 1
    return 0


def _time_nodes_from_cfg(cfg: Dict[str, Any]) -> List[Tuple[int, str]]:
    fy = float(cfg.get("served_years", 1.0))
    life = float(cfg.get("design_life", 26.0))
    trun = int((life - fy) / 5 + 1)
    trun = max(1, min(6, trun))  # Sheet12 blocks are fixed to N..N+25
    out: List[Tuple[int, str]] = []
    for y in range(1, trun + 1):
        if y == 1:
            out.append((y, "当前"))
        else:
            out.append((y, f"第{(y - 1) * 5}年"))
    return out


def build_node_plan_vba(
    forecast_wide_df: pd.DataFrame,
    cfg: Dict[str, Any],
    seed: int = 42,
) -> pd.DataFrame:
    """
    VBA equivalent of Sheet18.JIANYAN.
    Consumes Module 8 node-risk results and emits the Module 9 node plan.
    """
    out_cols = [
        "JoitID",
        "Brace",
        "JointType",
        "ConsequenceLevel",
        "CollapsePossLevel",
        "D",
        "c_delta",
        "c_a",
        "c_b",
        "m",
        "CTf",
        "beta",
        "Pf",
        "FatiguePossLevel",
        "PossLevel",
        "RiskGrade",
        "InspectLevel",
        "TimeNode",
    ]
    if forecast_wide_df.empty:
        return pd.DataFrame(columns=out_cols)

    horizon_tags = [("N", 1), ("N+5", 2), ("N+10", 3), ("N+15", 4), ("N+20", 5), ("N+25", 6)]
    active_nodes = _time_nodes_from_cfg(cfg)
    active_y = {y for y, _ in active_nodes}
    y_to_time = {y: t for y, t in active_nodes}

    rows: List[Dict[str, Any]] = []
    blocks: List[Tuple[int, int, int]] = []  # (Y, start_idx, end_idx)

    for tag, y in horizon_tags:
        if y not in active_y:
            continue
        needed = [
            f"{tag}_D",
            f"{tag}_c_delta",
            f"{tag}_c_a",
            f"{tag}_c_b",
            f"{tag}_m",
            f"{tag}_CTf",
            f"{tag}_beta",
            f"{tag}_Pf",
            f"{tag}_FatiguePossLevel",
            f"{tag}_PossLevel",
            f"{tag}_RiskGrade",
        ]
        if any(c not in forecast_wide_df.columns for c in needed):
            continue

        start = len(rows)
        for _, r in forecast_wide_df.iterrows():
            rows.append(
                {
                    "JoitID": _safe_str(r.get("JoitID")),
                    "Brace": r.get("Brace"),
                    "JointType": _safe_str(r.get("JointType")),
                    "ConsequenceLevel": r.get("ConsequenceLevel"),
                    "CollapsePossLevel": r.get("CollapsePossLevel"),
                    "D": r.get(f"{tag}_D"),
                    "c_delta": r.get(f"{tag}_c_delta"),
                    "c_a": r.get(f"{tag}_c_a"),
                    "c_b": r.get(f"{tag}_c_b"),
                    "m": r.get(f"{tag}_m"),
                    "CTf": r.get(f"{tag}_CTf"),
                    "beta": r.get(f"{tag}_beta"),
                    "Pf": r.get(f"{tag}_Pf"),
                    "FatiguePossLevel": r.get(f"{tag}_FatiguePossLevel"),
                    "PossLevel": r.get(f"{tag}_PossLevel"),
                    "RiskGrade": _safe_str(r.get(f"{tag}_RiskGrade")),
                    "InspectLevel": "",
                    "TimeNode": y_to_time[y],
                }
            )
        end = len(rows) - 1
        if end >= start:
            blocks.append((y, start, end))

    rng = random.Random(seed)
    recent_ii: Dict[str, int] = {}   # keys promoted from risk-2 to IV
    recent_iii: Dict[str, int] = {}  # keys promoted from risk-3 to III

    for y, s, e in blocks:
        sev2_cnt = 0
        sev3_cnt = 0
        for i in range(s, e + 1):
            sev = _risk_grade_severity(rows[i]["RiskGrade"])
            if sev == 2:
                sev2_cnt += 1
            elif sev == 3:
                sev3_cnt += 1

        target2 = int(sev2_cnt * 0.2)
        target3 = int(sev3_cnt * 0.2)
        mod2 = 0
        mod3 = 0
        promoted2: List[str] = []
        promoted3: List[str] = []

        for i in range(s, e + 1):
            row = rows[i]
            sev = _risk_grade_severity(row["RiskGrade"])
            key = f"{_safe_str(row['JoitID'])}{_safe_str(row['Brace'])}"

            if sev == 2:
                if mod2 < target2 and key not in recent_ii and rng.random() <= 0.2:
                    row["InspectLevel"] = "IV"
                    mod2 += 1
                    promoted2.append(key)
                else:
                    row["InspectLevel"] = "III"
            elif sev == 3:
                if mod3 < target3 and key not in recent_iii and rng.random() <= 0.2:
                    row["InspectLevel"] = "III"
                    mod3 += 1
                    promoted3.append(key)
                else:
                    row["InspectLevel"] = "II"
            elif sev == 1:
                row["InspectLevel"] = "IV"
            else:
                row["InspectLevel"] = "II"

        # VBA backfill when random pass did not meet 95% target.
        if y % 2 == 0:
            order = range(s, e + 1, 1)
        else:
            order = range(e, s - 1, -1)
        if y > 4:
            gl = 0.8
        elif y > 3:
            gl = 0.4
        else:
            gl = 0.25

        if mod2 < 0.95 * target2:
            for i in order:
                if mod2 >= target2:
                    break
                row = rows[i]
                sev = _risk_grade_severity(row["RiskGrade"])
                if sev != 2:
                    continue
                key = f"{_safe_str(row['JoitID'])}{_safe_str(row['Brace'])}"
                if key in recent_ii or _safe_str(row["InspectLevel"]).upper() == "IV":
                    continue
                if rng.random() <= gl:
                    row["InspectLevel"] = "IV"
                    mod2 += 1
                    promoted2.append(key)

        if mod3 < 0.95 * target3:
            for i in order:
                if mod3 >= target3:
                    break
                row = rows[i]
                sev = _risk_grade_severity(row["RiskGrade"])
                if sev != 3:
                    continue
                key = f"{_safe_str(row['JoitID'])}{_safe_str(row['Brace'])}"
                if key in recent_iii or _safe_str(row["InspectLevel"]).upper() == "III":
                    continue
                if rng.random() <= gl:
                    row["InspectLevel"] = "III"
                    mod3 += 1
                    promoted3.append(key)

        # Keep only last 4 periods.
        for k in promoted2:
            recent_ii[k] = y
        for k in promoted3:
            recent_iii[k] = y
        recent_ii = {k: yy for k, yy in recent_ii.items() if yy > y - 4}
        recent_iii = {k: yy for k, yy in recent_iii.items() if yy > y - 4}

        # VBA sorts each time block by inspect level desc.
        block = rows[s : e + 1]
        block.sort(key=lambda rr: _inspect_rank(_safe_str(rr.get("InspectLevel"))), reverse=True)
        rows[s : e + 1] = block

    return pd.DataFrame(rows, columns=out_cols)


def build_member_plan_vba(
    member_risk_df: pd.DataFrame,
    cfg: Dict[str, Any],
    seed: int = 42,
) -> pd.DataFrame:
    """
    VBA equivalent of Sheet19.MemberCheck.
    Consumes Module 8 member-risk results and emits the Module 9 member plan.
    """
    out_cols = ["JointA", "JointB", "MemberType", "ConsequenceLevel", "RiskGrade", "InspectLevel", "TimeNode"]
    if member_risk_df.empty:
        return pd.DataFrame(columns=out_cols)

    base_rows = []
    for _, r in member_risk_df.iterrows():
        base_rows.append(
            {
                "JointA": _safe_str(r.get("JointA")),
                "JointB": _safe_str(r.get("JointB")),
                "MemberType": _safe_str(r.get("MemberType")),
                "ConsequenceLevel": r.get("ConsequenceLevel"),
                "RiskGrade": _safe_str(r.get("RiskGrade")),
            }
        )

    sev2_cnt = sum(1 for r in base_rows if _risk_grade_severity(r["RiskGrade"]) == 2)
    sev3_cnt = sum(1 for r in base_rows if _risk_grade_severity(r["RiskGrade"]) == 3)
    target2 = int(sev2_cnt * 0.2)
    target3 = int(sev3_cnt * 0.2)

    rng = random.Random(seed + 10007)
    recent_ii: Dict[str, int] = {}
    recent_iii: Dict[str, int] = {}
    rows: List[Dict[str, Any]] = []

    for y, tnode in _time_nodes_from_cfg(cfg):
        if y == 1:
            gl = 0.2
        elif y == 2:
            gl = 0.25
        elif y == 3:
            gl = 0.34
        elif y == 4:
            gl = 0.5
        else:
            gl = 1.0

        mod2 = 0
        mod3 = 0
        promoted2: List[str] = []
        promoted3: List[str] = []

        s = len(rows)
        for r in base_rows:
            one = dict(r)
            one["TimeNode"] = tnode
            one["InspectLevel"] = ""
            sev = _risk_grade_severity(one["RiskGrade"])
            key = f"{_safe_str(one['JointA'])}{_safe_str(one['JointB'])}"

            if sev == 2:
                if mod2 < target2 and key not in recent_ii and rng.random() <= gl:
                    one["InspectLevel"] = "IV"
                    mod2 += 1
                    promoted2.append(key)
                else:
                    one["InspectLevel"] = "III"
            elif sev == 3:
                if mod3 < target3 and key not in recent_iii and rng.random() <= gl:
                    one["InspectLevel"] = "III"
                    mod3 += 1
                    promoted3.append(key)
                else:
                    one["InspectLevel"] = "II"
            elif sev == 1:
                one["InspectLevel"] = "IV"
            else:
                one["InspectLevel"] = "II"
            rows.append(one)

        e = len(rows) - 1

        for k in promoted2:
            recent_ii[k] = y
        for k in promoted3:
            recent_iii[k] = y
        recent_ii = {k: yy for k, yy in recent_ii.items() if yy > y - 4}
        recent_iii = {k: yy for k, yy in recent_iii.items() if yy > y - 4}

        block = rows[s : e + 1]
        block.sort(key=lambda rr: _inspect_rank(_safe_str(rr.get("InspectLevel"))), reverse=True)
        rows[s : e + 1] = block

    return pd.DataFrame(rows, columns=out_cols)


# =========================
# Inspection assignment
# =========================

def assign_by_distribution(ids: List[str], dist: Dict[str,float], seed: int) -> Dict[str,str]:
    rng = random.Random(seed)
    ids = list(ids)
    rng.shuffle(ids)
    n = len(ids)
    levels = list(dist.keys())
    ratios = [dist[k] for k in levels]
    counts=[]
    remaining=n
    for r in ratios[:-1]:
        cnt = int(round(n*r))
        cnt = min(cnt, remaining)
        counts.append(cnt)
        remaining -= cnt
    counts.append(remaining)

    out={}
    start=0
    for lvl,cnt in zip(levels,counts):
        for _id in ids[start:start+cnt]:
            out[_id] = lvl
        start += cnt
    return out


# =========================
# Structure classification (FindLegMember + Find_X_Joint)
# =========================

_RE_P8 = re.compile(r"^P\d8\d$", re.I)
_RE_P9 = re.compile(r"^P\d9\d$", re.I)
_RE_12L = re.compile(r"^\d12L$", re.I)
_RE_13L = re.compile(r"^\d13L$", re.I)
_RE_16L = re.compile(r"^16\dL$", re.I)


def classify_by_name(
    joints: pd.DataFrame,
    members: pd.DataFrame,
    waterline_z: float,
) -> Tuple[pd.DataFrame, pd.DataFrame]:
    """
    Port of VBA Sheet2_JointTYPE + Sheet3_MemberTYPE naming rules.
    """
    joints = joints.copy()
    members = members.copy()

    if not joints.empty:
        jz = pd.to_numeric(joints["Z"], errors="coerce")
        jid = joints["Joint"].astype(str)
        below = jz < float(waterline_z)
        joints.loc[below & jid.str.endswith("L"), "JointType"] = "LegJoint"
        joints.loc[below & jid.str.endswith("X"), "JointType"] = "X Joint"

    if not members.empty:
        a = members["A"].astype(str)
        b = members["B"].astype(str)
        z1 = pd.to_numeric(members["Z1"], errors="coerce")
        z2 = pd.to_numeric(members["Z2"], errors="coerce")
        below = (z1 < float(waterline_z)) & (z2 < float(waterline_z))

        # LEG: both joints end with L and same penultimate char.
        same_lane = a.str.slice(-2, -1) == b.str.slice(-2, -1)
        members.loc[below & a.str.endswith("L") & b.str.endswith("L") & same_lane, "MemberType"] = "LEG"

        # LEG: P* to P* except P#8# / P#9# cases.
        p_case = a.str.startswith("P") & b.str.startswith("P")
        p_excl = a.str.match(_RE_P8) | a.str.match(_RE_P9) | b.str.match(_RE_P8) | b.str.match(_RE_P9)
        members.loc[below & p_case & (~p_excl), "MemberType"] = "LEG"

        # X-Brace rules.
        excl_a = a.str.match(_RE_12L) | a.str.match(_RE_13L) | a.str.match(_RE_16L)
        excl_b = b.str.match(_RE_12L) | b.str.match(_RE_13L) | b.str.match(_RE_16L)
        members.loc[below & b.str.endswith("X") & (~excl_a), "MemberType"] = "X-Brace"
        members.loc[below & a.str.endswith("X") & (~excl_b), "MemberType"] = "X-Brace"

    return joints, members

def classify_structure(
    joints: pd.DataFrame,
    members: pd.DataFrame,
    work_points_xy: List[Tuple[float,float]],
    wp_z: float,
    min_leg_od: float,
    xy_tol: float = 1e-3,
    vertical_tol_deg: float = 8.0,
    xbrace_min_od: Optional[float] = None,
    x_angle_deviation: Optional[float] = None,
    apply_sheet2_jointtype: bool = True,
    apply_sheet3_membertype: bool = True,
) -> Tuple[pd.DataFrame, pd.DataFrame]:
    """
    Module 1 continuation: emulate the full VBA structure classification chain.
    1) Sheet1.FindLegMember
    2) Sheet1.Find_X_Joint
    3) Sheet2.JointTYPE
    4) Sheet3.Member

    Notes:
    - Sheet2/Sheet3 clear-and-rewrite type columns (VBA behavior).
    """
    joints = joints.copy()
    members = members.copy()

    if "JointType" not in joints.columns:
        joints["JointType"] = None
    if "MemberType" not in members.columns:
        members["MemberType"] = None

    j_key = joints["Joint"].astype(str).str.strip()
    jx = pd.to_numeric(joints["X"], errors="coerce")
    jy = pd.to_numeric(joints["Y"], errors="coerce")
    jz = pd.to_numeric(joints["Z"], errors="coerce")

    a_key = members["A"].astype(str).str.strip()
    b_key = members["B"].astype(str).str.strip()
    od_num = pd.to_numeric(members["OD"], errors="coerce")
    z1 = pd.to_numeric(members["Z1"], errors="coerce")
    z2 = pd.to_numeric(members["Z2"], errors="coerce")
    x_ang_tol = float(x_angle_deviation) if x_angle_deviation is not None else 15.0

    joint_xyz: Dict[str, Tuple[float, float, float]] = {}
    for _, jr in joints.iterrows():
        jid = str(jr.get("Joint", "")).strip()
        if jid == "" or jid in joint_xyz:
            continue
        xv = _as_float(jr.get("X"))
        yv = _as_float(jr.get("Y"))
        zv = _as_float(jr.get("Z"))
        if xv is None or yv is None or zv is None:
            continue
        joint_xyz[jid] = (float(xv), float(yv), float(zv))

    # ===== VBA Sheet1.FindLegMember =====
    # Mark LEG members by tracing from each work point both downward and upward.
    valid_joint = j_key.ne("") & jx.notna() & jy.notna() & jz.notna()

    def _same_xyz_mask(xv: float, yv: float, zv: float) -> pd.Series:
        return (
            np.isclose(jx, float(xv), rtol=0.0, atol=1e-9)
            & np.isclose(jy, float(yv), rtol=0.0, atol=1e-9)
            & np.isclose(jz, float(zv), rtol=0.0, atol=1e-9)
        )

    def _trace_leg(cur_x: float, cur_y: float, cur_z: float, down: bool) -> Tuple[float, float, float]:
        # Keep a hard guard to avoid pathological loops.
        for _ in range(5000):
            same_xyz = _same_xyz_mask(cur_x, cur_y, cur_z)
            j_list = j_key[same_xyz].tolist()
            if not j_list:
                break

            cand = (a_key.isin(j_list) | b_key.isin(j_list))
            zsum = z1 + z2
            if down:
                cand = cand & (zsum < 2.0 * float(cur_z))
            else:
                cand = cand & (zsum > 2.0 * float(cur_z))

            sel = members[cand]
            if sel.empty:
                break

            max_od_i = 0.0
            ja = None
            jb = None
            # VBA iterates rows and picks max OD manually.
            for _, r in sel.iterrows():
                odv = _as_float(r.get("OD"))
                if odv is None:
                    continue
                if float(odv) > max_od_i:
                    max_od_i = float(odv)
                    ja = str(r["A"]).strip()
                    jb = str(r["B"]).strip()

            if ja is None or jb is None or max_od_i < 100.0:
                break

            members.loc[(a_key == ja) & (b_key == jb), "MemberType"] = "LEG"

            next_id = jb if ja in set(j_list) else ja
            nxt = joints[j_key == next_id]
            if nxt.empty:
                break
            cur_x = float(pd.to_numeric(nxt.iloc[0]["X"], errors="coerce"))
            cur_y = float(pd.to_numeric(nxt.iloc[0]["Y"], errors="coerce"))
            cur_z = float(pd.to_numeric(nxt.iloc[0]["Z"], errors="coerce"))

        return cur_x, cur_y, cur_z

    for wp_x, wp_y in work_points_xy:
        if valid_joint.sum() == 0:
            break
        dist2 = (jx - float(wp_x)) ** 2 + (jy - float(wp_y)) ** 2 + (jz - float(wp_z)) ** 2
        dist2 = dist2.where(valid_joint, np.nan)
        if dist2.notna().sum() == 0:
            continue
        idx0 = dist2.idxmin()
        x0 = float(jx.loc[idx0])
        y0 = float(jy.loc[idx0])
        z0 = float(jz.loc[idx0])

        _trace_leg(x0, y0, z0, down=True)
        _trace_leg(x0, y0, z0, down=False)

    # LegJoint marking in two passes (A-side then B-side), same as VBA.
    def _mark_leg_joint_for_ids(ids: Sequence[str]) -> None:
        for jid in ids:
            if jid == "":
                continue
            conn = members[(a_key == jid) | (b_key == jid)]
            if conn.empty:
                continue
            conn_od = pd.to_numeric(conn["OD"], errors="coerce")
            max_od = float(conn_od.max()) if conn_od.notna().any() else 0.0
            min_od = max(0.2 * max_od, float(min_leg_od) / 10.0)
            cnt = int((conn_od >= min_od).sum())
            if cnt > 2:
                joints.loc[j_key == jid, "JointType"] = "LegJoint"

    leg_members = members[members["MemberType"].astype(str).str.strip() == "LEG"].copy()
    if not leg_members.empty:
        leg_a = set(leg_members["A"].astype(str).str.strip().tolist())
        leg_b = set(leg_members["B"].astype(str).str.strip().tolist())
        _mark_leg_joint_for_ids(j_key[j_key.isin(leg_a)].tolist())
        _mark_leg_joint_for_ids(j_key[j_key.isin(leg_b)].tolist())

    # ===== VBA Sheet1.Find_X_Joint =====
    # Candidate joints: not LegJoint and Z <= El.
    jt_now = joints["JointType"].fillna("").astype(str).str.strip()
    x_joint_candidates = joints[(jt_now != "LegJoint") & (jz <= float(wp_z))]["Joint"].astype(str).str.strip().tolist()

    for jid in x_joint_candidates:
        if jid == "":
            continue

        conn_mask = (a_key == jid) | (b_key == jid)
        if not bool(conn_mask.any()):
            continue
        conn_od = od_num[conn_mask]
        if not bool(conn_od.notna().any()):
            continue
        max_od = float(conn_od.max())

        near_mask = conn_mask & od_num.notna() & (od_num >= (max_od - 2.0))
        near_rows = members[near_mask]
        if near_rows.empty:
            continue

        segs: List[Dict[str, Any]] = []
        for _, mr in near_rows.iterrows():
            ja = str(mr.get("A", "")).strip()
            jb = str(mr.get("B", "")).strip()
            p1 = joint_xyz.get(ja)
            p2 = joint_xyz.get(jb)
            if p1 is None or p2 is None:
                continue
            segs.append(
                {
                    "a": ja,
                    "b": jb,
                    "p1": p1,
                    "p2": p2,
                    "remark": "",
                }
            )

        num_m = len(segs)
        if num_m < 4:
            continue

        angle_deviation = 2.0
        num_collinear = 0
        for i in range(num_m - 1):
            p1i = segs[i]["p1"]
            p2i = segs[i]["p2"]
            v1 = (p2i[0] - p1i[0], p2i[1] - p1i[1], p2i[2] - p1i[2])
            for j in range(i + 1, num_m):
                p1j = segs[j]["p1"]
                p2j = segs[j]["p2"]
                v2 = (p2j[0] - p1j[0], p2j[1] - p1j[1], p2j[2] - p1j[2])
                the_angle = vector_angle_degree_vba(v1, v2)
                if the_angle > 90.0:
                    the_angle = 180.0 - the_angle
                if the_angle <= angle_deviation and abs(v2[2]) != 0.0:
                    num_collinear += 1
                    remark = f"Pair-{num_collinear}"
                    segs[i]["remark"] = remark
                    segs[j]["remark"] = remark

        if num_collinear != 2:
            continue

        joints.loc[j_key == jid, "JointType"] = "X Joint"

        for seg in segs:
            if "Pair" not in str(seg.get("remark", "")):
                continue

            ja = str(seg["a"])
            jb = str(seg["b"])
            members.loc[(a_key == ja) & (b_key == jb), "MemberType"] = "X-Brace"

            start_joint = jid
            end_joint = jb if ja == jid else ja

            for _ in range(5000):
                start_xyz = joint_xyz.get(start_joint)
                end_xyz = joint_xyz.get(end_joint)
                if start_xyz is None or end_xyz is None:
                    break

                mt_now = members["MemberType"].fillna("").astype(str).str.strip()
                cand_mask = (
                    ((a_key == end_joint) | (b_key == end_joint))
                    & od_num.notna()
                    & (od_num >= (max_od - 2.0))
                    & (od_num <= (max_od + 2.0))
                    & (mt_now != "X-Brace")
                )
                cand_rows = members[cand_mask]
                if cand_rows.empty:
                    break

                found_next = False
                for _, cr in cand_rows.iterrows():
                    ca = str(cr.get("A", "")).strip()
                    cb = str(cr.get("B", "")).strip()
                    pca = joint_xyz.get(ca)
                    pcb = joint_xyz.get(cb)
                    if pca is None or pcb is None:
                        continue

                    v_member = (pcb[0] - pca[0], pcb[1] - pca[1], pcb[2] - pca[2])
                    v_ref = (
                        end_xyz[0] - start_xyz[0],
                        end_xyz[1] - start_xyz[1],
                        end_xyz[2] - start_xyz[2],
                    )
                    the_angle = vector_angle_degree_vba(v_member, v_ref)
                    if the_angle > 90.0:
                        the_angle = 180.0 - the_angle
                    if the_angle <= x_ang_tol:
                        if ca == end_joint:
                            start_joint = ca
                            end_joint = cb
                        else:
                            start_joint = cb
                            end_joint = ca
                        members.loc[(a_key == ca) & (b_key == cb), "MemberType"] = "X-Brace"
                        found_next = True
                        break

                if not found_next:
                    break

    if apply_sheet2_jointtype:
        # ===== VBA Sheet2.JointTYPE =====
        joints["JointType"] = None
        below = jz < float(wp_z)
        # DataFrame rows already represent real data rows only.
        # Do not skip index 0 here; doing so incorrectly drops the first
        # actual joint (for WC9-7 this wrongly leaves 001L untyped).
        joints.loc[below & j_key.str.endswith("L"), "JointType"] = "LegJoint"
        joints.loc[below & j_key.str.endswith("X"), "JointType"] = "X Joint"

    if apply_sheet3_membertype:
        # ===== VBA Sheet3.Member =====
        members["MemberType"] = None
        below_m = (z1 < float(wp_z)) & (z2 < float(wp_z))
        same_lane = a_key.str.slice(-2, -1) == b_key.str.slice(-2, -1)
        members.loc[below_m & a_key.str.endswith("L") & b_key.str.endswith("L") & same_lane, "MemberType"] = "LEG"

        p_case = a_key.str.startswith("P") & b_key.str.startswith("P")
        p_excl = a_key.str.match(_RE_P8) | a_key.str.match(_RE_P9) | b_key.str.match(_RE_P8) | b_key.str.match(_RE_P9)
        members.loc[below_m & p_case & (~p_excl), "MemberType"] = "LEG"

        excl_a = a_key.str.match(_RE_12L) | a_key.str.match(_RE_13L) | a_key.str.match(_RE_16L)
        excl_b = b_key.str.match(_RE_12L) | b_key.str.match(_RE_13L) | b_key.str.match(_RE_16L)
        members.loc[below_m & b_key.str.endswith("X") & (~excl_a), "MemberType"] = "X-Brace"
        members.loc[below_m & a_key.str.endswith("X") & (~excl_b), "MemberType"] = "X-Brace"

    return joints, members


def discover_data_bundle(data_dir: str | Path) -> Dict[str, Any]:
    root = Path(data_dir)
    if not root.exists():
        raise FileNotFoundError(f"data dir not found: {root}")

    model_candidates = sorted(root.rglob("sacinp*"), key=lambda p: len(str(p)))
    if not model_candidates:
        raise FileNotFoundError(f"no sacinp model file found under: {root}")

    clplog_files = sorted([p for p in root.rglob("clplog*") if p.is_file()], key=_natural_sort_key)
    if not clplog_files:
        raise FileNotFoundError(f"no clplog files found under: {root}")

    ftglst_files = sorted(
        [
            p for p in root.rglob("ftglst*")
            if p.is_file() and not p.name.lower().endswith(".runx")
        ],
        key=_natural_sort_key,
    )
    if not ftglst_files:
        raise FileNotFoundError(f"no ftglst files found under: {root}")

    ftginp_files = sorted([p for p in root.rglob("ftginp*") if p.is_file()], key=_natural_sort_key)

    return {
        "model_file": model_candidates[0],
        "clplog_files": clplog_files,
        "ftglst_files": ftglst_files,
        "ftginp_files": ftginp_files,
    }


def round_numeric_df(df: pd.DataFrame, digits: int = 2) -> pd.DataFrame:
    def _round_half_up(x: Any) -> Any:
        if pd.isna(x):
            return x
        try:
            q = Decimal("1").scaleb(-digits)
            d = Decimal(str(x))
            return float(d.quantize(q, rounding=ROUND_HALF_UP))
        except (InvalidOperation, ValueError, TypeError):
            return x

    out = df.copy()
    if out.empty:
        return out
    num_cols = out.select_dtypes(include=[np.number]).columns
    if len(num_cols) > 0:
        for c in num_cols:
            out[c] = out[c].map(_round_half_up)
    return out


# =========================
# Template-layout writer
# =========================

def _clear_sheet_block(ws, start_row: int, start_col: int, end_col: int) -> None:
    if ws.max_row < start_row:
        return
    for row in ws.iter_rows(min_row=start_row, max_row=ws.max_row, min_col=start_col, max_col=end_col):
        for cell in row:
            cell.value = None


def _write_df_block(ws, df: pd.DataFrame, start_row: int, start_col: int, cols: Sequence[str]) -> None:
    if df.empty:
        return
    r = start_row
    for _, row in df.iterrows():
        for j, c in enumerate(cols):
            ws.cell(r, start_col + j).value = row.get(c)
        r += 1


def write_template_layout_workbook(
    template_xlsm: str | Path,
    out_xlsx: str | Path,
    *,
    joints_df: pd.DataFrame,
    groups_df: pd.DataFrame,
    members_df: pd.DataFrame,
    sections_df: pd.DataFrame,
    collapse_df: pd.DataFrame,
    collapse_summary_df: pd.DataFrame,
    fatigue_df: pd.DataFrame,
    member_risk_df: pd.DataFrame,
    joint_risk_df: pd.DataFrame,
    forecast_wide_df: pd.DataFrame,
    node_plan_df: pd.DataFrame,
    member_plan_df: pd.DataFrame,
    clplog_paths: Sequence[Path],
) -> None:
    """
    Write outputs into the original template sheet layout to preserve styles/headers.
    Sheet index mapping follows the template workbook:
    6 collapse, 7 fatigue, 9 node risk, 10 node forecast, 11 node plan,
    12 member risk, 13 member plan, 15 joints, 16 groups, 17 members, 18 sections.
    """
    out_path = Path(out_xlsx)
    # Avoid extension/content-type mismatch:
    # only keep VBA parts when target is .xlsm.
    keep_vba = out_path.suffix.lower() == ".xlsm"
    wb = openpyxl.load_workbook(template_xlsm, data_only=False, keep_vba=keep_vba)

    ws_collapse = wb[wb.sheetnames[6]]
    ws_fatigue = wb[wb.sheetnames[7]]
    ws_node_risk = wb[wb.sheetnames[9]]
    ws_node_fcst = wb[wb.sheetnames[10]]
    ws_node_plan = wb[wb.sheetnames[11]]
    ws_member_risk = wb[wb.sheetnames[12]]
    ws_member_plan = wb[wb.sheetnames[13]]
    ws_joints = wb[wb.sheetnames[15]]
    ws_groups = wb[wb.sheetnames[16]]
    ws_members = wb[wb.sheetnames[17]]
    ws_sections = wb[wb.sheetnames[18]]

    # SACS sheets
    _clear_sheet_block(ws_joints, start_row=2, start_col=1, end_col=5)
    _write_df_block(ws_joints, joints_df, start_row=2, start_col=1, cols=["Joint", "X", "Y", "Z", "JointType"])

    _clear_sheet_block(ws_groups, start_row=2, start_col=1, end_col=3)
    _write_df_block(ws_groups, groups_df, start_row=2, start_col=1, cols=["ID", "OD", "Type"])

    _clear_sheet_block(ws_members, start_row=2, start_col=1, end_col=7)
    _write_df_block(ws_members, members_df, start_row=2, start_col=1, cols=["A", "B", "ID", "OD", "MemberType", "Z1", "Z2"])

    _clear_sheet_block(ws_sections, start_row=2, start_col=1, end_col=3)
    _write_df_block(ws_sections, sections_df, start_row=2, start_col=1, cols=["ID", "Type", "OD"])

    # Collapse sheet
    for r in range(2, 14):
        ws_collapse.cell(r, 6).value = None
    for i, p in enumerate(clplog_paths[:12], start=2):
        ws_collapse.cell(i, 2).value = str(Path(p).resolve())
    for _, rr in collapse_summary_df.iterrows():
        rid = _as_float(rr.get("LOADID"))
        if rid is None:
            continue
        row = int(rid) + 1
        if 2 <= row <= 13:
            ws_collapse.cell(row, 6).value = rr.get("LastLoadFactor")
    _clear_sheet_block(ws_collapse, start_row=16, start_col=1, end_col=5)
    ws_collapse.cell(15, 1).value = "LOADID"
    ws_collapse.cell(15, 2).value = "TYPE"
    ws_collapse.cell(15, 3).value = "LOCATION"
    ws_collapse.cell(15, 4).value = "FACTOR"
    ws_collapse.cell(15, 5).value = "REMARK"
    _write_df_block(
        ws_collapse,
        collapse_df,
        start_row=16,
        start_col=1,
        cols=["LOADID", "TYPE", "LOCATION", "FACTOR", "REMARK"],
    )

    # Fatigue sheet (A:T from row 5)
    _clear_sheet_block(ws_fatigue, start_row=5, start_col=1, end_col=20)
    _write_df_block(
        ws_fatigue,
        fatigue_df,
        start_row=5,
        start_col=1,
        cols=[
            "JOINT",
            "CHD_A",
            "CHD_B",
            "BRACE",
            "BRACE_TOP",
            "BRACE_TOP_LEFT",
            "BRACE_LEFT",
            "BRACE_BOT_LEFT",
            "BRACE_BOT",
            "BRACE_BOT_RIGHT",
            "BRACE_RIGHT",
            "BRACE_TOP_RIGHT",
            "CHORD_TOP",
            "CHORD_TOP_LEFT",
            "CHORD_LEFT",
            "CHORD_BOT_LEFT",
            "CHORD_BOT",
            "CHORD_BOT_RIGHT",
            "CHORD_RIGHT",
            "CHORD_TOP_RIGHT",
        ],
    )

    # Member risk (A:K from row 3)
    _clear_sheet_block(ws_member_risk, start_row=3, start_col=1, end_col=11)
    _write_df_block(
        ws_member_risk,
        member_risk_df,
        start_row=3,
        start_col=1,
        cols=[
            "JointA",
            "JointB",
            "MemberType",
            "ConsequenceLevel",
            "A",
            "B",
            "Rm",
            "VR",
            "Pf",
            "CollapsePossLevel",
            "RiskGrade",
        ],
    )

    # Member plan (A:G from row 2)
    _clear_sheet_block(ws_member_plan, start_row=2, start_col=1, end_col=7)
    _write_df_block(
        ws_member_plan,
        member_plan_df,
        start_row=2,
        start_col=1,
        cols=["JointA", "JointB", "MemberType", "ConsequenceLevel", "RiskGrade", "InspectLevel", "TimeNode"],
    )

    # Node risk (A:U from row 3)
    _clear_sheet_block(ws_node_risk, start_row=3, start_col=1, end_col=21)
    _write_df_block(
        ws_node_risk,
        joint_risk_df,
        start_row=3,
        start_col=1,
        cols=[
            "JoitID",
            "Brace",
            "JointType",
            "ConsequenceLevel",
            "A",
            "B",
            "Rm",
            "VR",
            "Pf_collapse",
            "CollapsePossLevel",
            "D",
            "c_delta",
            "c_a",
            "c_b",
            "m",
            "CTf",
            "beta_fatigue",
            "Pf_fatigue",
            "FatiguePossLevel",
            "PossLevel",
            "RiskGrade",
        ],
    )

    # Node forecast (A:BS from row 4)
    _clear_sheet_block(ws_node_fcst, start_row=4, start_col=1, end_col=71)
    fcst_cols: List[str] = ["JoitID", "Brace", "JointType", "ConsequenceLevel", "CollapsePossLevel"]
    for tag in ["N", "N+5", "N+10", "N+15", "N+20", "N+25"]:
        fcst_cols.extend(
            [
                f"{tag}_D",
                f"{tag}_c_delta",
                f"{tag}_c_a",
                f"{tag}_c_b",
                f"{tag}_m",
                f"{tag}_CTf",
                f"{tag}_beta",
                f"{tag}_Pf",
                f"{tag}_FatiguePossLevel",
                f"{tag}_PossLevel",
                f"{tag}_RiskGrade",
            ]
        )
    _write_df_block(ws_node_fcst, forecast_wide_df, start_row=4, start_col=1, cols=fcst_cols)

    # Node plan (A:R from row 3)
    _clear_sheet_block(ws_node_plan, start_row=3, start_col=1, end_col=18)
    _write_df_block(
        ws_node_plan,
        node_plan_df,
        start_row=3,
        start_col=1,
        cols=[
            "JoitID",
            "Brace",
            "JointType",
            "ConsequenceLevel",
            "CollapsePossLevel",
            "D",
            "c_delta",
            "c_a",
            "c_b",
            "m",
            "CTf",
            "beta",
            "Pf",
            "FatiguePossLevel",
            "PossLevel",
            "RiskGrade",
            "InspectLevel",
            "TimeNode",
        ],
    )

    _mkdir_for_file(out_path)
    wb.save(out_path)


# =========================
# One-click orchestration for Modules 1-9
# =========================

def run(
    template_xlsm: str | Path,
    model_file: str | Path,
    clplog_file: str | Path | Sequence[str | Path],
    ftglst_file: str | Path | Sequence[str | Path],
    out_xlsx: str | Path,
    policy_mode: str = "strict",
    seed: int = 42,
    params_json: str | Path | None = None,
    ftginp_files: Sequence[str | Path] | None = None,
    manual_fill_workbook: str | Path | None = None,
    interactive_manual_fill: bool = False,
    enable_topology_inference: bool = False,
    apply_sheet2_jointtype: bool = True,
    apply_sheet3_membertype: bool = True,
) -> None:
    """
    Execute the full 9-module workflow and write the intermediate workbook.
    This is the business entry called by sacs_to_report.py.
    """
    cfg = load_from_params_json(params_json) if params_json else load_from_template(template_xlsm)
    risk_pack: RiskMatrixPack = cfg["risk_pack"]

    joints, groups, members, sections = parse_sacinp(model_file)
    joints2, members2 = classify_structure(
        joints=joints,
        members=members,
        work_points_xy=cfg["work_points"],
        wp_z=cfg["wp_z"],
        min_leg_od=cfg["min_leg_od"],
        xy_tol=3.0,
        vertical_tol_deg=8.0,
        xbrace_min_od=None,
        x_angle_deviation=cfg.get("x_angle_deviation"),
        apply_sheet2_jointtype=apply_sheet2_jointtype,
        apply_sheet3_membertype=apply_sheet3_membertype,
    )

    clplog_paths = _resolve_multi_inputs(clplog_file, dir_pattern="clplog*")
    if len(clplog_paths) > MAX_VBA_COLLAPSE_FILES:
        raise ValueError(
            f"collapse files exceed VBA limit: got {len(clplog_paths)}, "
            f"max {MAX_VBA_COLLAPSE_FILES} (Sheet5 B2:B13)"
        )
    ftglst_paths = _resolve_multi_inputs(ftglst_file, dir_pattern="ftglst*")
    ftglst_paths = [p for p in ftglst_paths if not p.name.lower().endswith(".runx")]
    if not ftglst_paths:
        raise FileNotFoundError("no ftglst files resolved from input")
    if len(ftglst_paths) > MAX_VBA_FATIGUE_FILES:
        raise ValueError(
            f"fatigue result files exceed VBA limit: got {len(ftglst_paths)}, "
            f"max {MAX_VBA_FATIGUE_FILES} (Sheet1 B39:H39)"
        )

    ftginp_paths: List[Path] = []
    if ftginp_files:
        ftginp_paths = _resolve_multi_inputs(ftginp_files, dir_pattern="ftginp*")
        if len(ftginp_paths) > MAX_VBA_FTGINP_FILES:
            raise ValueError(
                f"fatigue input files exceed VBA Ringmember limit: got {len(ftginp_paths)}, "
                f"max {MAX_VBA_FTGINP_FILES} (Sheet1 row40 B:I)"
            )

    print(f"[INFO] collapse files resolved: {len(clplog_paths)}")
    for i, p in enumerate(clplog_paths, start=1):
        print(f"[INFO] collapse LOADID={i} -> {p}")

    print(f"[INFO] fatigue files resolved: {len(ftglst_paths)}")
    for i, p in enumerate(ftglst_paths, start=1):
        print(f"[INFO] fatigue ftglst[{i}] -> {p}")

    merge_cfg: Optional[Sequence[Dict[str, Any]]] = cfg.get("fatigue_merge")
    fatigue_selector_audit_df = pd.DataFrame(columns=FATIGUE_AUDIT_COLUMNS)
    manual_selector_overrides = load_manual_selector_overrides_from_workbook(manual_fill_workbook)
    if ftginp_paths:
        print(f"[INFO] fatigue input files resolved: {len(ftginp_paths)}")
        for i, p in enumerate(ftginp_paths, start=1):
            print(f"[INFO] fatigue ftginp[{i}] -> {p}")
        merge_cfg_raw = build_fatigue_merge_cfg_from_ftginp(ftginp_paths)
        if enable_topology_inference:
            merge_cfg, fatigue_selector_audit_df = enrich_fatigue_merge_cfg_by_topology(merge_cfg_raw, members2)
            if manual_selector_overrides:
                merge_cfg, manual_applied_df = apply_manual_selector_overrides(merge_cfg, manual_selector_overrides)
                if not manual_applied_df.empty:
                    fatigue_selector_audit_df = pd.concat(
                        [fatigue_selector_audit_df, manual_applied_df], ignore_index=True
                    )
            if not fatigue_selector_audit_df.empty:
                inferred_n = int((fatigue_selector_audit_df["Action"] == "inferred").sum())
                unresolved_n = int((fatigue_selector_audit_df["Action"] == "unresolved").sum())
                print(
                    f"[INFO] topology selector补全: inferred={inferred_n}, unresolved={unresolved_n}"
                )
        else:
            merge_cfg = list(merge_cfg_raw)
            if manual_selector_overrides:
                merge_cfg, manual_applied_df = apply_manual_selector_overrides(merge_cfg, manual_selector_overrides)
            else:
                manual_applied_df = pd.DataFrame(columns=FATIGUE_AUDIT_COLUMNS)
            fatigue_selector_audit_df = build_ringmember_manual_fill_audit(merge_cfg)
            if not manual_applied_df.empty:
                fatigue_selector_audit_df = pd.concat([fatigue_selector_audit_df, manual_applied_df], ignore_index=True)
            print("[INFO] strict VBA mode: topology inference disabled for fatigue merge.")
            manual_n = int((fatigue_selector_audit_df["Action"] == "manual_fill_needed").sum())
            if manual_n > 0:
                manual_csv = _write_manual_fill_csv(fatigue_selector_audit_df, out_xlsx)
                warn_msg = _build_manual_fill_warning(fatigue_selector_audit_df, manual_csv)
                print(
                    "[WARN] 请检查 O、Q、S 等列在 38 行后是否有撑杆未填入（VBA 红色标记场景），"
                    f"未映射 JSLC 节点数: {manual_n}"
                )
                if manual_csv is not None:
                    print(f"[WARN] manual fill csv: {manual_csv}")
                _show_warning_popup("疲劳输入检查", warn_msg)

                if interactive_manual_fill:
                    user_overrides_opt = _collect_manual_overrides_gui(fatigue_selector_audit_df)
                    if user_overrides_opt is None:
                        # GUI unavailable -> fallback to terminal input.
                        user_overrides = _collect_manual_overrides_interactive(fatigue_selector_audit_df)
                    else:
                        user_overrides = user_overrides_opt
                    if user_overrides:
                        merge_cfg, interactive_applied_df = apply_manual_selector_overrides(merge_cfg, user_overrides)
                        # Rebuild pending audit from updated selectors, then append applied log.
                        pending_after_df = build_ringmember_manual_fill_audit(merge_cfg)
                        if not interactive_applied_df.empty:
                            fatigue_selector_audit_df = pd.concat(
                                [pending_after_df, interactive_applied_df], ignore_index=True
                            )
                        else:
                            fatigue_selector_audit_df = pending_after_df
                        manual_left = int((pending_after_df["Action"] == "manual_fill_needed").sum())
                        print(
                            "[INFO] interactive manual fill applied: "
                            f"applied={len(interactive_applied_df)}, remaining={manual_left}"
                        )
        if len(ftginp_paths) != len(ftglst_paths):
            print(
                "[WARN] ftginp count != ftglst count; "
                "VBA-style merge config will be applied by file index order."
            )

    collapse_df, collapse_summary_df = parse_clplogs(clplog_paths)
    fatigue_df = parse_ftglst_details(ftglst_paths, merge_cfg=merge_cfg)

    rm = risk_pack.rm
    member_risk_df = build_member_risk_vba(
        members_df=members2,
        collapse_df=collapse_df,
        collapse_summary_df=collapse_summary_df,
        cfg=cfg,
        rm=rm,
    )
    joint_risk_df = build_joint_risk_vba(
        joints_df=joints2,
        fatigue_df=fatigue_df,
        collapse_df=collapse_df,
        collapse_summary_df=collapse_summary_df,
        cfg=cfg,
        rm=rm,
    )
    forecast_long_df = build_joint_forecast_vba(
        joint_risk_df=joint_risk_df,
        cfg=cfg,
        rm=rm,
    )
    forecast_df = build_joint_forecast_vba_wide(
        joint_risk_df=joint_risk_df,
        cfg=cfg,
        rm=rm,
    )
    node_plan_df = build_node_plan_vba(
        forecast_wide_df=forecast_df,
        cfg=cfg,
        seed=seed,
    )
    member_plan_df = build_member_plan_vba(
        member_risk_df=member_risk_df,
        cfg=cfg,
        seed=seed,
    )

    # Main output: write into template-like sheet layout to preserve format.
    write_template_layout_workbook(
        template_xlsm=template_xlsm,
        out_xlsx=out_xlsx,
        joints_df=joints2,
        groups_df=groups,
        members_df=members2,
        sections_df=sections,
        collapse_df=collapse_df,
        collapse_summary_df=collapse_summary_df,
        fatigue_df=fatigue_df,
        member_risk_df=member_risk_df,
        joint_risk_df=joint_risk_df,
        forecast_wide_df=forecast_df,
        node_plan_df=node_plan_df,
        member_plan_df=member_plan_df,
        clplog_paths=clplog_paths,
    )

    # Supplemental audit sheets for manual verification.
    out_path = Path(out_xlsx)
    keep_vba_out = out_path.suffix.lower() == ".xlsm"
    wb_out = openpyxl.load_workbook(out_path, data_only=False, keep_vba=keep_vba_out)
    for name in ("InputFiles", "FatigueSelectorAudit", "JointForecastLong(Python)", "节点检验策略(Python)_raw"):
        if name in wb_out.sheetnames:
            wb_out.remove(wb_out[name])

    ws_input = wb_out.create_sheet("InputFiles")
    input_rows: List[Dict[str, Any]] = []

    def _append_input_row(category: str, fp: str | Path) -> None:
        path_obj = Path(fp).resolve()
        input_rows.append(
            {
                "Category": category,
                "Path": str(path_obj),
                "SizeBytes": int(path_obj.stat().st_size),
                "LineCount": _file_line_count(path_obj),
            }
        )

    _append_input_row("model", model_file)
    for p in clplog_paths:
        _append_input_row("clplog", p)
    for p in ftglst_paths:
        _append_input_row("ftglst", p)
    if ftginp_paths:
        for p in ftginp_paths:
            _append_input_row("ftginp", p)
    if manual_fill_workbook:
        mfw = Path(manual_fill_workbook)
        if mfw.exists():
            _append_input_row("manual_fill_workbook", mfw)

    input_df = pd.DataFrame(input_rows, columns=["Category", "Path", "SizeBytes", "LineCount"])
    for j, c in enumerate(input_df.columns, start=1):
        ws_input.cell(1, j).value = c
    for i, (_, rr) in enumerate(input_df.iterrows(), start=2):
        for j, c in enumerate(input_df.columns, start=1):
            ws_input.cell(i, j).value = rr[c]

    if ftginp_paths:
        ws_audit = wb_out.create_sheet("FatigueSelectorAudit")
        audit_df = fatigue_selector_audit_df.copy()
        if "ManualBrace" not in audit_df.columns:
            audit_df["ManualBrace"] = ""
        if "Applied" not in audit_df.columns:
            audit_df["Applied"] = ""
        for j, c in enumerate(audit_df.columns, start=1):
            ws_audit.cell(1, j).value = c
        for i, (_, rr) in enumerate(audit_df.iterrows(), start=2):
            for j, c in enumerate(audit_df.columns, start=1):
                ws_audit.cell(i, j).value = rr[c]

    ws_long = wb_out.create_sheet("JointForecastLong(Python)")
    long_df = forecast_long_df.copy()
    for j, c in enumerate(long_df.columns, start=1):
        ws_long.cell(1, j).value = c
    for i, (_, rr) in enumerate(long_df.iterrows(), start=2):
        for j, c in enumerate(long_df.columns, start=1):
            ws_long.cell(i, j).value = rr[c]

    ws_plan_raw = wb_out.create_sheet("节点检验策略(Python)_raw")
    raw_df = node_plan_df.copy()
    for j, c in enumerate(raw_df.columns, start=1):
        ws_plan_raw.cell(1, j).value = c
    for i, (_, rr) in enumerate(raw_df.iterrows(), start=2):
        for j, c in enumerate(raw_df.columns, start=1):
            ws_plan_raw.cell(i, j).value = rr[c]

    wb_out.save(out_path)

    print(f"[OK] Wrote: {out_xlsx}")


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--template", required=True, help="xlsm template (to read matrices & params)")
    ap.add_argument("--data-dir", default="", help="Auto-discover inputs from data directory (model, clplogs, ftglsts, ftginps).")
    ap.add_argument("--model", default="", help="SACS model input (sacinp)")
    ap.add_argument(
        "--clplog",
        action="append",
        default=[],
        help=f"collapse log (clplog), can be provided multiple times, max {MAX_VBA_COLLAPSE_FILES}",
    )
    ap.add_argument(
        "--ftglst",
        action="append",
        default=[],
        help=f"fatigue list output (ftglst), can be provided multiple times, max {MAX_VBA_FATIGUE_FILES}",
    )
    ap.add_argument(
        "--ftginp",
        action="append",
        default=[],
        help=f"fatigue input file (ftginp), optional, can be provided multiple times, max {MAX_VBA_FTGINP_FILES}",
    )
    ap.add_argument(
        "--manual-fill-workbook",
        default="",
        help="Optional workbook path that contains user-filled FatigueSelectorAudit.ManualBrace values.",
    )
    ap.add_argument(
        "--interactive-manual-fill",
        action="store_true",
        help="Prompt ManualBrace by GUI input dialog (fallback terminal) when JSLC selector mapping is missing, then continue in one run.",
    )
    ap.add_argument("--out", required=True, help="output xlsx")
    ap.add_argument("--policy", default="strict", choices=["strict","loose"], help="inspection policy mode")
    ap.add_argument("--seed", type=int, default=42, help="random seed for sampling")
    ap.add_argument(
        "--enable-topology-inference",
        action="store_true",
        help="Enable non-VBA topology inference for fatigue selector completion (default: disabled for strict VBA).",
    )
    args = ap.parse_args()

    if args.data_dir:
        bundle = discover_data_bundle(args.data_dir)
        model_file = bundle["model_file"]
        clplogs = bundle["clplog_files"]
        ftglsts = bundle["ftglst_files"]
        ftginps = bundle["ftginp_files"]
        print(f"[INFO] model: {model_file}")
        print(f"[INFO] clplog files: {len(clplogs)}")
        print(f"[INFO] ftglst files: {len(ftglsts)}")
        print(f"[INFO] ftginp files: {len(ftginps)}")
    else:
        if not args.model:
            raise ValueError("--model is required when --data-dir is not provided")
        if not args.clplog:
            raise ValueError("at least one --clplog is required when --data-dir is not provided")
        if not args.ftglst:
            raise ValueError("at least one --ftglst is required when --data-dir is not provided")
        model_file = args.model
        clplogs = args.clplog
        ftglsts = args.ftglst
        ftginps = args.ftginp

    run(
        template_xlsm=args.template,
        model_file=model_file,
        clplog_file=clplogs,
        ftglst_file=ftglsts,
        out_xlsx=args.out,
        policy_mode=args.policy,
        seed=args.seed,
        ftginp_files=ftginps,
        manual_fill_workbook=(args.manual_fill_workbook or None),
        interactive_manual_fill=bool(args.interactive_manual_fill),
        enable_topology_inference=args.enable_topology_inference,
    )

if __name__ == "__main__":
    main()
