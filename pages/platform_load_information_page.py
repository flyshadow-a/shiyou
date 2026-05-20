# -*- coding: utf-8 -*-
# platform_load_information_page.py
#
# 本版改进：
# 1) 顶部表头固定不变：分公司、作业公司、油气田、设施编码、设施名称、设施类型、分类、投产时间、设计年限
#    下拉框在表头下方一行；选中后即显示在该单元格中（无需第三行）。
# 2) 主表前两行“所属信息区”仅保留【字段名 + 值】，不再显示绿色提示。
# 3) 主表从“序号0”开始的内容允许用户直接编辑录入（填表说明区不可编辑）。
# 4) 红色字段（Fx~Mz、操作工况、极端工况）支持：
#    - 用户直接编辑录入
#    - 或通过“读取结果文件”接口从CSV读取并回填
# 5) 新增曲线图页面：与当前页面平级（新Tab），标题动态为：
#    “设施编码 + 平台重量中心变化曲线”，展示 3×3 曲线图。
#
# 说明：本文件不依赖其它页面模块，直接可被 nav_config 引用。

import os
import ctypes
import random
import re
import subprocess
import sys
import openpyxl
from typing import List, Tuple, Dict, Optional

from PyQt5.QtWidgets import (
    QAction,
    QAbstractItemView,
    QApplication,
    QWidget, QHBoxLayout, QVBoxLayout, QPushButton, QComboBox, QLabel,
    QTableWidget, QTableWidgetItem, QScrollArea, QMessageBox,
    QHeaderView, QToolTip, QFileDialog, QGridLayout, QMenu, QSizePolicy, QFrame,
    QButtonGroup, QRadioButton, QCheckBox,
)
from PyQt5.QtCore import Qt, QEvent
from PyQt5.QtGui import QColor, QFont, QFontMetrics

from core.app_paths import existing_dirs, external_path, first_existing_path
from core.base_page import BasePage
from core.dropdown_bar import DropdownBar  # 复用平台基本信息页的顶部下拉条样式
from core.message_boxes import ask_yes_no
from pages.hover_tip_table import HoverTipTable
from pages.file_management_platforms import default_platform, sync_platform_dropdowns
from services.inspection_business_db_adapter import load_facility_profile
from services.inspection_business_db_adapter import (
    load_platform_load_information_items,
    replace_platform_load_information_items,
)

# 上部组块分项目计算表页面（右键重量/重心单元格跳转编辑）
try:
    from upper_block_subproject_calculation_table_page import UpperBlockSubprojectCalculationTablePage
except Exception:
    # 兼容 pages 包内运行
    from pages.upper_block_subproject_calculation_table_page import UpperBlockSubprojectCalculationTablePage


# matplotlib 嵌入
from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
from matplotlib.figure import Figure
import matplotlib as mpl
from matplotlib import font_manager

try:
    import pandas as pd
except ImportError:
    pd = None

def _setup_chinese_matplotlib_font():
    """为 matplotlib 设置中文字体（避免标题/坐标轴中文变成方块）。

    会在系统已安装字体里按优先级挑选一个可用的中文字体，并设置：
    - mpl.rcParams['font.sans-serif']
    - mpl.rcParams['axes.unicode_minus'] = False
    """
    preferred = [
        "Microsoft YaHei",      # Windows 常见
        "SimHei",               # Windows 常见
        "PingFang SC",          # macOS 常见
        "Heiti SC",             # macOS 常见
        "Noto Sans CJK SC",     # Linux 常见
        "WenQuanYi Zen Hei",    # Linux 常见
        "Source Han Sans SC",   # 思源黑体
        "Arial Unicode MS",
    ]
    installed = {f.name for f in font_manager.fontManager.ttflist}
    for name in preferred:
        if name in installed:
            mpl.rcParams['font.sans-serif'] = [name]
            mpl.rcParams['axes.unicode_minus'] = False
            mpl.rcParams['font.size'] = 12
            mpl.rcParams['axes.titlesize'] = 12
            mpl.rcParams['axes.labelsize'] = 12
            mpl.rcParams['xtick.labelsize'] = 12
            mpl.rcParams['ytick.labelsize'] = 12
            return
    # 若找不到中文字体，则至少保证负号显示正常
    mpl.rcParams['axes.unicode_minus'] = False
    mpl.rcParams['font.size'] = 12
    mpl.rcParams['axes.titlesize'] = 12
    mpl.rcParams['axes.labelsize'] = 12
    mpl.rcParams['xtick.labelsize'] = 12
    mpl.rcParams['ytick.labelsize'] = 12



class SimpleLineChart(FigureCanvas):
    """一个简单折线图控件（matplotlib 默认样式/颜色）。"""

    _font_inited = False
    SMALL_FOUR_PT = 12

    def __init__(self, title: str, x: List[float], y: List[float], xlabel: str = "改建次数", ylabel: str = "", parent=None):
        if not SimpleLineChart._font_inited:
            _setup_chinese_matplotlib_font()
            SimpleLineChart._font_inited = True
        fig = Figure(figsize=(3.2, 2.2), dpi=100)
        self.ax = fig.add_subplot(111)
        super().__init__(fig)
        self.setParent(parent)
        self.setWindowFlags(Qt.Widget)

        self._title = title
        self._xlabel = xlabel
        self._ylabel = ylabel
        self._line = None

        self._init_axes()
        self.update_data(x, y)

    def _init_axes(self):
        self.ax.set_title(self._title, fontsize=self.SMALL_FOUR_PT)
        self.ax.set_xlabel(self._xlabel, fontsize=self.SMALL_FOUR_PT)
        if self._ylabel:
            self.ax.set_ylabel(self._ylabel, fontsize=self.SMALL_FOUR_PT)
        self.ax.tick_params(axis="both", labelsize=self.SMALL_FOUR_PT)
        self.ax.grid(True, linewidth=0.6, alpha=0.6)

    def update_data(self, x: List[float], y: List[float]):
        if self._line is None:
            (self._line,) = self.ax.plot(x, y, marker="o")
        else:
            self._line.set_data(x, y)
        self.ax.relim()
        self.ax.autoscale_view()
        ticks = [int(v) for v in x]
        self.ax.set_xticks(ticks)
        if ticks:
            right = max(ticks) if max(ticks) > 0 else 1
            self.ax.set_xlim(0, right)
        else:
            self.ax.set_xlim(0, 1)
        self.figure.tight_layout()
        self.draw_idle()


class PlatformWeightCenterCurvePage(BasePage):
    """曲线图页面：3×3 网格（与截图一致）。"""
    def __init__(self, facility_code: str, series: Dict[str, List[float]], parent=None):
        super().__init__("", parent)
        self.facility_code = facility_code
        self.series = series
        self._build_ui()

    def _build_ui(self):
        # 整体滚动
        scroll = QScrollArea(self)
        scroll.setWidgetResizable(True)
        self.main_layout.addWidget(scroll, 1)

        container = QWidget()
        scroll.setWidget(container)
        root = QVBoxLayout(container)
        root.setContentsMargins(10, 10, 10, 10)
        root.setSpacing(10)

        grid_wrap = QWidget()
        grid = QGridLayout(grid_wrap)
        grid.setContentsMargins(0, 0, 0, 0)
        grid.setHorizontalSpacing(12)
        grid.setVerticalSpacing(12)

        x = list(range(len(self.series.get("idx", []))))
        # 9张图（3×3）
        charts = [
            ("上部组块重量t", "weight", "上部组块重量t"),
            ("上部组块重心x(m)", "cgx", "上部组块重心x(m)"),
            ("上部组块重心y(m)", "cgy", "上部组块重心y(m)"),
            ("极端工况最大载荷Fx(KN)", "fx", "极端工况最大载荷Fx(KN)"),
            ("极端工况最大载荷Fy(KN)", "fy", "极端工况最大载荷Fy(KN)"),
            ("极端工况最大载荷Fz(KN)", "fz", "极端工况最大载荷Fz(KN)"),
            ("极端工况最大载荷Mx(KN·m)", "mx", "极端工况最大载荷Mx(KN·m)"),
            ("极端工况最大载荷My(KN·m)", "my", "极端工况最大载荷My(KN·m)"),
            ("极端工况最大载荷Mz(KN·m)", "mz", "极端工况最大载荷Mz(KN·m)"),
        ]

        for i, (title, key, unit) in enumerate(charts):
            y = self.series.get(key, [])
            # 长度对齐
            if len(y) != len(x):
                y = (y + [0.0] * len(x))[:len(x)]
            canvas = SimpleLineChart(title, x, y, xlabel="改建次序", ylabel=unit)
            # 外框
            holder = QWidget()
            holder.setStyleSheet("background:#dfe9f6; border:1px solid #b6c2d6;")
            v = QVBoxLayout(holder)
            v.setContentsMargins(6, 6, 6, 6)
            v.addWidget(canvas)
            r, c = divmod(i, 3)
            grid.addWidget(holder, r, c)

        root.addWidget(grid_wrap, 1)


class PlatformWeightCenterCurveWidget(QWidget):
    """页面内嵌曲线区：3×3 网格。"""

    CHART_MIN_WIDTH = 360
    CHART_MIN_HEIGHT = 260

    CHARTS = [
        ("上部组块重量t", "weight", "上部组块重量t"),
        ("上部组块重心x(m)", "cgx", "上部组块重心x(m)"),
        ("上部组块重心y(m)", "cgy", "上部组块重心y(m)"),
        ("极端工况最大载荷Fx(KN)", "fx", "极端工况最大载荷Fx(KN)"),
        ("极端工况最大载荷Fy(KN)", "fy", "极端工况最大载荷Fy(KN)"),
        ("极端工况最大载荷Fz(KN)", "fz", "极端工况最大载荷Fz(KN)"),
        ("极端工况最大载荷Mx(KN·m)", "mx", "极端工况最大载荷Mx(KN·m)"),
        ("极端工况最大载荷My(KN·m)", "my", "极端工况最大载荷My(KN·m)"),
        ("极端工况最大载荷Mz(KN·m)", "mz", "极端工况最大载荷Mz(KN·m)"),
    ]

    def __init__(self, facility_code: str, series: Dict[str, List[float]], parent=None):
        super().__init__(parent)
        self.title_label: Optional[QLabel] = None
        self.grid: Optional[QGridLayout] = None
        self._chart_canvases: Dict[str, SimpleLineChart] = {}
        self._build_ui()
        self.update_series(facility_code, series)

    def _build_ui(self):
        root = QVBoxLayout(self)
        root.setContentsMargins(0, 0, 0, 0)
        root.setSpacing(10)

        self.title_label = QLabel()
        self.title_label.setStyleSheet("color:#1d2b3a;")
        root.addWidget(self.title_label, 0)

        grid_wrap = QWidget(self)
        self.grid = QGridLayout(grid_wrap)
        self.grid.setContentsMargins(0, 0, 0, 0)
        self.grid.setHorizontalSpacing(12)
        self.grid.setVerticalSpacing(12)
        root.addWidget(grid_wrap, 1)

    def update_series(self, facility_code: str, series: Dict[str, List[float]]):
        if self.title_label is not None:
            self.title_label.setFont(PlatformLoadInformationPage._songti_small_four_font(bold=True))
            self.title_label.setText(f"{facility_code}平台重量中心变化曲线")

        if self.grid is None:
            return

        x = list(range(len(series.get("idx", []))))
        for i, (title, key, unit) in enumerate(self.CHARTS):
            y = series.get(key, [])
            if len(y) != len(x):
                y = (y + [0.0] * len(x))[:len(x)]
            canvas = self._chart_canvases.get(key)
            if canvas is None:
                holder = QWidget(self)
                holder.setStyleSheet("background:#dfe9f6; border:1px solid #b6c2d6;")
                holder.setMinimumSize(self.CHART_MIN_WIDTH, self.CHART_MIN_HEIGHT)
                holder.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
                layout = QVBoxLayout(holder)
                layout.setContentsMargins(6, 6, 6, 6)
                canvas = SimpleLineChart(title, x, y, xlabel="改建次序", ylabel=unit, parent=holder)
                layout.addWidget(canvas)
                row, col = divmod(i, 3)
                self.grid.addWidget(holder, row, col)
                self._chart_canvases[key] = canvas
            else:
                canvas.update_data(x, y)


class PlatformLoadInformationPage(BasePage):
    """平台载荷信息页面（严格表格结构 + 顶部/所属信息联动 + 结果文件读取 + 曲线页面）。"""
    MAX_EXPAND_ROWS = 55  # 与 summary_information_table_page 保持一致
    DATA_START_ROW = 2

    TOP_KEY_ORDER: List[str] = [
        "branch", "op_company", "oilfield", "facility_code", "facility_name",
        "facility_type", "category", "start_time", "design_life",
    ]

    # 顶部固定表头（9字段）
    TOP_FIELDS: List[Tuple[str, str]] = [
        ("分公司", "湛江分公司"),
        ("作业公司", "文昌油田群作业公司"),
        ("油气田", "文昌19-1油田"),
        ("设施编码", "WC19-1WHPC"),
        ("设施名称", "文昌19-1WHPC井口平台"),
        ("设施类型", "平台"),
        ("分类", "井口平台"),
        ("投产时间", "2013-07-15"),
        ("设计年限", "15"),
    ]

    # DropdownBar 的 key 与“所属信息区”字段名（中文）映射
    KEY_TO_FIELD: Dict[str, str] = {
        "branch": "分公司",
        "op_company": "作业公司",
        "oilfield": "油气田",
        "facility_code": "设施编码",
        "facility_name": "设施名称",
        "facility_type": "设施类型",
        "category": "分类",
        "start_time": "投产时间",
        "design_life": "设计年限",
    }
    FIELD_TO_KEY: Dict[str, str] = {v: k for k, v in KEY_TO_FIELD.items()}

    @staticmethod
    def _songti_small_four_font(bold: bool = False) -> QFont:
        font = QFont("SimSun")
        font.setPointSize(12)
        font.setBold(bold)
        return font

    @staticmethod
    def _menu_qss() -> str:
        return """
            QMenu {
                background-color: #ffffff;
                color: #1d2b3a;
                border: 1px solid #cfd8e3;
                padding: 4px 0;
            }
            QMenu::item {
                padding: 6px 18px;
                background-color: transparent;
                color: #1d2b3a;
            }
            QMenu::item:selected {
                background-color: #dbe9ff;
                color: #1d2b3a;
            }
            QMenu::item:disabled {
                color: #8a94a6;
                background-color: #f7f9fc;
            }
        """


    def __init__(self, parent=None):
        super().__init__("", parent)
        self.data_dir = first_existing_path("data")
        self.output_data_dir = external_path("data")
        self._num_pat = r"[-+]?\d+(?:\.\d+)?(?:[eE][-+]?\d+)?"

        # === 红色字段 Excel 导入（仅当前行）===
        self.result_excel_path: Optional[str] = None
        self.red_field_mode: str = 'manual'  # 'manual' or 'excel'
        # 数据区行勾选（单选）：用于“读取结果文件”定位目标行
        self._row_radio_group: Optional[QButtonGroup] = None
        # 缓存子计算表用户输入的数据
        self._uppercalc_saved_data: Dict[int, dict] = {}
        self._result_factor_cache: Dict[tuple[str, float], Dict[str, object]] = {}
        self._switching_platform = False
        self._build_ui()
        self._ensure_demo_files()
        self._load_current_platform_data()

    # ---------------- UI ----------------
    def _build_ui(self):
        self.setObjectName("PlatformLoadInfoRoot")
        self.setStyleSheet("""
            QWidget#PlatformLoadInfoRoot,
            QWidget#PlatformLoadInfoTopWrap,
            QWidget#PlatformLoadInfoPageContainer {
                background: #f3f6fb;
            }

            QFrame#LoadInfoTablePanel {
                background: #ffffff;
                border: 1px solid #dbe5f3;
                border-radius: 18px;
            }

            QFrame#MetaInfoPanel {
                background: #f8fbff;
                border: 1px solid #e0e8f4;
                border-radius: 12px;
            }

            QLabel.MetaName {
                background: #edf4ff;
                color: #24476f;
                border: 1px solid #d8e5f5;
                border-radius: 6px;
                padding: 6px 10px;
                font-weight: bold;
            }

            QLabel.MetaValue {
                background: #ffffff;
                color: #253044;
                border: 1px solid #dfe7f1;
                border-radius: 6px;
                padding: 6px 10px;
                font-weight: bold;
            }

            QPushButton#TopActionBtn {
                background: #2563eb;
                color: #ffffff;
                border: 1px solid #1d4ed8;
                border-radius: 8px;
                padding: 7px 18px;
                font-family: "SimSun", "NSimSun", "宋体", "Microsoft YaHei UI", "Microsoft YaHei";
                font-size: 12pt;
                font-weight: bold;
            }
            QPushButton#TopActionBtn:hover { background: #1d4ed8; }
            QPushButton#TopActionBtn:pressed { background: #1e40af; }

            QTableWidget#MainTable {
                background: #ffffff;
                border: 1px solid #dce6f2;
                border-radius: 12px;
                gridline-color: #e8eef7;
                alternate-background-color: #fbfdff;
                selection-background-color: #dcecff;
                selection-color: #1d3557;
                font-family: "SimSun", "NSimSun", "宋体", "Microsoft YaHei UI", "Microsoft YaHei";
                font-size: 12pt;
                color: #253044;
            }
            QTableWidget#MainTable::item {
                padding: 6px;
                border-color: #e8eef7;
            }
            QTableWidget#MainTable::item:selected {
                background-color: #dcecff;
                color: #1d3557;
            }
            QTableWidget::item:focus { outline: none; }

            QWidget#RowSelectorCell {
                background: transparent;
            }
            QLabel#RowSeqLabel {
                color: #344054;
                background: transparent;
            }
            QCheckBox::indicator {
                width: 15px;
                height: 15px;
                border-radius: 4px;
                border: 1px solid #9fb2cc;
                background: #ffffff;
            }
            QCheckBox::indicator:hover {
                border: 1px solid #2563eb;
                background: #eff6ff;
            }
            QCheckBox::indicator:checked {
                border: 1px solid #2563eb;
                background: #2563eb;
            }
            QScrollBar:horizontal, QScrollBar:vertical {
                background: #f5f8fc;
                border: none;
                margin: 0px;
            }
            QScrollBar:horizontal { height: 12px; }
            QScrollBar:vertical { width: 12px; }
            QScrollBar::handle:horizontal, QScrollBar::handle:vertical {
                background: #c5d3e6;
                border-radius: 6px;
                min-width: 28px;
                min-height: 28px;
            }
            QScrollBar::handle:horizontal:hover, QScrollBar::handle:vertical:hover {
                background: #93acd0;
            }
            QScrollBar::add-line, QScrollBar::sub-line,
            QScrollBar::add-page, QScrollBar::sub-page {
                background: transparent;
                border: none;
            }

        """)

        page_scroll = QScrollArea(self)
        page_scroll.setWidgetResizable(True)
        page_scroll.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        page_scroll.setFrameShape(QFrame.NoFrame)
        self.main_layout.addWidget(page_scroll, 1)
        self.page_scroll = page_scroll

        page_container = QWidget()
        page_container.setObjectName("PlatformLoadInfoPageContainer")
        page_scroll.setWidget(page_container)
        page_root = QVBoxLayout(page_container)
        page_root.setContentsMargins(0, 0, 0, 0)
        page_root.setSpacing(0)

        # 顶部固定区：表头+下拉（2行） + 右侧按钮
        top_wrap = QWidget()
        top_wrap.setObjectName("PlatformLoadInfoTopWrap")
        top_layout = QHBoxLayout(top_wrap)
        top_layout.setContentsMargins(10, 10, 10, 0)
        top_layout.setSpacing(10)

        top_layout.setAlignment(Qt.AlignTop)
        platform_defaults = default_platform()
        profile = load_facility_profile(platform_defaults["facility_code"], defaults=platform_defaults)
        fallback_defaults = {
            "分公司": str(profile.get("branch") or self.TOP_FIELDS[0][1]),
            "作业公司": str(profile.get("op_company") or self.TOP_FIELDS[1][1]),
            "油气田": str(profile.get("oilfield") or self.TOP_FIELDS[2][1]),
            "设施编码": str(profile.get("facility_code") or self.TOP_FIELDS[3][1]),
            "设施名称": str(profile.get("facility_name") or self.TOP_FIELDS[4][1]),
            "设施类型": str(profile.get("facility_type") or self.TOP_FIELDS[5][1]),
            "分类": str(profile.get("category") or self.TOP_FIELDS[6][1]),
            "投产时间": str(profile.get("start_time") or self.TOP_FIELDS[7][1]),
            "设计年限": str(profile.get("design_life") or self.TOP_FIELDS[8][1]),
        }
        stretch_map = {
            "branch": 1,
            "op_company": 2,
            "oilfield": 2,
            "facility_code": 2,
            "facility_name": 3,
            "facility_type": 1,
            "category": 1,
            "start_time": 2,
            "design_life": 1,
        }
        fields = []
        for key in self.TOP_KEY_ORDER:
            label = self.KEY_TO_FIELD[key]
            fallback = fallback_defaults.get(label, "")
            opts = [fallback] if fallback else []
            default = fallback
            fields.append({
                "key": key,
                "label": label,
                "options": opts,
                "default": default,
                "stretch": stretch_map.get(key, 1),
            })

        self.dropdown_bar = DropdownBar(fields, parent=self)
        self.dropdown_bar.valueChanged.connect(self._on_top_key_changed)
        self._sync_platform_ui()
        top_layout.addWidget(self.dropdown_bar, 1)

        self.btn_save = QPushButton("保存")
        self.btn_export = QPushButton("导出数据")

        for b in (self.btn_save, self.btn_export):
            b.setObjectName("TopActionBtn")
            b.setFont(self._songti_small_four_font(bold=True))
            b.setMinimumHeight(32)

        # 按钮尺寸调整
        self.btn_save.setMinimumWidth(100)
        self.btn_export.setMinimumWidth(100)

        self.btn_save.clicked.connect(self._on_save)
        self.btn_export.clicked.connect(self._on_export)

        page_root.addWidget(top_wrap, 0)

        table_panel = QFrame()
        table_panel.setObjectName("LoadInfoTablePanel")
        table_panel.setMinimumHeight(520)
        root = QVBoxLayout(table_panel)
        root.setContentsMargins(18, 18, 18, 16)
        root.setSpacing(14)
        page_root.addWidget(table_panel, 0)
        self.table_panel = table_panel

        root.addWidget(self._build_meta_info_panel(), 0)

        self.table = self._build_main_table_skeleton()
        self.table.setObjectName("MainTable")
        self.table.setFont(self._songti_small_four_font())
        self.table.setAlternatingRowColors(True)
        self.table.setHorizontalScrollMode(QAbstractItemView.ScrollPerPixel)
        self.table.setVerticalScrollMode(QAbstractItemView.ScrollPerPixel)
        self.table.setHorizontalScrollBarPolicy(Qt.ScrollBarAsNeeded)
        self.table.setVerticalScrollBarPolicy(Qt.ScrollBarAsNeeded)

        # 开启右键菜单策略
        self.table.setContextMenuPolicy(Qt.CustomContextMenu)
        self.table.customContextMenuRequested.connect(self._on_table_context_menu)

        # 监听单元格变动，用于手动修改时恢复背景色
        self.table.itemChanged.connect(self._on_item_changed)

        # 主表允许编辑（但我们会用 item flags 精细控制哪些格可编辑）
        self.table.setEditTriggers(QTableWidget.DoubleClicked | QTableWidget.SelectedClicked | QTableWidget.EditKeyPressed)

        root.addWidget(self.table, 1)

        # 底部按钮区：保存、导出，横向居中排列
        bottom_btn_wrap = QWidget()
        bottom_btn_lay = QHBoxLayout(bottom_btn_wrap)
        bottom_btn_lay.setContentsMargins(0, 10, 0, 0)
        bottom_btn_lay.setSpacing(15)
        
        bottom_btn_lay.addStretch(1)
        bottom_btn_lay.addWidget(self.btn_save)
        bottom_btn_lay.addWidget(self.btn_export)
        bottom_btn_lay.addStretch(1)
        
        root.addWidget(bottom_btn_wrap, 0)

        self.curve_section = QWidget()
        curve_section_layout = QVBoxLayout(self.curve_section)
        curve_section_layout.setContentsMargins(10, 10, 10, 10)
        curve_section_layout.setSpacing(0)

        curve_wrap = QWidget()
        curve_wrap.setStyleSheet("background:#ffffff; border:none;")
        curve_layout = QVBoxLayout(curve_wrap)
        curve_layout.setContentsMargins(12, 12, 12, 12)
        curve_layout.setSpacing(10)
        self.curve_widget = PlatformWeightCenterCurveWidget(
            self._get_top_value("设施编码") or "XXXX",
            self._collect_series_for_curve(),
            curve_wrap,
        )
        curve_layout.addWidget(self.curve_widget)
        curve_section_layout.addWidget(curve_wrap)
        page_root.addWidget(self.curve_section, 0)

    # ---------------- 顶部表：固定表头 + 下拉行 ----------------
    def _build_top_header_combo_table(self) -> QTableWidget:
        fields = self.TOP_FIELDS
        col_count = len(fields)

        table = QTableWidget(2, col_count)
        table.verticalHeader().setVisible(False)
        table.horizontalHeader().setVisible(False)
        table.setEditTriggers(QTableWidget.NoEditTriggers)
        table.setSelectionMode(QTableWidget.NoSelection)
        table.setFocusPolicy(Qt.NoFocus)

        table.setRowHeight(0, 88)
        table.setRowHeight(1, 45)
        table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)

        self._top_combos: List[QComboBox] = []
        for c, (title, default) in enumerate(fields):
            # 固定表头（用 QLabel 作为 cellWidget，避免全局 QSS 覆盖 item 背景导致不生效）
            lab = QLabel(title)
            lab.setAlignment(Qt.AlignCenter)
            lab.setStyleSheet(
                "background-color:#00BFFF;"
                "color:#001018;"
                "border:1px solid #7f8ea3;"
                "font-weight:bold;"
            )
            table.setCellWidget(0, c, lab)

            # 下拉框（值）
            cb = QComboBox()
            cb.setEditable(False)
            cb.setMaximumHeight(28)
            cb.addItems(self._mock_top_options(title, default))
            cb.setCurrentText(default)
            cb.currentTextChanged.connect(lambda txt, field=title: self._on_top_field_changed(field, txt))
            table.setCellWidget(1, c, cb)
            self._top_combos.append(cb)

        return table

    def _on_top_field_changed(self, field: str, txt: str):
        """顶部下拉改变：同步回填主表所属信息区。"""
        self._sync_meta_value(field, txt)

    def _sync_meta_value(self, field: str, txt: str):
        if not hasattr(self, "_meta_value_items"):
            return
        it = self._meta_value_items.get(field)
        if it is None:
            return
        it.setText(txt)

    def _build_meta_info_panel(self) -> QFrame:
        panel = QFrame()
        panel.setObjectName("MetaInfoPanel")
        grid = QGridLayout(panel)
        grid.setContentsMargins(10, 10, 10, 10)
        grid.setHorizontalSpacing(8)
        grid.setVerticalSpacing(8)

        self._meta_value_items = {}
        defaults = self._current_top_defaults()
        rows = [
            [
                ("所属分公司", defaults.get("分公司", ""), "分公司"),
                ("所属作业单元", defaults.get("作业公司", ""), "作业公司"),
                ("所属油（气）田", defaults.get("油气田", ""), "油气田"),
            ],
            [
                ("设施名称", defaults.get("设施名称", ""), "设施名称"),
                ("投产时间", defaults.get("投产时间", ""), "投产时间"),
                ("设计年限", defaults.get("设计年限", ""), "设计年限"),
            ],
        ]

        for r, row in enumerate(rows):
            for group, (name, value, field_key) in enumerate(row):
                name_label = QLabel(name)
                name_label.setProperty("class", "MetaName")
                name_label.setAlignment(Qt.AlignRight | Qt.AlignVCenter)
                name_label.setFont(self._songti_small_four_font(bold=True))
                value_label = QLabel(value)
                value_label.setProperty("class", "MetaValue")
                value_label.setAlignment(Qt.AlignLeft | Qt.AlignVCenter)
                value_label.setFont(self._songti_small_four_font(bold=True))
                value_label.setWordWrap(True)

                base_col = group * 2
                grid.addWidget(name_label, r, base_col)
                grid.addWidget(value_label, r, base_col + 1)
                self._meta_value_items[field_key] = value_label

        for col in range(6):
            grid.setColumnStretch(col, 1 if col % 2 == 0 else 3)

        return panel

    def _normalize_top_value(self, value: object) -> str:
        txt = "" if value is None else str(value).strip()
        if (not txt) or (txt.lower() == "nan"):
            return ""
        if re.fullmatch(r"[-+]?\d+\.0+", txt):
            txt = txt.split(".", 1)[0]
        return txt

    def _load_top_records_from_excel(self) -> List[Dict[str, str]]:
        if (not getattr(self, "_excel_loaded", False)) or (not hasattr(self._excel_provider, "df")):
            return []

        df = self._excel_provider.df
        if df is None:
            return []

        fields = [f for f, _ in self.TOP_FIELDS]
        resolved: Dict[str, str] = {}
        for field in fields:
            col = self._excel_provider._resolve_col(field) if hasattr(self._excel_provider, "_resolve_col") else None
            if not col:
                return []
            resolved[field] = col

        records: List[Dict[str, str]] = []
        seen = set()
        for _, row in df.iterrows():
            rec: Dict[str, str] = {}
            for field, col in resolved.items():
                raw = self._excel_provider._clean(row[col]) if hasattr(self._excel_provider, "_clean") else row[col]
                rec[field] = self._normalize_top_value(raw)

            if not any(rec.get(k) for k in ("分公司", "作业公司", "油气田", "设施编码", "设施名称")):
                continue

            uniq = tuple(rec.get(f, "") for f in fields)
            if uniq in seen:
                continue
            seen.add(uniq)
            records.append(rec)

        return records

    def _unique_record_values(self, records: List[Dict[str, str]], field: str) -> List[str]:
        out: List[str] = []
        seen = set()
        for rec in records:
            v = self._normalize_top_value(rec.get(field, ""))
            if (not v) or (v in seen):
                continue
            seen.add(v)
            out.append(v)
        return out

    def _pick_option(self, options: List[str], preferred: str = "") -> str:
        p = self._normalize_top_value(preferred)
        if p and p in options:
            return p
        return options[0] if options else ""

    def _sync_platform_ui(self, changed_key: str | None = None):
        if not hasattr(self, "dropdown_bar"):
            return
        platform = sync_platform_dropdowns(self.dropdown_bar, changed_key=changed_key)
        profile = load_facility_profile(
            platform["facility_code"],
            defaults={
                "branch": platform["branch"],
                "op_company": platform["op_company"],
                "oilfield": platform["oilfield"],
                "facility_code": platform["facility_code"],
                "facility_name": platform["facility_name"],
                "facility_type": platform["facility_type"],
                "category": platform["category"],
                "start_time": platform["start_time"],
                "design_life": platform["design_life"],
            },
        )
        self.dropdown_bar.set_options("branch", [profile["branch"]], profile["branch"])
        self.dropdown_bar.set_options("op_company", [profile["op_company"]], profile["op_company"])
        self.dropdown_bar.set_options("oilfield", [profile["oilfield"]], profile["oilfield"])
        self.dropdown_bar.set_options("facility_type", [profile["facility_type"]], profile["facility_type"])
        self.dropdown_bar.set_options("category", [profile["category"]], profile["category"])
        self.dropdown_bar.set_options("start_time", [profile["start_time"]], profile["start_time"])
        self.dropdown_bar.set_options("design_life", [profile["design_life"]], profile["design_life"])
        self._sync_all_top_meta_values()

    def _sync_all_top_meta_values(self):
        for field_cn in self.FIELD_TO_KEY.keys():
            self._sync_meta_value(field_cn, self._get_top_value(field_cn))

    def _current_top_defaults(self) -> Dict[str, str]:
        defaults = {k: v for k, v in self.TOP_FIELDS}
        if hasattr(self, "dropdown_bar"):
            for field_cn in defaults.keys():
                cur = self._get_top_value(field_cn)
                if cur:
                    defaults[field_cn] = cur
        return defaults

    def _mock_top_options(self, field: str, default: str) -> List[str]:
        options_map = {
            "分公司": ["湛江分公司", "深圳分公司", "上海分公司", "海南分公司", "天津分公司"],
            "作业公司": ["文昌油田群作业公司", "涠洲作业公司", "珠江作业公司", "渤海作业公司"],
            "油气田": ["文昌19-1油田", "文昌19-2油田", "涠洲油田", "珠江口油田"],
            "设施编码": ["WC19-1WHPC", "WC19-2WHPC", "WC9-7DPP", "WC19-1DPPA"],
            "设施名称": ["文昌19-1WHPC井口平台", "文昌19-2WHPC井口平台", "WC9-7DPP井口平台"],
            "设施类型": ["平台", "导管架", "浮式"],
            "分类": ["井口平台", "生产平台", "生活平台"],
            "投产时间": ["2013-07-15", "2008-06-26", "2010-03-10"],
            "设计年限": ["10", "15", "20", "25", "30"],
        }
        opts = options_map.get(field, [default])
        return opts if default in opts else [default] + opts

    def _on_top_key_changed(self, key: str, txt: str):
        """
        顶部 DropdownBar 改变：
        - 级联刷新下拉值
        - 同步回填主表所属信息区（字段名用中文）
        """
        if key in {"branch", "op_company", "oilfield", "facility_code", "facility_name"}:
            self._sync_platform_ui(changed_key=key)
            self._load_current_platform_data()
            return

        field = self.KEY_TO_FIELD.get(key, key)
        self._sync_meta_value(field, txt)
        if key == "facility_code":
            self._load_current_platform_data()

    def _get_top_value(self, field: str) -> str:
        """获取顶部 DropdownBar 当前值（field 为中文表头名，如“设施编码”）。"""
        if hasattr(self, "dropdown_bar"):
            key = self.FIELD_TO_KEY.get(field)
            if key:
                return self.dropdown_bar.get_value(key)
        return ""


    # ---------------- 主表（严格结构） ----------------
    def _columns(self) -> List[str]:
        return [
            "序号",
            "改扩建项目名称",
            "改扩建时间",
            "改扩建内容",
            "上部组块总\n操作重量\n(MT)",
            "上部组块不\n可超越重量\n(MT)",
            "重量变化\n(MT)",
            "上部组块重心\n(x,y,z)\n(m)",
            "上部组块重心\n不可超越半径\n(m)",
            "Fx\n(KN)",
            "Fy\n(KN)",
            "Fz\n(KN)",
            "Mx\n(KN·m)",
            "My\n(KN·m)",
            "Mz\n(KN·m)",
            "操作工况",
            "极端工况",
            "是否整体\n评估",
            "评估机构",
        ]

    def _mk_item(self, text: str, *, editable: bool=False, bold: bool=False,
                 fg: QColor=None, bg: QColor=None,
                 align: Qt.AlignmentFlag = Qt.AlignCenter) -> QTableWidgetItem:
        it = QTableWidgetItem(text)
        flags = it.flags()
        if editable:
            flags |= Qt.ItemIsEditable
        else:
            flags &= ~Qt.ItemIsEditable
        it.setFlags(flags)
        it.setTextAlignment(align)
        it.setFont(self._songti_small_four_font(bold=bold))
        if fg is not None:
            it.setForeground(fg)
        if bg is not None:
            it.setBackground(bg)
        return it

    def _set_header_widget(
        self,
        table: HoverTipTable,
        row: int,
        col: int,
        text: str,
        *,
        bg: str = "#f6f7f9",
        align: Qt.AlignmentFlag = Qt.AlignCenter,
        bold: bool = False,
    ) -> QTableWidgetItem:
        item = self._mk_item(text, bold=bold, bg=QColor(bg), align=align)
        item.setForeground(QColor("#183b66"))
        table.setItem(row, col, item)
        return item

    def _build_main_table_skeleton(self) -> HoverTipTable:
        cols = self._columns()
        col_count = len(cols)

        # 行：0-1 表头；所属信息已拆到表格上方的表单头。
        base_rows = self.DATA_START_ROW
        table = HoverTipTable(base_rows, col_count)
        table.setFont(self._songti_small_four_font())
        table.verticalHeader().setVisible(False)
        table.horizontalHeader().setVisible(False)
        # table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)

        table.setRowHeight(0, 36)
        table.setRowHeight(1, 68)

        # ===== 分组表头（row2）=====
        table.setSpan(0, 0, 2, 1)
        self._set_header_widget(table, 0, 0, "序号", bg="#dbeafe")

        # table.setSpan(2, 1, 1, 3)
        # table.setItem(2, 1, self._mk_item("改扩建", bold=True, bg=bg))

        # table.setSpan(2, 4, 1, 5)
        # table.setItem(2, 4, self._mk_item("上部组块重控", bold=True, bg=bg))

        table.setSpan(0, 9, 1, 6)
        self._set_header_widget(table, 0, 9, "极端工况最大载荷", bg="#fed7aa")

        table.setSpan(0, 15, 1, 2)
        self._set_header_widget(table, 0, 15, "桩基承载力安全系数（最小）", bg="#bbf7d0")

        table.setSpan(0, 17, 2, 1)
        self._set_header_widget(table, 0, 17, "是否整体\n评估", bg="#dbeafe")
        table.setSpan(0, 18, 2, 1)
        self._set_header_widget(table, 0, 18, "评估机构", bg="#dbeafe")

        # ===== 子表头（row3）=====
        for i in range(1,9):
            table.setSpan(0, i, 2, 1)
        self._set_header_widget(table, 0, 1, "改扩建项目名称", bg="#dbeafe")
        self._set_header_widget(table, 0, 2, "改扩建时间", bg="#dbeafe")
        self._set_header_widget(table, 0, 3, "改扩建内容", bg="#dbeafe")

        self._set_header_widget(table, 0, 4, "上部组块总操作重量\n(MT)", bg="#dbeafe")
        self._set_header_widget(table, 0, 5, "上部组块不可超越重量\n(MT)", bg="#dbeafe")
        self._set_header_widget(table, 0, 6, "重量变化\n(MT)", bg="#dbeafe")
        self._set_header_widget(table, 0, 7, "上部组块重心\n(x,y,z)\n(m)", bg="#bbf7d0")
        self._set_header_widget(table, 0, 8, "上部组块重心\n不可超越半径\n(m)", bg="#dbeafe")

        # 结果字段（严格：Mz 纵向显示）
        fx_headers = [
            ("Fx", "Fx\n(KN)"),
            ("Fy", "Fy\n(KN)"),
            ("Fz", "Fz\n(KN)"),
            ("Mx", "Mx\n(KN·m)"),
            ("My", "My\n(KN·m)"),
            ("Mz", "Mz\n(KN·m)"),
        ]
        red_cols = list(range(9, 15))
        for (src, shown), c in zip(fx_headers, red_cols):
            self._set_header_widget(table, 1, c, shown, bg="#ffedd5")

        self._set_header_widget(table, 1, 15, "操作工况", bg="#dcfce7")
        self._set_header_widget(table, 1, 16, "极端工况", bg="#dcfce7")

        # 补齐背景（未被 span 覆盖的单元格）
        for r in range(base_rows):
            for c in range(col_count):
                if table.item(r, c) is None and table.cellWidget(r, c) is None:
                    self._set_header_widget(table, r, c, "", bg="#f6f9fd")
        return table

    # ---------------- 数据加载 ----------------
    def _blank_table_row(self) -> List[str]:
        row = [""] * len(self._columns())
        row[0] = "0"
        return row

    def _db_rows_to_table_rows(self, rows: List[Dict[str, object]]) -> List[List[str]]:
        table_rows: List[List[str]] = []
        for index, row in enumerate(rows):
            table_rows.append([
                str(row.get("seq_no", index)),
                str(row.get("project_name") or ""),
                str(row.get("rebuild_time") or ""),
                str(row.get("rebuild_content") or ""),
                str(row.get("total_weight_mt") or ""),
                str(row.get("weight_limit_mt") or ""),
                str(row.get("weight_delta_mt") or ""),
                str(row.get("center_xyz") or ""),
                str(row.get("center_radius_m") or ""),
                str(row.get("fx_kn") or ""),
                str(row.get("fy_kn") or ""),
                str(row.get("fz_kn") or ""),
                str(row.get("mx_kn_m") or ""),
                str(row.get("my_kn_m") or ""),
                str(row.get("mz_kn_m") or ""),
                str(row.get("safety_op") or ""),
                str(row.get("safety_extreme") or ""),
                str(row.get("overall_assessment") or ""),
                str(row.get("assessment_org") or ""),
            ])
        return table_rows

    def _format_weight_delta_text(self, value: object) -> str:
        text = "" if value is None else str(value).strip()
        if not text or text in ("\\", "/"):
            return text
        if text.startswith("+") or text.startswith("-"):
            return text
        normalized = text.replace(",", "")
        try:
            number = float(normalized)
        except ValueError:
            return text
        if number > 0:
            return f"+{text}"
        return text

    def _collect_table_rows_for_db(self) -> List[Dict[str, str]]:
        rows: List[Dict[str, str]] = []
        for sort_order, row in enumerate(range(self.DATA_START_ROW, self._find_data_end_row()), start=1):
            rows.append(
                {
                    "seq_no": self._cell_text(row, 0).strip() or str(sort_order - 1),
                    "project_name": self._cell_text(row, 1).strip(),
                    "rebuild_time": self._cell_text(row, 2).strip(),
                    "rebuild_content": self._cell_text(row, 3).strip(),
                    "total_weight_mt": self._cell_text(row, 4).strip(),
                    "weight_limit_mt": self._cell_text(row, 5).strip(),
                    "weight_delta_mt": self._cell_text(row, 6).strip(),
                    "center_xyz": self._cell_text(row, 7).strip(),
                    "center_radius_m": self._cell_text(row, 8).strip(),
                    "fx_kn": self._cell_text(row, 9).strip(),
                    "fy_kn": self._cell_text(row, 10).strip(),
                    "fz_kn": self._cell_text(row, 11).strip(),
                    "mx_kn_m": self._cell_text(row, 12).strip(),
                    "my_kn_m": self._cell_text(row, 13).strip(),
                    "mz_kn_m": self._cell_text(row, 14).strip(),
                    "safety_op": self._cell_text(row, 15).strip(),
                    "safety_extreme": self._cell_text(row, 16).strip(),
                    "overall_assessment": self._cell_text(row, 17).strip(),
                    "assessment_org": self._cell_text(row, 18).strip(),
                    "sort_order": str(sort_order),
                }
            )
        return rows

    def _load_current_platform_data(self):
        if not hasattr(self, "table") or self._switching_platform:
            return
        facility_code = self._get_top_value("设施编码").strip()
        if not facility_code:
            return

        self._switching_platform = True
        try:
            rows = load_platform_load_information_items(facility_code)
            if rows:
                self._apply_data(self._db_rows_to_table_rows(rows))
            else:
                self._apply_data([self._blank_table_row()])
        except Exception as exc:
            QMessageBox.warning(self, "读取失败", f"读取平台载荷信息数据库数据失败：\n{exc}")
            self._apply_data([self._blank_table_row()])
        finally:
            self._switching_platform = False

    def _find_data_end_row(self) -> int:
        """找到填表说明起始行（不含），作为数据区结束行。"""
        for r in range(self.table.rowCount()):
            it = self.table.item(r, 0)
            if it and it.text().startswith("填表说明"):
                return r
        return self.table.rowCount()

    def _rebuild_row_checkbox_selectors(self, data_start: int, data_end: int):
        """在数据区每行首列放置复选框，用于批量操作。"""
        # 移除旧的 cellWidget
        for r in range(self.table.rowCount()):
            self.table.removeCellWidget(r, 0)

        self._row_checkboxes: Dict[int, QCheckBox] = {}

        for row_idx in range(data_start, data_end):
            seq_item = self.table.item(row_idx, 0)
            seq_text = seq_item.text().strip() if seq_item else ""
            if seq_item is not None:
                seq_item.setText("")

            cb = QCheckBox()
            cb.setCursor(Qt.PointingHandCursor)
            # 记录复选框引用
            self._row_checkboxes[row_idx] = cb

            holder = QWidget(self.table)
            holder.setObjectName("RowSelectorCell")
            lay = QHBoxLayout(holder)
            lay.setContentsMargins(10, 0, 8, 0)
            lay.setSpacing(7)
            lay.addWidget(cb, 0, Qt.AlignCenter)

            seq_lab = QLabel(seq_text)
            seq_lab.setObjectName("RowSeqLabel")
            seq_lab.setAlignment(Qt.AlignCenter)
            seq_lab.setFont(self._songti_small_four_font())
            seq_lab.setMinimumWidth(22)
            lay.addWidget(seq_lab, 0, Qt.AlignCenter)
            lay.addStretch(1)

            self.table.setCellWidget(row_idx, 0, holder)

    def _on_item_changed(self, item: QTableWidgetItem):
        """处理单元格变动：如果是红色列的手动修改，恢复背景色为白色。"""
        if getattr(self, "_loading_data", False):
            return

        row = item.row()
        col = item.column()
        base_rows = self.DATA_START_ROW
        data_end = self._find_data_end_row()

        # 仅处理数据区的红色字段列：9..16
        if (base_rows <= row < data_end) and (9 <= col <= 16):
            # 如果当前背景色是淡蓝色（导入态），手动改动后应切回白色
            if item.background().color() == QColor("#e1f5fe"):
                self.table.blockSignals(True)
                item.setBackground(QColor("white"))
                self.table.blockSignals(False)

        if (base_rows <= row < data_end) and col == 6:
            formatted = self._format_weight_delta_text(item.text())
            if formatted != item.text():
                self.table.blockSignals(True)
                item.setText(formatted)
                self.table.blockSignals(False)

        if base_rows <= row < data_end:
            self._refresh_curve_view()

    def _text_pixel_width(self, text: str, fm: QFontMetrics) -> int:
        lines = str(text).splitlines() or [str(text)]
        return max(fm.horizontalAdvance(line) for line in lines)

    def _auto_fit_main_table_columns(self):
        """按文字内容自适应主表列宽（忽略跨列单元格，避免异常拉宽）。"""
        if not hasattr(self, "table") or self.table is None:
            return

        table = self.table
        col_count = table.columnCount()
        column_widths = {
            0: 104, 1: 170, 2: 116, 3: 280, 4: 132, 5: 140, 6: 120,
            7: 168, 8: 142, 9: 104, 10: 104, 11: 104, 12: 112, 13: 112,
            14: 112, 15: 118, 16: 118, 17: 112, 18: 168,
        }

        for c in range(col_count):
            table.setColumnWidth(c, column_widths.get(c, 110))

        table.setMinimumWidth(0)
        table.setMaximumWidth(16777215)
        table.setHorizontalScrollBarPolicy(Qt.ScrollBarAsNeeded)

    def _apply_data(self, rows: List[List[str]]):
        self._loading_data = True
        base_rows = self.DATA_START_ROW
        bg = QColor("#f8fafc")
        black = QColor("#111827")
        row_hint_bg = QColor("#f6faff")
        import_bg = QColor("#fff7ed")

        explain = [
            "填表说明：",
            "1. 改扩建时间：0行填原设计，后续填写改扩建完工的年月（如2023年10月）。",
            "2. 改扩建内容：平台本次改扩建增加或拆除的结构或设备。0行不填。",
            "3. 上部组块重量：0行为组块操作重量，后续为上一行重量与本行重量变化之和。",
            "4. 重量变化：0行为0，后续为重量增加填“+增加吨数”，如重量减少填“-减少吨数”。",
            "5. 极端工况最大载荷：填写最大值。",
        ]

        try:
            total = base_rows + len(rows) + len(explain)
            self.table.setRowCount(total)

            # 数据行高度
            for rr in range(base_rows, base_rows + len(rows)):
                self.table.setRowHeight(rr, 44)

            red_cols = set(range(9, 17))
            calc_bg = QColor("#eef8ef")

            for i, row in enumerate(rows):
                rr = base_rows + i
                for c, val in enumerate(row):
                    if c == 6:
                        val = self._format_weight_delta_text(val)
                    editable = (c != 0)
                    it = self._mk_item(val, editable=editable)
                    if c == 7:
                        it.setBackground(calc_bg)
                    elif c in red_cols:
                        it.setBackground(import_bg)
                        it.setToolTip("双击可手动输入数据；右键点击可读取本行对应的分析结果文件。")
                    elif i in (0, 1):
                        it.setBackground(row_hint_bg)
                    if val:
                        it.setForeground(black)
                    self.table.setItem(rr, c, it)

            start = base_rows + len(rows)
            for k, line in enumerate(explain):
                rr = start + k
                self.table.setRowHeight(rr, 24 if k else 28)
                self.table.setSpan(rr, 0, 1, self.table.columnCount())
                it = self._mk_item(line, bg=bg, align=Qt.AlignLeft | Qt.AlignVCenter, editable=False)
                self.table.setItem(rr, 0, it)

            data_end = start
            for r in range(0, self.table.rowCount()):
                for c in range(self.table.columnCount()):
                    if self.table.item(r, c) is None and self.table.cellWidget(r, c) is None:
                        editable = (base_rows <= r < data_end) and (c != 0)
                        item_bg = calc_bg if c == 7 else import_bg if (base_rows <= r < data_end) and (c in red_cols) else bg
                        it = self._mk_item("", bg=item_bg, editable=editable)
                        if (base_rows <= r < data_end) and (c in red_cols):
                            it.setToolTip("双击可手动输入数据；右键点击可读取本行对应的分析结果文件。")
                        self.table.setItem(r, c, it)

            data_start = base_rows
            data_end = self._find_data_end_row()
            for idx, row_idx in enumerate(range(data_start, data_end)):
                seq_item = self.table.item(row_idx, 0)
                if seq_item is not None:
                    seq_item.setText(str(idx))

            self._rebuild_row_checkbox_selectors(data_start, data_end)

            self.table.setMinimumHeight(0)
            self.table.setMaximumHeight(16777215)
            self.table.setVerticalScrollBarPolicy(Qt.ScrollBarAsNeeded)

            self._auto_fit_main_table_columns()
        finally:
            self._loading_data = False

        self._refresh_curve_view()
    # ---------------- 结果文件读取接口 ----------------

    def _on_table_context_menu(self, pos):
        """表格右键菜单逻辑：首列支持行操作，红色列支持读取结果文件。"""
        base_rows = self.DATA_START_ROW
        data_end = self._find_data_end_row()
        row = self.table.rowAt(pos.y())
        col = self.table.columnAt(pos.x())
        if not (base_rows <= row < data_end):
            return

        if col == 0:
            self._show_row_context_menu(row, pos)
            return

        if col == 7:
            menu = QMenu(self.table)
            menu.setStyleSheet(self._menu_qss())
            action = menu.addAction("打开上部组块分项目计算表")
            action.triggered.connect(lambda: self._open_upper_block_subproject_page(src_row=row))
            menu.exec_(self.table.viewport().mapToGlobal(pos))
            return

        # 红色字段列：9..16 (Fx~Mz, 操作工况, 极端工况)
        if 9 <= col <= 16:
            menu = QMenu(self.table)
            menu.setStyleSheet(self._menu_qss())
            action = menu.addAction("读取该行关联的结果文件 (psilst.factor)")
            action.triggered.connect(lambda: self._on_import_result(target_row=row))
            menu.exec_(self.table.viewport().mapToGlobal(pos))

    def _checked_data_rows(self) -> List[int]:
        data_end = self._find_data_end_row()
        rows: List[int] = []
        for row_idx, checkbox in getattr(self, "_row_checkboxes", {}).items():
            if row_idx < data_end and checkbox.isChecked():
                rows.append(row_idx)
        return sorted(rows)

    def _show_row_context_menu(self, row: int, pos):
        menu = QMenu(self.table)
        menu.setStyleSheet(self._menu_qss())

        add_above_action = QAction("在上方新增一行", menu)
        add_below_action = QAction("在下方新增一行", menu)
        add_above_action.triggered.connect(lambda _=False, target=row: self._insert_row_at(target))
        add_below_action.triggered.connect(lambda _=False, target=row + 1: self._insert_row_at(target))
        menu.addAction(add_above_action)
        menu.addAction(add_below_action)

        checked_rows = self._checked_data_rows()
        if checked_rows:
            delete_action = QAction(f"删除已勾选行（{len(checked_rows)}）", menu)
            delete_action.triggered.connect(lambda _=False, rows=checked_rows: self._delete_checked_rows(rows))
            menu.addAction(delete_action)
        else:
            hint_action = QAction("请先勾选要删除的行", menu)
            hint_action.setEnabled(False)
            menu.addAction(hint_action)

        menu.exec_(self.table.viewport().mapToGlobal(pos))

    def _on_import_result(self, target_row: int = None):
        """读取结果文件（psilst.factor）：由用户选择本地文件并回填该行红色字段。"""
        base_rows = self.DATA_START_ROW
        data_end = self._find_data_end_row()
        row = target_row if target_row is not None else self.table.currentRow()
        if row < base_rows or row >= data_end:
            QMessageBox.information(self, "提示", "请先在主表数据区选中或右键点击一行。")
            return

        start_dir = self.data_dir if os.path.exists(self.data_dir) else self.output_data_dir
        path, _ = QFileDialog.getOpenFileName(
            self,
            "选择结果文件（psilst.factor）",
            start_dir,
            "Factor Files (psilst.factor *.psilst.factor);;All Files (*)",
        )
        if not path:
            return

        file_name = os.path.basename(path).lower()
        if file_name != "psilst.factor" and not file_name.endswith(".psilst.factor"):
            QMessageBox.warning(self, "读取失败", "请选择名为 psilst.factor 的结果文件。")
            return

        try:
            QApplication.setOverrideCursor(Qt.WaitCursor)
            values = self._get_cached_result_factor_values(path)
            keys = ['Fx', 'Fy', 'Fz', 'Mx', 'My', 'Mz', '操作工况', '极端工况']
            if not any(str(values.get(k, '')).strip() for k in keys):
                QMessageBox.warning(self, "读取失败", "在 psilst.factor 文件中未解析到有效字段。")
                return
            self._apply_excel_values_to_row(row, values, bg_color=QColor("#e1f5fe"))
            QMessageBox.information(self, "读取结果文件", f"结果文件读取并回填成功。\n文件：{path}")
        except Exception as e:
            QMessageBox.warning(self, "读取失败", f"读取或解析失败：{e}")
        finally:
            QApplication.restoreOverrideCursor()

    def _on_red_field_mode_changed(self, idx: int):
        self.red_field_mode = 'manual' if idx == 0 else 'excel'
        self.btn_pick_excel_result.setEnabled(self.red_field_mode == 'excel')
        self.btn_import_excel_result.setEnabled(self.red_field_mode == 'excel')
        self._apply_red_fields_editability()

    def _apply_red_fields_editability(self):
        red_cols = list(range(9, 17))
        base_rows, data_end = self.DATA_START_ROW, self._find_data_end_row()
        for r in range(base_rows, data_end):
            for c in red_cols:
                it = self.table.item(r, c)
                if not it:
                    it = self._mk_item('', editable=True)
                    self.table.setItem(r, c, it)
                flags = it.flags()
                if self.red_field_mode == 'manual':
                    it.setFlags(flags | Qt.ItemIsEditable)
                    it.setBackground(QColor('white'))
                else:
                    it.setFlags(flags & ~Qt.ItemIsEditable)
                    it.setBackground(QColor('#f2f2f2'))

    def _pick_result_excel(self):
        start_dir = self.data_dir if os.path.exists(self.data_dir) else self.output_data_dir
        path, _ = QFileDialog.getOpenFileName(self, '选择结果文件（Excel）', start_dir, 'Excel Files (*.xlsx *.xlsm);;All Files (*)')
        if path: self.result_excel_path = path

    def _import_excel_to_current_row(self):
        if self.red_field_mode != 'excel': return
        if not self.result_excel_path: return
        row = self.table.currentRow()
        if row < self.DATA_START_ROW or row >= self._find_data_end_row(): return
        try:
            values = self._read_result_excel_generic(self.result_excel_path)
            if values:
                self._apply_excel_values_to_row(row, values, bg_color=QColor("#e1f5fe"))
                QMessageBox.information(self, '完成', '已导入Excel数据到当前行。')
        except Exception as e: QMessageBox.warning(self, '导入失败', str(e))

    def _apply_excel_values_to_row(self, row: int, values: Dict[str, object], bg_color: QColor = None):
        keys = ['Fx', 'Fy', 'Fz', 'Mx', 'My', 'Mz', '操作工况', '极端工况']
        cols = [9, 10, 11, 12, 13, 14, 15, 16]
        for k, c in zip(keys, cols):
            v = values.get(k, '')
            it = self.table.item(row, c)
            if not it:
                it = self._mk_item('', editable=True)
                self.table.setItem(row, c, it)
            it.setText('' if v is None else str(v))
            it.setForeground(QColor("#111827"))
            if bg_color: it.setBackground(bg_color)
        self._auto_fit_main_table_columns()
        self._refresh_curve_view()

    def _read_result_excel_generic(self, path: str) -> Dict[str, object]:
        wb = openpyxl.load_workbook(path, data_only=True)
        ws = wb.active
        keys = ['Fx', 'Fy', 'Fz', 'Mx', 'My', 'Mz', '操作工况', '极端工况']
        result = {}
        for r in range(1, ws.max_row + 1):
            for c in range(1, ws.max_column + 1):
                v = ws.cell(r, c).value
                if isinstance(v, str) and v.strip() in keys:
                    result[v.strip()] = ws.cell(r, c + 1).value or ws.cell(r + 1, c).value
        return result

    def _to_float(self, value: object) -> Optional[float]:
        try: return float(str(value).strip())
        except: return None

    def _fmt_float(self, value: object) -> str:
        fv = self._to_float(value)
        return f"{fv:.6f}".rstrip("0").rstrip(".") if fv is not None else (str(value) if value else "")

    def _fmt_result_load(self, value: object) -> str:
        fv = self._to_float(value)
        return f"{fv:.0f}" if fv is not None else (str(value) if value else "")

    def _fmt_result_safety(self, value: object) -> str:
        fv = self._to_float(value)
        return f"{fv:.2f}" if fv is not None else (str(value) if value else "")

    # ---------------- 表格行操作（新增/删除） ----------------
    def _insert_row_at(self, row: int):
        """在数据区指定位置新增一行。"""
        base_rows = self.DATA_START_ROW
        data_end = self._find_data_end_row()
        insert_row = max(base_rows, min(row, data_end))
        self.table.insertRow(insert_row)
        
        # 初始化新行的样式
        cols = self.table.columnCount()
        calc_bg = QColor("#eef8ef")
        red_cols = set(range(9, 17))
        
        for c in range(cols):
            editable = (c != 0)
            bg = calc_bg if c == 7 else QColor("white")
            it = self._mk_item("", bg=bg, editable=editable)
            if c in red_cols:
                it.setToolTip("双击可手动输入数据；右键点击可读取本行对应的分析结果文件。")
            self.table.setItem(insert_row, c, it)
            
        self._refresh_table_layout_and_seq()

    def _delete_checked_rows(self, rows: Optional[List[int]] = None):
        """删除已勾选的数据行。"""
        base_rows = self.DATA_START_ROW
        data_end = self._find_data_end_row()
        target_rows = sorted(rows if rows is not None else self._checked_data_rows())

        if not target_rows:
            QMessageBox.information(self, "提示", "请先勾选要删除的行。")
            return
        if target_rows[0] < base_rows or target_rows[-1] >= data_end:
            return
        if len(target_rows) >= data_end - base_rows:
            QMessageBox.information(self, "提示", "表格至少保留一条数据行。")
            return

        if not ask_yes_no(
            self,
            "确认删除",
            f"确定删除选中的 {len(target_rows)} 行信息吗？",
        ):
            return

        for row in reversed(target_rows):
            self.table.removeRow(row)

        self._refresh_table_layout_and_seq()

    def _refresh_table_layout_and_seq(self):
        """统一刷新序号、单选框、行高及表格总高度。"""
        base_rows = self.DATA_START_ROW
        data_end = self._find_data_end_row()
        
        # 1. 重新设置序号（连续编号）
        for idx, row_idx in enumerate(range(base_rows, data_end)):
            it = self.table.item(row_idx, 0)
            if it: it.setText(str(idx))
            self.table.setRowHeight(row_idx, 44)
            
        # 2. 重新构建复选框
        self._rebuild_row_checkbox_selectors(base_rows, data_end)
        
        self.table.setMinimumHeight(0)
        self.table.setMaximumHeight(16777215)
        self.table.setVerticalScrollBarPolicy(Qt.ScrollBarAsNeeded)
            
        self._auto_fit_main_table_columns()
        self._refresh_curve_view()

    def _extract_numbers(self, text: str) -> List[float]:
        return [float(s) for s in re.findall(self._num_pat, text or "") if self._to_float(s) is not None]

    def _read_text_file_with_fallback(self, path: str) -> str:
        for enc in ["utf-8", "gb18030", "gbk", "latin-1"]:
            try:
                with open(path, "r", encoding=enc) as f: return f.read()
            except: continue
        return ""

    def _extract_loads_from_text(self, text: str) -> Dict[str, str]:
        loads = {}
        for key in ["Fx", "Fy", "Fz", "Mx", "My", "Mz"]:
            m = re.search(rf"(?i)\b{key}\b.*?\b({self._num_pat})\b", text)
            if m: loads[key] = self._fmt_result_load(m.group(1))
        return loads

    def _extract_safety_from_text(self, text: str) -> Dict[str, str]:
        out = {}
        m_op = re.search(rf"(?im)操作工况.*?({self._num_pat})", text)
        if m_op: out["操作工况"] = self._fmt_result_safety(m_op.group(1))
        m_ex = re.search(rf"(?im)极端工况.*?({self._num_pat})", text)
        if m_ex: out["极端工况"] = self._fmt_result_safety(m_ex.group(1))
        return out

    def _parse_combined_load_cases_from_lines(self, lines: List[str]) -> List[Dict[str, object]]:
        rows: List[Dict[str, object]] = []
        reading = False
        skip = 0
        current: Optional[Dict[str, object]] = None

        for raw in lines:
            line = raw.replace("\x0c", "")
            if "***** SEASTATE COMBINED LOAD CASES *****" in raw:
                reading = True
                skip = 6
                current = None
                continue

            if not reading:
                continue

            if skip > 0:
                skip -= 1
                continue

            if "*****" in raw and "SEASTATE COMBINED LOAD CASES" not in raw:
                reading = False
                current = None
                continue

            if not line.strip():
                continue

            prefix = line[:12].strip()
            if prefix.isdigit():
                current = {
                    "load_case": line[14:18].strip(),
                    "fx": None,
                    "fy": None,
                    "fz": None,
                    "mx": None,
                    "my": None,
                    "mz": None,
                }
                rows.append(current)
                continue

            if current is None:
                continue

            left28 = line[:28]
            if "TOTAL" in left28 and line.strip()[-1:].isdigit():
                current["fx"] = self._to_float(line[28:42].strip())
                current["fy"] = self._to_float(line[42:55].strip())
                current["fz"] = self._to_float(line[55:68].strip())
                current["mx"] = self._to_float(line[68:81].strip())
                current["my"] = self._to_float(line[81:94].strip())
                current["mz"] = self._to_float(line[94:107].strip())

        return [row for row in rows if row.get("load_case")]

    def _parse_combined_load_summary_from_lines(self, lines: List[str]) -> List[Dict[str, object]]:
        rows: List[Dict[str, object]] = []
        reading = False
        skip = 0

        for raw in lines:
            line = raw.replace("\x0c", "")
            if "***** SEASTATE COMBINED LOAD CASE SUMMARY *****" in raw:
                reading = True
                skip = 6
                continue

            if not reading:
                continue

            if skip > 0:
                skip -= 1
                continue

            if "\x0c" in raw:
                reading = False
                continue

            if not line.strip():
                continue

            m = re.match(
                rf"^\s*(\d+)\s+([A-Z0-9]+)\s+({self._num_pat})\s+({self._num_pat})\s+({self._num_pat})\s+({self._num_pat})\s+({self._num_pat})\s+({self._num_pat})\s*$",
                line,
            )
            if not m:
                continue

            rows.append(
                {
                    "load_case": m.group(2).strip(),
                    "fx": self._to_float(m.group(3)),
                    "fy": self._to_float(m.group(4)),
                    "fz": self._to_float(m.group(5)),
                    "mx": self._to_float(m.group(6)),
                    "my": self._to_float(m.group(7)),
                    "mz": self._to_float(m.group(8)),
                }
            )

        return rows

    def _parse_combined_case_types_from_lines(self, lines: List[str]) -> Dict[str, str]:
        types: Dict[str, str] = {}
        for raw in lines:
            line = raw.replace("\x0c", "")
            m = re.match(
                r"^\s*\d+\s+([A-Z0-9]+)\s+\S+\s+\S+\s+\S+\s+([-+]?\d+(?:\.\d+)?)\s+([-+]?\d+(?:\.\d+)?)\s*$",
                line,
            )
            if not m:
                continue
            load_case = m.group(1).strip()
            amod_value = self._to_float(m.group(3))
            if not load_case or amod_value is None:
                continue
            types[load_case] = "Extreme" if abs(amod_value - 1.33) < 1e-6 else "Operation"

        return types

    def _pick_max_abs_loads_by_type(self, rows: List[Dict[str, object]], case_type: str) -> Dict[str, str]:
        fields = [("fx", "Fx"), ("fy", "Fy"), ("fz", "Fz"), ("mx", "Mx"), ("my", "My"), ("mz", "Mz")]
        out: Dict[str, str] = {}
        typed_rows = [row for row in rows if row.get("case_type") == case_type]
        for field_key, out_key in fields:
            best_row = None
            best_abs = -1.0
            for row in typed_rows:
                value = row.get(field_key)
                if value is None:
                    continue
                abs_value = abs(float(value))
                if abs_value > best_abs:
                    best_abs = abs_value
                    best_row = row
            if best_row is not None:
                out[out_key] = self._fmt_result_load(best_row.get(field_key))
        return out

    def _extract_vba_style_max_loads(self, text: str) -> Dict[str, str]:
        lines = text.splitlines()
        load_rows = self._parse_combined_load_summary_from_lines(lines)
        if not load_rows:
            load_rows = self._parse_combined_load_cases_from_lines(lines)
        if not load_rows:
            return {}

        case_types = self._parse_combined_case_types_from_lines(lines)
        for row in load_rows:
            row["case_type"] = case_types.get(str(row.get("load_case") or ""), "")

        extreme_loads = self._pick_max_abs_loads_by_type(load_rows, "Extreme")
        if extreme_loads:
            return extreme_loads
        return self._pick_max_abs_loads_by_type(load_rows, "Operation")

    def _parse_pile_safety_summary_from_lines(self, lines: List[str]) -> Dict[str, Dict[str, float]]:
        piles: Dict[str, Dict[str, float]] = {}
        reading = False
        skip = 0

        for raw in lines:
            line = raw.replace("\x0c", "")
            if "S O I L  M A X I M U M  A X I A L  C A P A C I T Y  S U M M A R Y" in raw:
                reading = True
                skip = 6
                continue

            if not reading:
                continue

            if skip > 0:
                skip -= 1
                continue

            if "*****" in raw and "S O I L  M A X I M U M" not in raw:
                reading = False
                continue

            if not line.strip():
                continue

            pile_id = line[:4].strip()
            if not pile_id:
                continue

            piles[pile_id] = {
                "weight": self._to_float(line[21:28].strip()) or 0.0,
                "comb_capacity": abs(self._to_float(line[34:44].strip()) or 0.0),
                "ten_capacity": abs(self._to_float(line[76:86].strip()) or 0.0),
            }

        return piles

    def _parse_pile_head_forces_from_lines(self, lines: List[str], case_types: Dict[str, str]) -> Dict[str, Dict[str, List[float]]]:
        forces: Dict[str, Dict[str, List[float]]] = {}
        i = 0
        total = len(lines)

        while i < total:
            raw = lines[i]
            if "INTERNAL FORCES ON STRUCTURE" not in raw:
                i += 1
                continue

            load_case = raw.strip()[-4:]
            j = i + 1
            found_header = False
            for _ in range(4):
                if j >= total:
                    break
                if "PILE HEAD COORDINATES" in lines[j]:
                    found_header = True
                    break
                j += 1

            if not found_header:
                i += 1
                continue

            j += 5
            case_type = case_types.get(load_case, "")
            while j < total:
                current_raw = lines[j]
                current = current_raw.replace("\x0c", "")
                if "\x0c" in current_raw or "INTERNAL FORCES ON STRUCTURE" in current_raw:
                    break
                if current.strip():
                    pile_tokens = current[:10].split()
                    pile_id = pile_tokens[0].strip() if pile_tokens else ""
                    force_value = self._to_float(current[:33][-14:].strip())
                    if pile_id and force_value is not None and case_type:
                        pile_bucket = forces.setdefault(pile_id, {"Operation": [], "Extreme": []})
                        pile_bucket.setdefault(case_type, []).append(-force_value)
                j += 1

            i = j

        return forces

    def _compute_vba_style_min_pile_safety(self, text: str) -> Dict[str, str]:
        lines = text.splitlines()
        case_types = self._parse_combined_case_types_from_lines(lines)
        pile_capacities = self._parse_pile_safety_summary_from_lines(lines)
        pile_forces = self._parse_pile_head_forces_from_lines(lines, case_types)

        if not pile_capacities or not pile_forces:
            return {}

        minima: Dict[str, Optional[float]] = {"Operation": None, "Extreme": None}
        for pile_id, capacities in pile_capacities.items():
            force_groups = pile_forces.get(pile_id)
            if not force_groups:
                continue

            weight = capacities.get("weight", 0.0)
            comb_capacity = capacities.get("comb_capacity", 0.0)
            ten_capacity = capacities.get("ten_capacity", 0.0)

            for case_type in ("Operation", "Extreme"):
                values = force_groups.get(case_type, [])
                if not values:
                    continue

                case_sfs: List[float] = []
                max_compression = min(values)
                if comb_capacity > 0:
                    denom = abs(max_compression) + weight
                    if denom > 0:
                        case_sfs.append(comb_capacity / denom)

                max_tension = max(values)
                if max_tension >= 0 and ten_capacity > 0:
                    denom = max_tension - weight
                    if denom > 0:
                        case_sfs.append(ten_capacity / denom)

                if case_sfs:
                    best = min(case_sfs)
                    current = minima[case_type]
                    minima[case_type] = best if current is None else min(current, best)

        out: Dict[str, str] = {}
        if minima["Operation"] is not None:
            out["操作工况"] = self._fmt_result_safety(minima["Operation"])
        if minima["Extreme"] is not None:
            out["极端工况"] = self._fmt_result_safety(minima["Extreme"])
        return out

    def _extract_vba_style_result_values(self, text: str) -> Dict[str, str]:
        result: Dict[str, str] = {}

        lines = text.splitlines()
        load_rows = self._parse_combined_load_cases_from_lines(lines)
        if load_rows:
            case_types = self._parse_combined_case_types_from_lines(lines)
            for row in load_rows:
                row["case_type"] = case_types.get(str(row.get("load_case") or ""), "")

            # 当前页面 9~14 列表头为“极端工况最大载荷”，优先取 Extreme。
            extreme_loads = self._pick_max_abs_loads_by_type(load_rows, "Extreme")
            if extreme_loads:
                result.update(extreme_loads)
            else:
                result.update(self._pick_max_abs_loads_by_type(load_rows, "Operation"))

        result.update(self._compute_vba_style_min_pile_safety(text))
        return result

    def _get_cached_result_factor_values(self, path: str) -> Dict[str, object]:
        normalized = os.path.normpath(path)
        mtime = os.path.getmtime(normalized)
        cache_key = (normalized, mtime)
        cached = self._result_factor_cache.get(cache_key)
        if cached is not None:
            return dict(cached)

        values = self._read_result_factor_generic(normalized)
        self._result_factor_cache = {cache_key: dict(values)}
        return values

    def _read_result_factor_generic(self, path: str) -> Dict[str, object]:
        text = self._read_text_file_with_fallback(path)
        res = self._extract_vba_style_result_values(text)
        if not any(str(res.get(k, '')).strip() for k in ['Fx', 'Fy', 'Fz', 'Mx', 'My', 'Mz', '操作工况', '极端工况']):
            res.update(self._extract_loads_from_text(text))
            res.update(self._extract_safety_from_text(text))
        return res

    def _collect_series_for_curve(self) -> Dict[str, List[float]]:
        base_rows, data_end = self.DATA_START_ROW, self._find_data_end_row()
        idxs, weight, cgx, cgy, fx, fy, fz, mx, my, mz = [], [], [], [], [], [], [], [], [], []
        for r in range(base_rows, data_end):
            seq = self._cell_text(r, 0).strip()
            if not seq.isdigit(): continue
            idxs.append(len(idxs))
            weight.append(self._to_float(self._cell_text(r, 4)) or 0.0)
            xyz = self._cell_text(r, 7).split(",")
            cgx.append(self._to_float(xyz[0]) if len(xyz)>0 else 0.0)
            cgy.append(self._to_float(xyz[1]) if len(xyz)>1 else 0.0)
            fx.append(self._to_float(self._cell_text(r, 9)) or 0.0)
            fy.append(self._to_float(self._cell_text(r, 10)) or 0.0)
            fz.append(self._to_float(self._cell_text(r, 11)) or 0.0)
            mx.append(self._to_float(self._cell_text(r, 12)) or 0.0)
            my.append(self._to_float(self._cell_text(r, 13)) or 0.0)
            mz.append(self._to_float(self._cell_text(r, 14)) or 0.0)
        return {"idx": idxs, "weight": weight, "cgx": cgx, "cgy": cgy, "fx": fx, "fy": fy, "fz": fz, "mx": mx, "my": my, "mz": mz}

    def _refresh_curve_view(self):
        if getattr(self, "_loading_data", False):
            return
        if not hasattr(self, "curve_widget"):
            return
        code = self._get_top_value("设施编码") or "XXXX"
        self.curve_widget.update_series(code, self._collect_series_for_curve())

    def _cell_text(self, row: int, col: int) -> str:
        if col == 0 and self.DATA_START_ROW <= row < self._find_data_end_row():
            return str(row - self.DATA_START_ROW)
        it = self.table.item(row, col)
        return it.text() if it else ""

    def _open_upper_block_subproject_page(self, src_row: int):
        mw = self.window()
        if not hasattr(mw, "tab_widget"): return
        key = f"uppercalc::{src_row}"
        if key in getattr(mw, "page_tab_map", {}):
            idx = mw.tab_widget.indexOf(mw.page_tab_map[key])
            if idx != -1: mw.tab_widget.setCurrentIndex(idx); return
        
        row_data = {c: self._cell_text(src_row, c) for c in range(self.table.columnCount())}
        page = UpperBlockSubprojectCalculationTablePage(main_window=mw, parent=mw)
        page.set_context(source_row=src_row, source_row_data=row_data)
        if src_row in self._uppercalc_saved_data: page.load_table_data(self._uppercalc_saved_data[src_row])
        page.saved.connect(self._on_upper_page_saved)
        title = f"{(row_data.get(1) or '序号'+str(src_row))}-上部组块分项目计算表"
        idx = mw.tab_widget.addTab(page, title)
        mw.tab_widget.setCurrentIndex(idx)
        if hasattr(mw, "page_tab_map"): mw.page_tab_map[key] = page

    def _on_upper_page_saved(self, payload: dict):
        src_row = payload.get("source_row")
        if src_row is None: return
        if "table_data" in payload: self._uppercalc_saved_data[src_row] = payload["table_data"]
        calc_bg = QColor("#eef8ef")
        wb = payload.get("write_back", {})
        for col, val in wb.items():
            it = self.table.item(src_row, int(col))
            if not it:
                it = self._mk_item("", editable=(int(col) != 0))
                self.table.setItem(src_row, int(col), it)
            it.setText(str(val))
            if int(col) == 7: it.setBackground(calc_bg)
        self._auto_fit_main_table_columns()
        self._refresh_curve_view()

    def _ensure_demo_files(self):
        # 打包环境下演示数据位于 _internal/data，无需在 exe 同级预创建空 data 目录。
        return

    def _on_save(self):
        facility_code = self._get_top_value("设施编码").strip()
        if not facility_code:
            QMessageBox.warning(self, "保存失败", "当前缺少设施编码，无法保存平台载荷信息。")
            return

        try:
            replace_platform_load_information_items(facility_code, self._collect_table_rows_for_db())
            self._notify_summary_pages_refresh()
            QMessageBox.information(self, "保存成功", "平台载荷信息已保存到数据库。")
        except Exception as exc:
            QMessageBox.critical(self, "保存失败", f"平台载荷信息保存失败：\n{exc}")

    def _notify_summary_pages_refresh(self):
        mw = self.window()
        tab_widget = getattr(mw, "tab_widget", None)
        if tab_widget is None:
            return
        for index in range(tab_widget.count()):
            page = tab_widget.widget(index)
            refresh = getattr(page, "refresh_from_database", None)
            if callable(refresh):
                refresh()

    def _on_export(self):
        if pd is None:
            QMessageBox.warning(self, "导出失败", "当前导出 Excel 需要安装 pandas 和 openpyxl。")
            return

        header = self._columns()
        rows = []
        for r in range(self.DATA_START_ROW, self._find_data_end_row()):
            row = [self._cell_text(r, c) for c in range(self.table.columnCount())]
            if any(str(value or "").strip() for value in row):
                rows.append(row)
        if not rows:
            QMessageBox.information(self, "导出数据", "当前无数据可导出。")
            return

        default_name = f"{self._get_top_value('设施编码').strip() or '平台'}_平台载荷信息.xlsx"
        file_path, _ = QFileDialog.getSaveFileName(
            self,
            "导出平台载荷信息",
            default_name,
            "Excel 文件 (*.xlsx)",
        )
        if not file_path:
            return
        if not file_path.lower().endswith(".xlsx"):
            file_path += ".xlsx"

        try:
            pd.DataFrame(rows, columns=header).to_excel(file_path, index=False)
        except Exception as exc:
            QMessageBox.critical(self, "导出失败", f"写入 Excel 失败：\n{exc}")
            return

        self._show_exported_file(file_path)

    def _show_exported_file(self, file_path: str):
        normalized = os.path.normpath(file_path)
        try:
            if sys.platform.startswith("win"):
                subprocess.Popen(["explorer.exe", f"/select,{normalized}"])
                self._raise_explorer_window()
            else:
                folder = os.path.dirname(normalized) or "."
                if sys.platform == "darwin":
                    subprocess.Popen(["open", folder])
                else:
                    subprocess.Popen(["xdg-open", folder])
        except Exception:
            pass

    def _raise_explorer_window(self):
        if not sys.platform.startswith("win"):
            return

        def _activate():
            try:
                user32 = ctypes.windll.user32
                handles = []

                @ctypes.WINFUNCTYPE(ctypes.c_bool, ctypes.c_void_p, ctypes.c_void_p)
                def enum_proc(hwnd, _lparam):
                    if not user32.IsWindowVisible(hwnd):
                        return True
                    class_name = ctypes.create_unicode_buffer(256)
                    user32.GetClassNameW(hwnd, class_name, 256)
                    if class_name.value in ("CabinetWClass", "ExploreWClass"):
                        handles.append(hwnd)
                    return True

                user32.EnumWindows(enum_proc, 0)
                if handles:
                    hwnd = handles[-1]
                    user32.ShowWindow(hwnd, 9)  # SW_RESTORE
                    user32.SetForegroundWindow(hwnd)
            except Exception:
                pass

        QTimer.singleShot(400, _activate)
