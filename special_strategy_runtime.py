from __future__ import annotations

import copy
import json
import sys
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from app_paths import external_path
from file_db_adapter import is_file_db_configured, list_files_by_prefix
from pages.file_management_platforms import find_platform
from special_strategy_state_db import (
    list_strategy_runs,
    load_latest_strategy_result_snapshot,
    load_latest_strategy_run,
    load_strategy_result_snapshot_by_run,
    load_strategy_run_by_id,
    save_strategy_result_snapshot,
    save_strategy_run,
    update_strategy_report,
)


OUTPUT_SPECIAL_STRATEGY_DIR = Path(__file__).resolve().parent / "pages" / "output_special_strategy"
if str(OUTPUT_SPECIAL_STRATEGY_DIR) not in sys.path:
    sys.path.insert(0, str(OUTPUT_SPECIAL_STRATEGY_DIR))

from inspection_tool import run as run_inspection_pipeline  # type: ignore  # noqa: E402
from report_jinja2_generator import (  # type: ignore  # noqa: E402
    build_appendix_pdf_plan,
    build_appendix_sections,
    insert_appendix_pdf_images,
    load_context_from_workbook,
    render_report,
)


FACILITY_CONFIG_MAP = {
    "WC19-1D": OUTPUT_SPECIAL_STRATEGY_DIR / "wc19_1d_run_config.json",
    "WC9-7": OUTPUT_SPECIAL_STRATEGY_DIR / "wc9_7_run_config.json",
}

TIME_CURRENT = "当前"


def normalize_facility_code(facility_code: str) -> str:
    code = (facility_code or "").strip().upper()
    if code in FACILITY_CONFIG_MAP:
        return code
    return "WC19-1D"


def runtime_dir(facility_code: str) -> Path:
    code = normalize_facility_code(facility_code)
    return Path(external_path("special_strategy_runtime", code)).resolve()


def state_path(facility_code: str) -> Path:
    return runtime_dir(facility_code) / "runtime_state.json"


def runtime_paths(facility_code: str) -> dict[str, Path]:
    root = runtime_dir(facility_code)
    root.mkdir(parents=True, exist_ok=True)
    return {
        "root": root,
        "params_json": root / "runtime_params.json",
        "intermediate_workbook": root / "special_strategy.pipeline.xlsx",
        "output_report": root / "special_strategy.docx",
        "state_json": root / "runtime_state.json",
    }


def load_base_config(facility_code: str) -> dict[str, Any]:
    path = FACILITY_CONFIG_MAP[normalize_facility_code(facility_code)]
    return json.loads(path.read_text(encoding="utf-8"))


def load_default_params(facility_code: str) -> dict[str, Any]:
    cfg = load_base_config(facility_code)
    params_path = Path(str(cfg.get("params_json", ""))).resolve()
    if not params_path.exists():
        return {}
    return json.loads(params_path.read_text(encoding="utf-8"))


def load_latest_strategy_params(facility_code: str) -> dict[str, Any]:
    params = copy.deepcopy(load_default_params(facility_code))
    latest = load_latest_strategy_run(normalize_facility_code(facility_code))
    if latest and isinstance(latest.get("params_json"), dict):
        params.update(latest["params_json"])
    return params


def default_metadata(facility_code: str) -> dict[str, str]:
    platform = find_platform(facility_code=normalize_facility_code(facility_code))
    return {
        "platform_name": platform["facility_name"],
        "report_date": datetime.now().strftime("%Y年%m月%d日"),
    }


def _write_json(path: Path, payload: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def load_runtime_state(facility_code: str) -> dict[str, Any] | None:
    db_state = load_latest_strategy_run(normalize_facility_code(facility_code))
    if db_state:
        return _state_from_run_payload(db_state)
    path = state_path(facility_code)
    if not path.exists():
        return None
    return json.loads(path.read_text(encoding="utf-8"))


def _state_from_run_payload(run_payload: dict[str, Any]) -> dict[str, Any]:
    return {
        "db_run_id": run_payload.get("id"),
        "facility_code": run_payload.get("facility_code"),
        "metadata": run_payload.get("metadata_json") or {},
        "params": run_payload.get("params_json") or {},
        "intermediate_workbook": str(run_payload.get("intermediate_workbook") or ""),
        "output_report": str(run_payload.get("output_report") or ""),
        "config_path": str(run_payload.get("config_path") or ""),
        "updated_at": run_payload.get("updated_at").isoformat(timespec="seconds") if run_payload.get("updated_at") else "",
        "inputs": run_payload.get("inputs_json") or {},
    }


def _normalize_logical_path(value: str) -> str:
    return str(value or "").replace("\\", "/").strip().strip("/")


def _row_sort_key(row: dict[str, Any]) -> tuple[str, str, str]:
    return (
        _normalize_logical_path(str(row.get("logical_path", ""))).lower(),
        str(row.get("original_name", "")).lower(),
        str(row.get("uploaded_at", "")),
    )


def _logical_has_segment(logical_path: str, segment: str) -> bool:
    path_parts = [part for part in _normalize_logical_path(logical_path).lower().split("/") if part]
    target_parts = [part for part in _normalize_logical_path(segment).lower().split("/") if part]
    if not target_parts or len(path_parts) < len(target_parts):
        return False
    for idx in range(len(path_parts) - len(target_parts) + 1):
        if path_parts[idx : idx + len(target_parts)] == target_parts:
            return True
    return False


def _existing_paths(paths: list[Any] | None) -> list[str]:
    out: list[str] = []
    for value in paths or []:
        text = str(value or "").strip()
        if not text:
            continue
        try:
            resolved = str(Path(text).resolve())
        except Exception:
            resolved = text
        if Path(resolved).exists():
            out.append(resolved)
    return out


def _apply_input_overrides(
    resolved_inputs: dict[str, Any],
    input_overrides: dict[str, Any] | None,
) -> dict[str, Any]:
    merged = {
        "model": str(resolved_inputs.get("model", "")),
        "clplog": [str(x) for x in resolved_inputs.get("clplog", [])],
        "ftglst": [str(x) for x in resolved_inputs.get("ftglst", [])],
        "ftginp": [str(x) for x in resolved_inputs.get("ftginp", [])],
    }
    if not input_overrides:
        return merged

    model_path = str(input_overrides.get("model") or "").strip()
    if model_path and Path(model_path).exists():
        merged["model"] = str(Path(model_path).resolve())

    for key in ("clplog", "ftglst", "ftginp"):
        override_paths = _existing_paths(input_overrides.get(key))
        if override_paths:
            merged[key] = override_paths
    return merged


def _should_override_fatigue_groups(
    default_ftglst: list[str],
    default_ftginp: list[str],
    candidate_ftglst_rows: list[dict[str, Any]],
    candidate_ftginp_rows: list[dict[str, Any]],
) -> bool:
    result_count = len(candidate_ftglst_rows)
    input_count = len(candidate_ftginp_rows)
    if result_count <= 0 or input_count <= 0:
        return False

    expected_result_count = len(default_ftglst)
    expected_input_count = len(default_ftginp)
    if expected_result_count > 0 and result_count < expected_result_count:
        return False
    if expected_input_count > 0 and input_count < expected_input_count:
        return False
    return True


def resolve_current_model_inputs(facility_code: str, cfg: dict[str, Any]) -> dict[str, Any]:
    resolved = {
        "model": str(cfg["model"]),
        "clplog": [str(x) for x in cfg.get("clplog", [])],
        "ftglst": [str(x) for x in cfg.get("ftglst", [])],
        "ftginp": [str(x) for x in cfg.get("ftginp", [])],
    }
    if not is_file_db_configured():
        return resolved

    rows = list_files_by_prefix(
        module_code="model_files",
        logical_path_prefix=f"{normalize_facility_code(facility_code)}/当前模型",
        facility_code=normalize_facility_code(facility_code),
    )
    if not rows:
        return resolved

    model_path = ""
    collapse_rows: list[dict[str, Any]] = []
    ftglst_rows: list[dict[str, Any]] = []
    ftginp_rows: list[dict[str, Any]] = []

    for row in sorted(rows, key=_row_sort_key):
        logical_path = _normalize_logical_path(str(row.get("logical_path", "")))
        storage_path = str(row.get("storage_path", "")).strip()
        if not storage_path or not Path(storage_path).exists():
            continue
        if _logical_has_segment(logical_path, "当前模型/结构模型") and not model_path:
            model_path = storage_path
        elif _logical_has_segment(logical_path, "当前模型/倒塌分析"):
            collapse_rows.append(row)
        elif (
            _logical_has_segment(logical_path, "当前模型/疲劳分析")
            and _logical_has_segment(logical_path, "结果")
        ):
            ftglst_rows.append(row)
        elif (
            _logical_has_segment(logical_path, "当前模型/疲劳分析")
            and _logical_has_segment(logical_path, "输入")
        ):
            ftginp_rows.append(row)

    if model_path:
        resolved["model"] = model_path
    if collapse_rows:
        resolved["clplog"] = [str(row["storage_path"]) for row in sorted(collapse_rows, key=_row_sort_key)]
    if _should_override_fatigue_groups(resolved["ftglst"], resolved["ftginp"], ftglst_rows, ftginp_rows):
        resolved["ftglst"] = [str(row["storage_path"]) for row in sorted(ftglst_rows, key=_row_sort_key)]
        resolved["ftginp"] = [str(row["storage_path"]) for row in sorted(ftginp_rows, key=_row_sort_key)]
    return resolved


def merge_runtime_params(facility_code: str, overrides: dict[str, Any] | None = None) -> dict[str, Any]:
    params = copy.deepcopy(load_default_params(facility_code))
    overrides = overrides or {}
    for key in (
        "life_safety_level",
        "failure_consequence_level",
        "x_angle_deviation",
        "min_leg_od",
        "wp_z",
        "no_legs",
        "global_level_tag",
        "region",
        "collapse_a_const",
        "collapse_b_const",
        "served_years",
        "design_life",
    ):
        if key in overrides and overrides[key] not in ("", None):
            params[key] = overrides[key]
    work_points = overrides.get("work_points")
    if work_points:
        params["work_points"] = work_points
    return params


def _context_from_workbook(workbook_path: Path, cfg: dict[str, Any], metadata: dict[str, Any]) -> dict[str, Any]:
    context = load_context_from_workbook(
        workbook_path,
        metadata,
        row_limits=None,
        strict_vba_algorithms=not bool(cfg.get("non_vba_post_filters", False)),
        apply_vba_member_delete_rules=True,
        apply_vba_joint_delete_rules_current=True,
        apply_vba_joint_delete_rules_future=True,
    )
    return context


def _to_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _iter_rows(ws, min_row: int, max_col: int):
    if ws is None:
        return
    for row in ws.iter_rows(min_row=min_row, max_col=max_col, values_only=True):
        values = list(row[:max_col])
        if not any(_to_text(v) for v in values):
            continue
        yield values


def _load_detail_rows(workbook_path: Path) -> dict[str, list[dict[str, str]]]:
    wb = load_workbook(workbook_path, data_only=True, read_only=True)
    ws_node_risk = wb["节点失效风险等级"] if "节点失效风险等级" in wb.sheetnames else None
    ws_node_strategy = wb["节点检验策略"] if "节点检验策略" in wb.sheetnames else None
    ws_member_risk = wb["构件失效风险等级"] if "构件失效风险等级" in wb.sheetnames else None

    node_level_map: dict[tuple[str, str], str] = {}
    if ws_node_strategy is not None:
        for row in _iter_rows(ws_node_strategy, min_row=3, max_col=18):
            time_node = _to_text(row[17])
            if time_node not in ("", TIME_CURRENT):
                continue
            key = (_to_text(row[0]), _to_text(row[1]))
            node_level_map[key] = _to_text(row[15])

    node_rows: list[dict[str, str]] = []
    if ws_node_risk is not None:
        for row in _iter_rows(ws_node_risk, min_row=3, max_col=10):
            key = (_to_text(row[0]), _to_text(row[1]))
            node_rows.append(
                {
                    "joint_a": _to_text(row[0]),
                    "joint_b": _to_text(row[1]),
                    "weld_type": _to_text(row[2]),
                    "consequence_level": _to_text(row[3]),
                    "a": _to_text(row[4]),
                    "b": _to_text(row[5]),
                    "rm": _to_text(row[6]),
                    "vr": _to_text(row[7]),
                    "pf": _to_text(row[8]),
                    "collapse_prob_level": _to_text(row[9]),
                    "risk_level": node_level_map.get(key, ""),
                }
            )

    member_rows: list[dict[str, str]] = []
    if ws_member_risk is not None:
        for row in _iter_rows(ws_member_risk, min_row=3, max_col=11):
            member_rows.append(
                {
                    "joint_a": _to_text(row[0]),
                    "joint_b": _to_text(row[1]),
                    "member_type": _to_text(row[2]),
                    "consequence_level": _to_text(row[3]),
                    "a": _to_text(row[4]),
                    "b": _to_text(row[5]),
                    "rm": _to_text(row[6]),
                    "vr": _to_text(row[7]),
                    "pf": _to_text(row[8]),
                    "collapse_prob_level": _to_text(row[9]),
                    "risk_level": _to_text(row[10]),
                }
            )
    wb.close()
    return {"member": member_rows, "node": node_rows}


def _build_result_bundle_from_workbook(
    *,
    facility_code: str,
    workbook_path: Path,
    cfg: dict[str, Any],
    metadata: dict[str, Any],
    state: dict[str, Any],
) -> dict[str, Any]:
    context = _context_from_workbook(workbook_path, cfg, metadata)
    detail_rows = _load_detail_rows(workbook_path)
    return {
        "state": state,
        "context": context,
        "member_risk_rows_full": detail_rows["member"],
        "node_risk_rows_full": detail_rows["node"],
    }


def run_special_strategy_calculation(
    facility_code: str,
    *,
    param_overrides: dict[str, Any] | None = None,
    input_overrides: dict[str, Any] | None = None,
    metadata: dict[str, Any] | None = None,
) -> dict[str, Any]:
    code = normalize_facility_code(facility_code)
    cfg = load_base_config(code)
    paths = runtime_paths(code)
    params = merge_runtime_params(code, param_overrides)
    _write_json(paths["params_json"], params)

    metadata_payload = default_metadata(code)
    if metadata:
        metadata_payload.update({k: v for k, v in metadata.items() if v not in ("", None)})

    resolved_inputs = _apply_input_overrides(
        resolve_current_model_inputs(code, cfg),
        input_overrides,
    )
    config_xlsm = Path(str(cfg.get("config_xlsm") or cfg["template_xlsm"])).resolve()

    run_inspection_pipeline(
        template_xlsm=config_xlsm,
        model_file=Path(str(resolved_inputs["model"])).resolve(),
        clplog_file=[Path(str(x)).resolve() for x in resolved_inputs["clplog"]],
        ftglst_file=[Path(str(x)).resolve() for x in resolved_inputs["ftglst"]],
        out_xlsx=paths["intermediate_workbook"],
        policy_mode=str(cfg.get("policy", "strict")),
        seed=int(cfg.get("seed", 42)),
        params_json=paths["params_json"],
        ftginp_files=[Path(str(x)).resolve() for x in resolved_inputs["ftginp"]],
        manual_fill_workbook=None,
        interactive_manual_fill=True,
        enable_topology_inference=bool(cfg.get("enable_topology_inference", False)),
    )

    state = {
        "facility_code": code,
        "metadata": metadata_payload,
        "params_json": str(paths["params_json"]),
        "intermediate_workbook": str(paths["intermediate_workbook"]),
        "output_report": str(paths["output_report"]),
        "config_path": str(FACILITY_CONFIG_MAP[code]),
        "inputs": resolved_inputs,
        "updated_at": datetime.now().isoformat(timespec="seconds"),
    }
    run_id = save_strategy_run(
        facility_code=code,
        params=params,
        metadata=metadata_payload,
        inputs=resolved_inputs,
        intermediate_workbook=str(paths["intermediate_workbook"]),
        output_report=str(paths["output_report"]),
        config_path=str(FACILITY_CONFIG_MAP[code]),
        status="completed",
    )
    state["db_run_id"] = run_id
    result_bundle = _build_result_bundle_from_workbook(
        facility_code=code,
        workbook_path=paths["intermediate_workbook"],
        cfg=cfg,
        metadata=metadata_payload,
        state=state,
    )
    snapshot_id = save_strategy_result_snapshot(
        facility_code=code,
        run_id=run_id,
        result_payload={
            "context": result_bundle["context"],
            "member_risk_rows_full": result_bundle["member_risk_rows_full"],
            "node_risk_rows_full": result_bundle["node_risk_rows_full"],
            "state": {
                "facility_code": state["facility_code"],
                "metadata": state["metadata"],
                "params_json": str(paths["params_json"]),
                "intermediate_workbook": str(paths["intermediate_workbook"]),
                "output_report": str(paths["output_report"]),
                "config_path": state["config_path"],
                "inputs": state["inputs"],
                "updated_at": state["updated_at"],
                "db_run_id": run_id,
            },
        },
    )
    state["db_snapshot_id"] = snapshot_id
    _write_json(paths["state_json"], state)
    result_bundle["state"] = state
    return result_bundle


def load_result_bundle(facility_code: str, run_id: int | None = None) -> dict[str, Any] | None:
    code = normalize_facility_code(facility_code)
    run_payload = load_strategy_run_by_id(run_id) if run_id else None
    if run_id and run_payload is None:
        return None
    state = _state_from_run_payload(run_payload) if run_payload else load_runtime_state(code)
    if not state:
        return None
    if run_id and run_payload and normalize_facility_code(str(run_payload.get("facility_code") or "")) != code:
        return None
    snapshot = load_strategy_result_snapshot_by_run(run_id) if run_id else load_latest_strategy_result_snapshot(code)
    if snapshot and isinstance(snapshot.get("result_json"), dict):
        payload = dict(snapshot["result_json"])
        payload["state"] = payload.get("state") or state
        return payload
    workbook_path = Path(str(state.get("intermediate_workbook", ""))).resolve()
    if not workbook_path.exists():
        return None
    cfg = load_base_config(code)
    metadata = dict(default_metadata(code))
    metadata.update(state.get("metadata") or {})
    return _build_result_bundle_from_workbook(
        facility_code=code,
        workbook_path=workbook_path,
        cfg=cfg,
        metadata=metadata,
        state=state,
    )


def load_latest_result_bundle(facility_code: str) -> dict[str, Any] | None:
    return load_result_bundle(facility_code)


def list_result_run_history(facility_code: str, limit: int = 50) -> list[dict[str, Any]]:
    code = normalize_facility_code(facility_code)
    return list_strategy_runs(code, limit=limit)


def generate_special_strategy_report(
    facility_code: str,
    *,
    run_id: int | None = None,
    metadata: dict[str, Any] | None = None,
) -> Path:
    code = normalize_facility_code(facility_code)
    run_payload = load_strategy_run_by_id(run_id) if run_id else None
    if run_id and run_payload is None:
        raise FileNotFoundError(f"No calculated result found for {code} run_id={run_id}.")
    if run_id and run_payload and normalize_facility_code(str(run_payload.get("facility_code") or "")) != code:
        raise FileNotFoundError(f"No calculated result found for {code} run_id={run_id}.")
    state = _state_from_run_payload(run_payload) if run_payload else load_runtime_state(code)
    if not state:
        raise FileNotFoundError(f"No calculated result found for {code}.")
    workbook_path = Path(str(state["intermediate_workbook"])).resolve()
    if not workbook_path.exists():
        raise FileNotFoundError(f"Intermediate workbook not found: {workbook_path}")

    cfg = load_base_config(code)
    metadata_payload = dict(default_metadata(code))
    metadata_payload.update(state.get("metadata") or {})
    if metadata:
        metadata_payload.update({k: v for k, v in metadata.items() if v not in ("", None)})

    context = _context_from_workbook(workbook_path, cfg, metadata_payload)
    context["appendix_sections"] = build_appendix_sections(
        appendix_a_file=str(cfg.get("appendix_a_file", "")).strip(),
        appendix_b_file=str(cfg.get("appendix_b_file", "")).strip(),
        appendix_c_dirs=[str(x) for x in cfg.get("appendix_c_dirs", [])],
    )
    context["appendix_pdf_plan"] = build_appendix_pdf_plan(
        appendix_a_file=str(cfg.get("appendix_a_file", "")).strip(),
        appendix_b_file=str(cfg.get("appendix_b_file", "")).strip(),
        appendix_c_dirs=[str(x) for x in cfg.get("appendix_c_dirs", [])],
    )
    context["include_word_plan_detail_tables"] = bool(cfg.get("include_word_plan_detail_tables", False))

    paths = runtime_paths(code)
    report_output_path = (
        paths["root"] / f"special_strategy_run_{int(run_id)}.docx"
        if run_id
        else paths["output_report"]
    )
    report_template = Path(str(cfg["report_template"])).resolve()
    render_report(report_template, report_output_path, context)
    insert_appendix_pdf_images(report_output_path, context)

    state["metadata"] = metadata_payload
    state["output_report"] = str(report_output_path)
    state["report_generated_at"] = datetime.now().isoformat(timespec="seconds")
    state_run_id = state.get("db_run_id")
    if state_run_id:
        update_strategy_report(int(state_run_id), output_report=str(report_output_path))
    if not run_id:
        _write_json(paths["state_json"], state)
    return report_output_path
